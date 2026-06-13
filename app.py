"""
Revisi — мульти-тенантная система инвентаризации.
Фаза 1: Аутентификация и регистрация компаний.
"""
from flask import (
    Flask, request, render_template_string, redirect, url_for, flash, jsonify,
    send_file, abort, session,
)
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    current_user, login_required,
)
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
import secrets
import os
import re
import csv
import io
import json
import click

from config import Config
from models import (
    db, Organization, User, Location, OwnerUser, LoginAttempt,
    Category, Product, ProductNorm, LocationProduct, Revision, RevisionItem,
    OwnerAuditLog, SupportTicket, SupportTicketReply, CatalogProduct,
)

# ============== ИНИЦИАЛИЗАЦИЯ ==============
app = Flask(__name__)
app.config.from_object(Config)

# gzip-сжатие HTML/JSON ответов — ~70% экономии трафика на наших страницах
try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    pass  # Flask-Compress опционален

db.init_app(app)
migrate = Migrate(app, db)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Пожалуйста, войдите для доступа.'
# Отключаем basic protection — иначе смена IP/UA (мобильный интернет, апгрейд браузера)
# выбрасывает пользователя. Для инвентаризации достаточно обычной сессии.
login_manager.session_protection = None

# Render/любой reverse-proxy: доверяем X-Forwarded-* чтобы Flask знал настоящий scheme
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Время запуска приложения (для health-check uptime)
_APP_STARTED_AT = datetime.utcnow()


# ============== HOSTNAME-РОУТИНГ ==============
# Лендинг живёт на корневом домене, само приложение — на поддомене app.
LANDING_HOSTS = {'revisi.ru', 'www.revisi.ru'}
APP_HOST = 'app.revisi.ru'


@app.after_request
def add_owner_security_headers(response):
    """Owner-панель не должна попадать в поиск, кеши и архивы."""
    if request.path.startswith('/owner'):
        response.headers['X-Robots-Tag'] = 'noindex, nofollow, nosnippet, noarchive'
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['Referrer-Policy'] = 'no-referrer'
    return response


@app.before_request
def route_by_host():
    """revisi.ru/ → лендинг. revisi.ru/<path> → 301 на app.revisi.ru/<path>. app.revisi.ru/* → приложение."""
    # /static/* и /healthz работают одинаково на обоих хостах без редиректов
    if request.path.startswith('/static/') or request.path == '/healthz':
        return None

    host = (request.host or '').split(':')[0].lower()

    if host in LANDING_HOSTS:
        if request.path == '/':
            from flask import send_from_directory
            return send_from_directory('static/landing', 'index.html')
        new_url = f'https://{APP_HOST}{request.full_path.rstrip("?")}'
        return redirect(new_url, code=301)

    return None


# Health-check endpoint для Render — проверяет что приложение живо И БД отвечает.
# Render автоматически рестартит сервис, если /healthz возвращает 5xx.
@app.route('/healthz')
def healthz():
    from sqlalchemy import text as _sql_text
    try:
        db.session.execute(_sql_text('SELECT 1'))
        return jsonify(status='ok'), 200
    except Exception as e:
        return jsonify(status='error', detail=str(e)[:120]), 503


_PWA_HEAD = '''
<link rel="icon" type="image/svg+xml" href="/static/icon.svg?v=4">
<link rel="apple-touch-icon" href="/static/icon-180.png?v=4">
<link rel="manifest" href="/static/manifest.json?v=4">
<meta name="theme-color" content="#0f1725">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Revisi">
'''

_OWNER_PWA_HEAD = '''
<link rel="icon" type="image/svg+xml" href="/static/icon-owner.svg?v=4">
<link rel="apple-touch-icon" href="/static/icon-owner-180.png?v=4">
<link rel="manifest" href="/static/manifest-owner.json?v=4">
<meta name="theme-color" content="#0f1725">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Revisi Owner">
'''


# ── SVG-иконки (Lucide-style, currentColor) ──────────────────────────
# Заменяют emoji в UI чтобы интерфейс выглядел профессионально.
_ICONS = {
    'menu':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><line x1="3" y1="6" x2="21" y2="6"/><line x1="3" y1="12" x2="21" y2="12"/><line x1="3" y1="18" x2="21" y2="18"/></svg>',
    'mappin':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M21 10c0 7-9 13-9 13s-9-6-9-13a9 9 0 0 1 18 0z"/><circle cx="12" cy="10" r="3"/></svg>',
    'folder':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M22 19a2 2 0 0 1-2 2H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h5l2 3h9a2 2 0 0 1 2 2z"/></svg>',
    'box':        '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M21 16V8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16z"/><polyline points="3.27 6.96 12 12.01 20.73 6.96"/><line x1="12" y1="22.08" x2="12" y2="12"/></svg>',
    'users':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M16 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="8.5" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg>',
    'ruler':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M21 3 3 21l-2-2 18-18z"/><path d="M7 17l2-2M10 14l2-2M13 11l2-2M16 8l2-2"/></svg>',
    'download':   '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>',
    'inbox':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><polyline points="22 12 16 12 14 15 10 15 8 12 2 12"/><path d="M5.45 5.11 2 12v6a2 2 0 0 0 2 2h16a2 2 0 0 0 2-2v-6l-3.45-6.89A2 2 0 0 0 16.76 4H7.24a2 2 0 0 0-1.79 1.11z"/></svg>',
    'archive':    '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><polyline points="21 8 21 21 3 21 3 8"/><rect x="1" y="3" width="22" height="5"/><line x1="10" y1="12" x2="14" y2="12"/></svg>',
    'pencil':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M12 20h9"/><path d="M16.5 3.5a2.12 2.12 0 0 1 3 3L7 19l-4 1 1-4z"/></svg>',
    'trash':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"/><path d="M19 6 17.7 20.1A2 2 0 0 1 15.7 22H8.3a2 2 0 0 1-2-1.9L5 6"/><path d="M10 11v6M14 11v6"/><path d="M9 6V4a2 2 0 0 1 2-2h2a2 2 0 0 1 2 2v2"/></svg>',
    'chat':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"/></svg>',
    'logout':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/></svg>',
    'chart':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M3 3v18h18"/><rect x="7" y="13" width="3" height="5"/><rect x="12" y="9" width="3" height="9"/><rect x="17" y="6" width="3" height="12"/></svg>',
    'shop':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9 4.5 4h15L21 9"/><path d="M4 9v11a1 1 0 0 0 1 1h14a1 1 0 0 0 1-1V9"/><path d="M9 21v-6h6v6"/></svg>',
    'building':   '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><rect x="4" y="3" width="16" height="18" rx="2"/><line x1="9" y1="9" x2="9" y2="9.01"/><line x1="15" y1="9" x2="15" y2="9.01"/><line x1="9" y1="13" x2="9" y2="13.01"/><line x1="15" y1="13" x2="15" y2="13.01"/><line x1="9" y1="17" x2="9" y2="17.01"/><line x1="15" y1="17" x2="15" y2="17.01"/></svg>',
    'rotate':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M3 12a9 9 0 0 1 15.5-6.3L21 8"/><path d="M21 3v5h-5"/></svg>',
    'reset':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M3 12a9 9 0 1 0 3-6.7L3 8"/><path d="M3 3v5h5"/></svg>',
    'plus':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="5" x2="12" y2="19"/><line x1="5" y1="12" x2="19" y2="12"/></svg>',
    'close':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/></svg>',
    'check':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>',
    'send':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><line x1="22" y1="2" x2="11" y2="13"/><polygon points="22 2 15 22 11 13 2 9 22 2"/></svg>',
    'search':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="7"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>',
    'lock':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="11" width="18" height="11" rx="2"/><path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg>',
    'ticket':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9V6a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2v3a2 2 0 0 0 0 4v3a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-3a2 2 0 0 0 0-4z"/><line x1="13" y1="5" x2="13" y2="7"/><line x1="13" y1="11" x2="13" y2="13"/><line x1="13" y1="17" x2="13" y2="19"/></svg>',
    'sliders':    '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><line x1="4" y1="21" x2="4" y2="14"/><line x1="4" y1="10" x2="4" y2="3"/><line x1="12" y1="21" x2="12" y2="12"/><line x1="12" y1="8" x2="12" y2="3"/><line x1="20" y1="21" x2="20" y2="16"/><line x1="20" y1="12" x2="20" y2="3"/><line x1="1" y1="14" x2="7" y2="14"/><line x1="9" y1="8" x2="15" y2="8"/><line x1="17" y1="16" x2="23" y2="16"/></svg>',
    'file':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>',
    'megaphone':  '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="m3 11 18-5v12L3 14v-3z"/><path d="M11.6 16.8a3 3 0 1 1-5.8-1.6"/></svg>',
    'database':   '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><ellipse cx="12" cy="5" rx="9" ry="3"/><path d="M3 5v14a9 3 0 0 0 18 0V5"/><path d="M3 12a9 3 0 0 0 18 0"/></svg>',
    'heartpulse': '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M20.84 4.61a5.5 5.5 0 0 0-7.78 0L12 5.67l-1.06-1.06a5.5 5.5 0 0 0-7.78 7.78L12 21.23l8.84-8.84a5.5 5.5 0 0 0 0-7.78z"/><path d="m3.66 12 4.34 0 1.5-3 3 6 1.5-3 4.34 0"/></svg>',
    'terminal':   '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><polyline points="4 17 10 11 4 5"/><line x1="12" y1="19" x2="20" y2="19"/></svg>',
    'alert':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M10.29 3.86 1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/><line x1="12" y1="9" x2="12" y2="13"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg>',
    'arrow-left': '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><line x1="19" y1="12" x2="5" y2="12"/><polyline points="12 19 5 12 12 5"/></svg>',
    'mail':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><rect x="2" y="4" width="20" height="16" rx="2"/><path d="m22 7-10 6L2 7"/></svg>',
    'key':        '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><circle cx="7.5" cy="15.5" r="4.5"/><path d="m10.7 12.3 9.3-9.3M18 5l2 2M15 8l2 2"/></svg>',
    'user':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg>',
    'save':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"/><polyline points="17 21 17 13 7 13 7 21"/><polyline points="7 3 7 8 15 8"/></svg>',
    'target':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><circle cx="12" cy="12" r="6"/><circle cx="12" cy="12" r="2"/></svg>',
    'help':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><path d="M9.1 9a3 3 0 0 1 5.8 1c0 2-3 3-3 3"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg>',
    'upload':     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>',
    'bulb':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M9 18h6M10 22h4"/><path d="M12 2a7 7 0 0 0-4 12.7c.6.5 1 1.3 1 2.3h6c0-1 .4-1.8 1-2.3A7 7 0 0 0 12 2z"/></svg>',
    'utensils':   '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M3 2v7a3 3 0 0 0 6 0V2M6 9v13M16 2c-1.5 0-3 1.5-3 4v6h4V2z"/></svg>',
    'check-circle':'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg>',
    'clipboard':  '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><rect x="8" y="2" width="8" height="4" rx="1"/><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"/></svg>',
    'mask':       '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M2 9a3 3 0 0 1 3-3h14a3 3 0 0 1 3 3v3a7 7 0 0 1-7 7 4 4 0 0 1-3-1.5A4 4 0 0 1 9 19a7 7 0 0 1-7-7z"/><circle cx="8" cy="11" r="1"/><circle cx="16" cy="11" r="1"/></svg>',
    'clock':      '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"/><polyline points="12 7 12 12 15.5 14"/></svg>',
}


def _icon(name, size=18):
    svg = _ICONS.get(name, '')
    if not svg:
        return ''
    return svg.replace('<svg ', f'<svg width="{size}" height="{size}" style="vertical-align:-3px;flex-shrink:0;" ', 1)


# ── Время: храним в UTC, показываем в часовом поясе компании ──────────
_TZ_CHOICES = [
    ('Europe/Kaliningrad', 'Калининград (UTC+2)'),
    ('Europe/Moscow', 'Москва (UTC+3)'),
    ('Europe/Samara', 'Самара (UTC+4)'),
    ('Asia/Yekaterinburg', 'Екатеринбург (UTC+5)'),
    ('Asia/Omsk', 'Омск (UTC+6)'),
    ('Asia/Krasnoyarsk', 'Красноярск (UTC+7)'),
    ('Asia/Irkutsk', 'Иркутск (UTC+8)'),
    ('Asia/Yakutsk', 'Якутск (UTC+9)'),
    ('Asia/Vladivostok', 'Владивосток (UTC+10)'),
    ('Asia/Magadan', 'Магадан (UTC+11)'),
    ('Asia/Kamchatka', 'Камчатка (UTC+12)'),
    ('Asia/Almaty', 'Алматы (UTC+5)'),
    ('Asia/Tashkent', 'Ташкент (UTC+5)'),
    ('Asia/Bishkek', 'Бишкек (UTC+6)'),
    ('Asia/Tbilisi', 'Тбилиси (UTC+4)'),
    ('Asia/Yerevan', 'Ереван (UTC+4)'),
    ('Asia/Baku', 'Баку (UTC+4)'),
]
_TZ_VALID = {c[0] for c in _TZ_CHOICES}


def _fmt_dt(dt, org=None, fmt='%d.%m.%Y %H:%M'):
    """Отформатировать UTC-время в часовом поясе компании (по умолчанию МСК)."""
    if not dt:
        return ''
    from datetime import timezone as _utc
    try:
        from zoneinfo import ZoneInfo
        tzname = (getattr(org, 'timezone', None) or 'Europe/Moscow') if org else 'Europe/Moscow'
        if tzname not in _TZ_VALID:
            tzname = 'Europe/Moscow'
        return dt.replace(tzinfo=_utc.utc).astimezone(ZoneInfo(tzname)).strftime(fmt)
    except Exception:
        # Фолбэк: жёсткое смещение МСК (+3), если zoneinfo/tzdata недоступны
        from datetime import timedelta as _td
        return (dt + _td(hours=3)).strftime(fmt)


@app.context_processor
def _inject_pwa():
    return {
        'pwa_head': _PWA_HEAD,
        'owner_pwa_head': _OWNER_PWA_HEAD,
        'is_impersonating': bool(session.get('impersonating_owner_id')),
        'icon': _icon,
    }


@app.template_filter('qty')
def _qty_fmt(v):
    """Округлить до 3 знаков и отформатировать с запятой (RU)."""
    if v is None:
        return ''
    try:
        n = round(float(v), 3)
    except (TypeError, ValueError):
        return str(v)
    if n == int(n):
        return str(int(n))
    return f'{n:g}'.replace('.', ',')


@app.before_request
def make_session_permanent():
    session.permanent = True


def log_owner_action(action, target_org_id=None, target_user_id=None, details=None):
    """Записать действие владельца в audit-log. Безопасно: исключения подавляются."""
    try:
        owner_id = None
        try:
            if hasattr(current_user, 'is_owner') and current_user.is_owner:
                owner_id = current_user.raw_id
        except Exception:
            pass
        ip = request.headers.get('X-Forwarded-For', request.remote_addr) if request else None
        if ip and ',' in ip:
            ip = ip.split(',')[0].strip()
        entry = OwnerAuditLog(
            owner_id=owner_id, action=action,
            target_org_id=target_org_id, target_user_id=target_user_id,
            details=details, ip_address=ip,
        )
        db.session.add(entry)
        db.session.commit()
    except Exception as _e:
        try:
            db.session.rollback()
        except Exception:
            pass


# ============== ОБЁРТКИ ДЛЯ FLASK-LOGIN ==============
class AuthUser(UserMixin):
    """Обёртка над User или OwnerUser для flask-login с префиксом типа."""

    def __init__(self, record, user_type):
        self.record = record
        self.user_type = user_type  # 'user' или 'owner'

    def get_id(self):
        prefix = 'u' if self.user_type == 'user' else 'o'
        return f"{prefix}:{self.record.id}"

    # Удобный доступ к полям
    @property
    def username(self):
        if self.user_type == 'user':
            return self.record.username
        return self.record.email

    @property
    def role(self):
        if self.user_type == 'user':
            return self.record.role
        return 'owner'

    @property
    def organization(self):
        if self.user_type == 'user':
            return self.record.organization
        return None

    @property
    def user(self):
        """Актуальная запись User (None для владельцев)."""
        return self.record if self.user_type == 'user' else None

    @property
    def raw_id(self):
        return self.record.id


@login_manager.user_loader
def load_user(uid):
    try:
        prefix, raw_id = uid.split(':', 1)
        rid = int(raw_id)
    except (ValueError, AttributeError):
        return None

    if prefix == 'u':
        u = db.session.get(User, rid)
        return AuthUser(u, 'user') if u else None
    if prefix == 'o':
        o = db.session.get(OwnerUser, rid)
        return AuthUser(o, 'owner') if o else None
    return None


# ============== HELPERS ==============
def trial_days_left(org):
    if not org or not org.trial_ends_at:
        return 0
    delta = org.trial_ends_at - datetime.utcnow()
    return max(0, delta.days)


def count_recent_failed_attempts(ip):
    """Считаем неудачные попытки с этого IP за последние LOGIN_LOCKOUT_MINUTES минут."""
    cutoff = datetime.utcnow() - timedelta(minutes=Config.LOGIN_LOCKOUT_MINUTES)
    return LoginAttempt.query.filter(
        LoginAttempt.ip_address == ip,
        LoginAttempt.success == False,  # noqa: E712
        LoginAttempt.attempted_at >= cutoff,
    ).count()


def log_attempt(ip, username, success):
    try:
        attempt = LoginAttempt(ip_address=ip, username=username, success=success)
        db.session.add(attempt)
        db.session.commit()
    except Exception:
        db.session.rollback()


# ============== ROOT ==============
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.user_type == 'user' and current_user.role == 'admin':
            return redirect('/admin')
        if current_user.user_type == 'user':
            return redirect('/revision')
    return redirect('/login')


# ============== РЕГИСТРАЦИЯ ==============
@app.route('/register', methods=['GET', 'POST'])
def register():
    error = None
    if request.method == 'POST':
        company = (request.form.get('company') or '').strip()
        email = (request.form.get('email') or '').strip().lower()
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        password2 = request.form.get('password2') or ''
        terms = request.form.get('terms')
        pos_system = (request.form.get('pos_system') or '').strip().lower()
        if pos_system not in ('iiko', 'rkeeper', '1c', 'excel', 'other'):
            pos_system = None

        if not (company and email and username and password):
            error = 'Заполните все поля.'
        elif len(password) < 6:
            error = 'Пароль должен быть не короче 6 символов.'
        elif password != password2:
            error = 'Пароли не совпадают.'
        elif not terms:
            error = 'Необходимо согласиться с условиями использования.'
        elif Organization.query.filter_by(owner_email=email).first():
            error = 'Компания с таким email уже зарегистрирована.'

        if error is None:
            try:
                token = secrets.token_urlsafe(32)
                org = Organization(
                    name=company,
                    owner_email=email,
                    email_verified=False,
                    email_verify_token=token,
                    plan='trial',
                    trial_ends_at=datetime.utcnow() + timedelta(days=Config.TRIAL_DAYS),
                    pos_system=pos_system,
                )
                db.session.add(org)
                db.session.flush()

                admin = User(
                    org_id=org.id,
                    username=username,
                    email=email,
                    role='admin',
                )
                admin.set_password(password)
                db.session.add(admin)

                for order, loc_name in enumerate(('Склад', 'Кухня', 'Островок')):
                    db.session.add(Location(org_id=org.id, name=loc_name, sort_order=order))

                db.session.commit()

                # TODO: отправлять email в продакшне
                return redirect(url_for('register_success', token=token))
            except Exception as e:
                db.session.rollback()
                error = f'Ошибка регистрации: {e}'

    return render_template_string(register_html, error=error)


@app.route('/register/success')
def register_success():
    token = request.args.get('token', '')
    verify_url = url_for('verify_email', token=token) if token else None
    return render_template_string(register_success_html, verify_url=verify_url)


@app.route('/verify/<token>')
def verify_email(token):
    org = Organization.query.filter_by(email_verify_token=token).first()
    if not org:
        return render_template_string(
            message_html,
            title='Ссылка недействительна',
            text='Токен подтверждения не найден или уже использован.',
        ), 404

    org.email_verified = True
    org.email_verify_token = None
    db.session.commit()

    admin = User.query.filter_by(org_id=org.id, role='admin').first()
    if admin:
        login_user(AuthUser(admin, 'user'), remember=True)
        admin.last_login_at = datetime.utcnow()
        db.session.commit()
        return redirect('/admin')
    return redirect('/login')


# ============== ЛОГИН / ЛОГАУТ ==============
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    ip = request.remote_addr or 'unknown'

    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''

        # Защита от брутфорса
        if count_recent_failed_attempts(ip) >= Config.MAX_LOGIN_ATTEMPTS:
            error = (
                f'Слишком много неудачных попыток. '
                f'Попробуйте через {Config.LOGIN_LOCKOUT_MINUTES} минут.'
            )
            return render_template_string(
                login_html, error=error,
                now=datetime.now().strftime('%d.%m %H:%M'),
            )

        org = Organization.query.filter_by(owner_email=email).first()
        user = None
        if org:
            user = User.query.filter_by(org_id=org.id, username=username).first()

        if not user or not user.check_password(password):
            log_attempt(ip, username, False)
            error = 'Неверный email компании, логин или пароль.'
        elif org.is_blocked:
            log_attempt(ip, username, False)
            error = 'Компания заблокирована. Обратитесь в поддержку.'
        elif not org.email_verified:
            log_attempt(ip, username, False)
            error = 'Email компании не подтверждён. Проверьте почту.'
        else:
            log_attempt(ip, username, True)
            user.last_login_at = datetime.utcnow()
            db.session.commit()
            login_user(AuthUser(user, 'user'), remember=True)
            if user.role == 'admin':
                return redirect('/admin')
            return redirect('/revision')

    return render_template_string(
        login_html, error=error,
        now=datetime.now().strftime('%d.%m %H:%M'),
    )


@app.route('/logout')
def logout():
    logout_user()
    return redirect('/login')


# ============== АДМИН: ХЕЛПЕРЫ ==============
def _current_org():
    """Возвращает Organization текущего залогиненного пользователя (или None)."""
    if not current_user.is_authenticated:
        return None
    if current_user.user_type != 'user':
        return None
    u = current_user.user
    if not u:
        return None
    return db.session.get(Organization, u.org_id)


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect('/login')
        if current_user.user_type != 'user':
            return redirect('/login')
        u = current_user.user
        if not u or u.role != 'admin':
            return redirect('/login')
        return fn(*args, **kwargs)
    return wrapper


def login_required_user(fn):
    """Доступ для любого User (админ или оператор), но не OwnerUser."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect('/login')
        if current_user.user_type != 'user':
            return redirect('/login')
        if not current_user.user:
            return redirect('/login')
        return fn(*args, **kwargs)
    return wrapper


def _get_or_create_active_revision(org_id, location_id, user_id):
    """Возвращает активную (in_progress) ревизию для (org, location) или создаёт.
    На случай гонки (два оператора одновременно жмут добавить первую позицию)
    после insert-исключения перезапрашиваем существующую запись.
    """
    from sqlalchemy.exc import IntegrityError
    rev = Revision.query.filter_by(
        org_id=org_id, location_id=location_id, status='in_progress'
    ).first()
    if rev:
        return rev
    rev = Revision(
        org_id=org_id, user_id=user_id,
        location_id=location_id, status='in_progress',
    )
    db.session.add(rev)
    try:
        db.session.flush()
    except IntegrityError:
        db.session.rollback()
        rev = Revision.query.filter_by(
            org_id=org_id, location_id=location_id, status='in_progress'
        ).first()
        if rev:
            return rev
        raise
    return rev


def _seed_location_assortment_if_empty(org_id):
    """Если у компании ещё ни на одной локации не настроен ассортимент
    (нет ни одной строки в LocationProduct) — заполняем его «всё × все»:
    каждый существующий товар приписываем ко всем существующим локациям.
    Это сохраняет старое поведение для компаний, у которых уже что-то накоплено,
    до тех пор пока админ не начнёт настраивать списки вручную.

    Идемпотентно и устойчиво к гонке: при IntegrityError откатываемся и
    выходим — значит параллельный запрос уже засеялся.
    """
    from sqlalchemy.exc import IntegrityError
    loc_ids = [l.id for l in Location.query.filter_by(org_id=org_id).all()]
    if not loc_ids:
        return False
    has_any = db.session.query(LocationProduct.id).filter(
        LocationProduct.location_id.in_(loc_ids)
    ).first()
    if has_any:
        return False
    prod_ids = [p.id for p in Product.query.filter_by(org_id=org_id, is_active=True).all()]
    if not prod_ids:
        return False
    for lid in loc_ids:
        for pid in prod_ids:
            db.session.add(LocationProduct(location_id=lid, product_id=pid))
    try:
        db.session.commit()
        return True
    except IntegrityError:
        db.session.rollback()
        return False


def _product_in_location(location_id, product_id):
    return db.session.query(LocationProduct.id).filter_by(
        location_id=location_id, product_id=product_id
    ).first() is not None


def _link_product_to_open_revisions(org_id, product_id, user_id):
    """Создаёт placeholder RevisionItem(qty=0) для каждой открытой (in_progress)
    ревизии компании — но только там, где товар входит в ассортимент локации.
    Возвращает количество затронутых ревизий.
    """
    open_revs = Revision.query.filter_by(org_id=org_id, status='in_progress').all()
    count = 0
    for rev in open_revs:
        if not rev.location_id:
            continue
        if not _product_in_location(rev.location_id, product_id):
            continue
        exists = RevisionItem.query.filter_by(
            revision_id=rev.id, product_id=product_id
        ).first()
        if exists:
            continue
        placeholder = RevisionItem(
            revision_id=rev.id,
            product_id=product_id,
            location_id=rev.location_id,
            quantity=0,
            added_by_user_id=user_id,
        )
        db.session.add(placeholder)
        count += 1
    return count


def _generate_product_code(org_id):
    """Сгенерировать уникальный код товара формата P{org_id}-{NNN}."""
    prefix = f'P{org_id}-'
    max_num = 0
    existing = Product.query.filter_by(org_id=org_id).all()
    pat = re.compile(r'^P' + str(org_id) + r'-(\d+)$')
    for p in existing:
        if p.code:
            m = pat.match(p.code)
            if m:
                try:
                    n = int(m.group(1))
                    if n > max_num:
                        max_num = n
                except ValueError:
                    pass
    next_num = max_num + 1
    return f'{prefix}{next_num:03d}'


# ============== АДМИН: ПАНЕЛЬ ==============
@app.route('/admin')
@admin_required
def admin_panel():
    org = _current_org()
    locations = Location.query.filter_by(org_id=org.id).order_by(Location.sort_order, Location.name).all()
    categories = Category.query.filter_by(org_id=org.id).order_by(Category.sort_order, Category.name).all()
    products = Product.query.filter_by(org_id=org.id).order_by(Product.name).all()
    users = User.query.filter_by(org_id=org.id).order_by(User.created_at).all()

    # Если ассортимент локаций ещё не настраивался — один раз сидим «всё × все»,
    # чтобы старые компании не получили пустые локации после выкатки фичи.
    _seed_location_assortment_if_empty(org.id)

    # Нормы: {(product_id, location_id): qty}
    norms_map = {}
    if products and locations:
        product_ids = [p.id for p in products]
        for n in ProductNorm.query.filter(ProductNorm.product_id.in_(product_ids)).all():
            norms_map[(n.product_id, n.location_id)] = n.norm_qty

    # Ассортимент: {location_id: [product_id, ...]} — для встроенного редактора
    assortment_map = {}
    if locations:
        loc_ids = [l.id for l in locations]
        for lp in LocationProduct.query.filter(
            LocationProduct.location_id.in_(loc_ids)
        ).all():
            assortment_map.setdefault(lp.location_id, []).append(lp.product_id)

    # Группировка товаров по категориям
    cat_map = {c.id: c for c in categories}
    grouped = {}
    for p in products:
        key = p.category_id
        grouped.setdefault(key, []).append(p)

    user_limit_reached = (
        org.plan == 'free' and len(users) >= Config.FREE_MAX_USERS
    )

    # Новый сгенерированный пароль (однократный показ)
    new_user_pwd = request.args.get('new_pwd')
    new_user_name = request.args.get('new_user')

    # Ревизии: запросы (pending) и история (completed)
    pending_revs = Revision.query.filter_by(org_id=org.id, status='pending').order_by(Revision.created_at.desc()).all()
    completed_revs = Revision.query.filter_by(org_id=org.id, status='completed').order_by(Revision.finished_at.desc()).limit(50).all()
    loc_map = {l.id: l for l in locations}
    user_map = {u.id: u for u in users}

    # N+1 fix: одним запросом считаем количество позиций для всех ревизий
    from sqlalchemy import func as _sa_func
    _rev_ids = [r.id for r in pending_revs] + [r.id for r in completed_revs]
    if _rev_ids:
        _items_counts = dict(
            db.session.query(RevisionItem.revision_id, _sa_func.count(RevisionItem.id))
            .filter(RevisionItem.revision_id.in_(_rev_ids))
            .group_by(RevisionItem.revision_id)
            .all()
        )
    else:
        _items_counts = {}

    def _rev_info(r):
        loc = loc_map.get(r.location_id)
        u = user_map.get(r.user_id)
        return {
            'id': r.id,
            'location': loc.name if loc else '—',
            'user': u.username if u else '—',
            'created_at': _fmt_dt(r.created_at, org),
            'finished_at': _fmt_dt(r.finished_at, org),
            'items_count': _items_counts.get(r.id, 0),
        }

    pending_list = [_rev_info(r) for r in pending_revs]
    completed_list = [_rev_info(r) for r in completed_revs]

    # История: группируем завершённые ревизии в «Ревизию ресторана». Локации,
    # подтверждённые одним действием, имеют идентичный finished_at — по нему и
    # группируем. Каждая группа = одна строка истории со списком локаций, общей
    # суммой позиций и кнопкой скачать общий отчёт.
    from collections import OrderedDict as _OD
    _hist_groups = _OD()
    for r in completed_revs:
        _key = r.finished_at.strftime('%Y%m%d%H%M%S') if r.finished_at else ('id%d' % r.id)
        _hist_groups.setdefault(_key, []).append(r)
    completed_groups = []
    for _key, _grp in _hist_groups.items():
        _locn = []
        for _r in _grp:
            _l = loc_map.get(_r.location_id)
            _nm = _l.name if _l else '—'
            if _nm not in _locn:
                _locn.append(_nm)
        _ops = []
        for _r in _grp:
            _u = user_map.get(_r.user_id)
            _un = _u.username if _u else '—'
            if _un not in _ops:
                _ops.append(_un)
        _ref = _grp[0]
        completed_groups.append({
            'download_id': _ref.id,
            'locations_str': ', '.join(_locn),
            'loc_count': len(_locn),
            'operators_str': ', '.join(_ops),
            'finished_at': _fmt_dt(_ref.finished_at, org),
            'total_items': sum(_items_counts.get(_r.id, 0) for _r in _grp),
        })

    # Схлопываем все ожидающие локации в ОДИН запрос «Ревизия ресторана».
    # Оператор завершает зоны по одной, но админ должен видеть единый запрос
    # на завершение всей ревизии — с суммой позиций и списком локаций.
    if pending_list:
        _loc_names = []
        for _r in pending_revs:
            _l = loc_map.get(_r.location_id)
            _nm = _l.name if _l else '—'
            if _nm not in _loc_names:
                _loc_names.append(_nm)
        _ops = []
        for _r in pending_revs:
            _u = user_map.get(_r.user_id)
            _un = _u.username if _u else '—'
            if _un not in _ops:
                _ops.append(_un)
        _dates = [r.created_at for r in pending_revs if r.created_at]
        pending_group = {
            'confirm_id': pending_revs[0].id,   # confirm закроет все pending разом
            'locations': _loc_names,
            'locations_str': ', '.join(_loc_names),
            'loc_count': len(_loc_names),
            'operators_str': ', '.join(_ops),
            'total_items': sum(p['items_count'] for p in pending_list),
            'created_at': _fmt_dt(min(_dates), org) if _dates else '',
        }
    else:
        pending_group = None

    return render_template_string(
        admin_html,
        org=org,
        username=current_user.username,
        locations=locations,
        categories=categories,
        products=products,
        grouped=grouped,
        cat_map=cat_map,
        users=users,
        current_user_id=current_user.raw_id,
        norms_map=norms_map,
        assortment_map=assortment_map,
        days_left=trial_days_left(org),
        user_limit_reached=user_limit_reached,
        free_max_users=Config.FREE_MAX_USERS,
        new_user_pwd=new_user_pwd,
        new_user_name=new_user_name,
        pending_revs=pending_list,
        pending_group=pending_group,
        completed_revs=completed_list,
        completed_groups=completed_groups,
        tz_choices=_TZ_CHOICES,
    )


# ============== АДМИН: ЛОКАЦИИ ==============
@app.route('/admin/locations/add', methods=['POST'])
@admin_required
def admin_add_location():
    org = _current_org()
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('Введите название локации.', 'error')
        return redirect('/admin#tab-locations')
    if Location.query.filter_by(org_id=org.id, name=name).first():
        flash('Локация с таким именем уже существует.', 'error')
        return redirect('/admin#tab-locations')
    try:
        max_order = db.session.query(db.func.max(Location.sort_order)).filter_by(org_id=org.id).scalar() or 0
        db.session.add(Location(org_id=org.id, name=name, sort_order=max_order + 1))
        db.session.commit()
        flash(f'Локация «{name}» добавлена.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-locations')


@app.route('/admin/locations/delete/<int:loc_id>', methods=['POST'])
@admin_required
def admin_delete_location(loc_id):
    org = _current_org()
    loc = Location.query.filter_by(id=loc_id, org_id=org.id).first()
    if not loc:
        flash('Локация не найдена.', 'error')
        return redirect('/admin#tab-locations')
    try:
        from models import RevisionItem, Revision
        # Нормы и привязки ассортимента актуальны только для живой локации —
        # их удаляем безопасно.
        ProductNorm.query.filter_by(location_id=loc.id).delete(synchronize_session=False)
        LocationProduct.query.filter_by(location_id=loc.id).delete(synchronize_session=False)
        # Открытые ревизии на этой локации больше не имеют смысла — отменяем их.
        open_revs = Revision.query.filter(
            Revision.location_id == loc.id,
            Revision.status.in_(['in_progress', 'pending']),
        ).all()
        for _rev in open_revs:
            RevisionItem.query.filter_by(revision_id=_rev.id).delete(synchronize_session=False)
            _rev.status = 'cancelled'
            _rev.location_id = None
            _rev.finished_at = datetime.utcnow()
        # Для завершённых/отменённых ревизий историю сохраняем — обнуляем только
        # ссылки на локацию (как в заголовке ревизии, так и в её позициях),
        # чтобы FK не падала.
        RevisionItem.query.filter(
            RevisionItem.location_id == loc.id,
        ).update({'location_id': None}, synchronize_session=False)
        Revision.query.filter_by(location_id=loc.id).update(
            {'location_id': None}, synchronize_session=False,
        )
        db.session.delete(loc)
        db.session.commit()
        flash('Локация удалена. Завершённые ревизии остались в истории, незавершённые — отменены.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Не удалось удалить локацию: {e}', 'error')
    return redirect('/admin#tab-locations')


# ============== АДМИН: АССОРТИМЕНТ ЛОКАЦИЙ ==============
@app.route('/admin/locations/<int:loc_id>/assortment/toggle/<int:pid>', methods=['POST'])
@admin_required
def admin_assortment_toggle(loc_id, pid):
    """Включить/выключить товар в ассортименте локации. Возвращает новое состояние."""
    org = _current_org()
    loc = Location.query.filter_by(id=loc_id, org_id=org.id).first()
    prod = Product.query.filter_by(id=pid, org_id=org.id).first()
    if not loc or not prod:
        return jsonify({'ok': False, 'error': 'Локация или товар не найдены.'}), 404
    try:
        existing = LocationProduct.query.filter_by(
            location_id=loc.id, product_id=prod.id
        ).first()
        if existing:
            db.session.delete(existing)
            # Сохраняем синхронизацию ассортимент ↔ отчёт: товар, убранный из
            # локации, должен исчезнуть и из её открытых ревизий целиком —
            # и placeholder-нули, и фактические подсчёты. Завершённые ревизии
            # не трогаем, это исторический отчёт.
            open_rev_ids = [r.id for r in Revision.query.filter_by(
                org_id=org.id, location_id=loc.id, status='in_progress'
            ).all()]
            removed_counts = 0
            if open_rev_ids:
                removed_counts = RevisionItem.query.filter(
                    RevisionItem.revision_id.in_(open_rev_ids),
                    RevisionItem.product_id == prod.id,
                ).delete(synchronize_session=False)
            db.session.commit()
            return jsonify({'ok': True, 'selected': False, 'removed_counts': removed_counts or 0})
        else:
            db.session.add(LocationProduct(location_id=loc.id, product_id=prod.id))
            # Если есть открытая ревизия — сразу создаём placeholder, чтобы товар
            # появился у оператора и попал в отчёт.
            linked = _link_product_to_open_revisions(org.id, prod.id, current_user.raw_id)
            db.session.commit()
            return jsonify({'ok': True, 'selected': True, 'linked_to_revisions': linked})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/admin/locations/<int:loc_id>/assortment/clear', methods=['POST'])
@admin_required
def admin_assortment_clear(loc_id):
    """Полностью очищает ассортимент локации одним действием.
    Из открытых ревизий этой локации удаляются ВСЕ позиции (и placeholder-нули,
    и фактические подсчёты) — чтобы состояние «товара нет в ассортименте»
    однозначно отражалось в отчёте. Завершённые ревизии не трогаем —
    это исторический отчёт.
    """
    org = _current_org()
    loc = Location.query.filter_by(id=loc_id, org_id=org.id).first()
    if not loc:
        return jsonify({'ok': False, 'error': 'Локация не найдена.'}), 404
    try:
        removed = LocationProduct.query.filter_by(location_id=loc.id).count()
        LocationProduct.query.filter_by(location_id=loc.id).delete(synchronize_session=False)
        open_rev_ids = [r.id for r in Revision.query.filter_by(
            org_id=org.id, location_id=loc.id, status='in_progress'
        ).all()]
        removed_counts = 0
        if open_rev_ids:
            removed_counts = RevisionItem.query.filter(
                RevisionItem.revision_id.in_(open_rev_ids),
            ).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({
            'ok': True, 'removed': removed, 'removed_counts': removed_counts or 0,
            'location': loc.name,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/admin/locations/<int:loc_id>/assortment', methods=['GET'])
@admin_required
def admin_assortment_get(loc_id):
    """Возвращает текущее множество выбранных товаров для локации (id-шников)."""
    org = _current_org()
    loc = Location.query.filter_by(id=loc_id, org_id=org.id).first()
    if not loc:
        return jsonify({'ok': False, 'error': 'Локация не найдена.'}), 404
    selected = [lp.product_id for lp in LocationProduct.query.filter_by(
        location_id=loc.id
    ).all()]
    return jsonify({'ok': True, 'selected': selected})


# ============== АДМИН: КАТЕГОРИИ ==============
@app.route('/admin/categories/add', methods=['POST'])
@admin_required
def admin_add_category():
    org = _current_org()
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('Введите название категории.', 'error')
        return redirect('/admin#tab-categories')
    if Category.query.filter_by(org_id=org.id, name=name).first():
        flash('Категория уже существует.', 'error')
        return redirect('/admin#tab-categories')
    try:
        max_order = db.session.query(db.func.max(Category.sort_order)).filter_by(org_id=org.id).scalar() or 0
        db.session.add(Category(org_id=org.id, name=name, sort_order=max_order + 1))
        db.session.commit()
        flash(f'Категория «{name}» добавлена.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-categories')


@app.route('/admin/categories/delete/<int:cat_id>', methods=['POST'])
@admin_required
def admin_delete_category(cat_id):
    org = _current_org()
    cat = Category.query.filter_by(id=cat_id, org_id=org.id).first()
    if not cat:
        flash('Категория не найдена.', 'error')
        return redirect('/admin#tab-categories')
    has_products = Product.query.filter_by(org_id=org.id, category_id=cat.id).count()
    if has_products:
        flash(f'Нельзя удалить категорию: в ней есть товары ({has_products}). Сначала удалите товары.', 'error')
        return redirect('/admin#tab-categories')
    try:
        db.session.delete(cat)
        db.session.commit()
        flash('Категория удалена.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-categories')


# ============== АДМИН: ТОВАРЫ ==============
@app.route('/admin/products/add', methods=['POST'])
@admin_required
def admin_add_product():
    org = _current_org()
    name = _sanitize_product_name((request.form.get('name') or '').strip())
    category_id = request.form.get('category_id') or ''
    new_category = (request.form.get('new_category') or '').strip()
    unit = (request.form.get('unit') or 'шт').strip() or 'шт'
    code = (request.form.get('code') or '').strip()

    is_xhr = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if not name:
        if is_xhr:
            return jsonify({'ok': False, 'error': 'Введите название товара.'}), 400
        flash('Введите название товара.', 'error')
        return redirect('/admin#tab-products')

    try:
        # Создать категорию inline, если нужно
        cat_id = None
        if new_category:
            existing = Category.query.filter_by(org_id=org.id, name=new_category).first()
            if existing:
                cat_id = existing.id
            else:
                max_order = db.session.query(db.func.max(Category.sort_order)).filter_by(org_id=org.id).scalar() or 0
                c = Category(org_id=org.id, name=new_category, sort_order=max_order + 1)
                db.session.add(c)
                db.session.flush()
                cat_id = c.id
        elif category_id:
            c = Category.query.filter_by(id=int(category_id), org_id=org.id).first()
            if c:
                cat_id = c.id

        if not code:
            code = _generate_product_code(org.id)
        else:
            if Product.query.filter_by(org_id=org.id, code=code).first():
                if is_xhr:
                    db.session.rollback()
                    return jsonify({'ok': False, 'error': f'Товар с кодом «{code}» уже существует.'}), 400
                flash(f'Товар с кодом «{code}» уже существует.', 'error')
                db.session.rollback()
                return redirect('/admin#tab-products')

        p = Product(org_id=org.id, name=name, category_id=cat_id, unit=unit, code=code)
        db.session.add(p)
        db.session.flush()

        # Привязываем новый товар ко всем открытым ревизиям компании (placeholder qty=0),
        # чтобы он сразу попал в подсчёт оператора и в итоговый отчёт.
        linked_revs = _link_product_to_open_revisions(org.id, p.id, current_user.raw_id)

        db.session.commit()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            cat_name = ''
            if cat_id:
                c2 = Category.query.get(cat_id)
                cat_name = c2.name if c2 else ''
            msg = f'Товар «{name}» добавлен.'
            if linked_revs:
                msg += f' Добавлен в открытые ревизии ({linked_revs}).'
            return jsonify({'ok': True, 'msg': msg, 'product': {
                'id': p.id, 'name': p.name, 'code': p.code or '—', 'unit': p.unit,
                'category_id': p.category_id, 'category_name': cat_name,
            }})
        flash(f'Товар «{name}» добавлен (код {code}).', 'success')
    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'error': str(e)}), 400
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-products')


@app.route('/admin/products/edit/<int:pid>', methods=['POST'])
@admin_required
def admin_edit_product(pid):
    is_xhr = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    org = _current_org()
    p = Product.query.filter_by(id=pid, org_id=org.id).first()
    if not p:
        if is_xhr:
            return jsonify({'ok': False, 'error': 'Товар не найден.'}), 404
        flash('Товар не найден.', 'error')
        return redirect('/admin#tab-products')
    name = _sanitize_product_name((request.form.get('name') or '').strip())
    category_id = request.form.get('category_id') or ''
    unit = (request.form.get('unit') or 'шт').strip() or 'шт'
    code = (request.form.get('code') or '').strip()
    if not name:
        if is_xhr:
            return jsonify({'ok': False, 'error': 'Название не может быть пустым.'}), 400
        flash('Название не может быть пустым.', 'error')
        return redirect('/admin#tab-products')
    try:
        p.name = name
        p.unit = unit
        if category_id:
            c = Category.query.filter_by(id=int(category_id), org_id=org.id).first()
            p.category_id = c.id if c else None
        else:
            p.category_id = None
        if code and code != p.code:
            dup = Product.query.filter(
                Product.org_id == org.id, Product.code == code, Product.id != p.id
            ).first()
            if dup:
                if is_xhr:
                    db.session.rollback()
                    return jsonify({'ok': False, 'error': 'Код уже используется другим товаром.'}), 400
                flash('Код уже используется другим товаром.', 'error')
                db.session.rollback()
                return redirect('/admin#tab-products')
            p.code = code
        elif not code:
            p.code = _generate_product_code(org.id) if not p.code else p.code
        db.session.commit()
        if is_xhr:
            return jsonify({'ok': True, 'msg': 'Товар обновлён.'})
        flash('Товар обновлён.', 'success')
    except Exception as e:
        db.session.rollback()
        if is_xhr:
            return jsonify({'ok': False, 'error': str(e)}), 400
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-products')


@app.route('/admin/products/delete/<int:pid>', methods=['POST'])
@admin_required
def admin_delete_product(pid):
    is_xhr = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    org = _current_org()
    p = Product.query.filter_by(id=pid, org_id=org.id).first()
    if not p:
        if is_xhr:
            return jsonify({'ok': False, 'error': 'Товар не найден.'}), 404
        flash('Товар не найден.', 'error')
        return redirect('/admin#tab-products')
    try:
        from models import RevisionItem
        # Считаем, в скольких открытых ревизиях есть позиции этого товара —
        # это число и покажем админу для подтверждения изменений.
        open_rev_ids = [r.id for r in Revision.query.filter_by(
            org_id=org.id, status='in_progress'
        ).all()]
        removed_from_open = 0
        if open_rev_ids:
            removed_from_open = RevisionItem.query.filter(
                RevisionItem.product_id == p.id,
                RevisionItem.revision_id.in_(open_rev_ids),
            ).count()
        prod_name = p.name
        RevisionItem.query.filter_by(product_id=p.id).delete(synchronize_session=False)
        ProductNorm.query.filter_by(product_id=p.id).delete(synchronize_session=False)
        LocationProduct.query.filter_by(product_id=p.id).delete(synchronize_session=False)
        db.session.delete(p)
        db.session.commit()
        if is_xhr:
            msg = f'Товар «{prod_name}» удалён.'
            if removed_from_open:
                msg += f' Убран из открытых ревизий ({removed_from_open}).'
            return jsonify({'ok': True, 'msg': msg})
        flash('Товар удалён.', 'success')
    except Exception as e:
        db.session.rollback()
        if is_xhr:
            return jsonify({'ok': False, 'error': str(e)}), 400
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-products')


# ============== АДМИН: ПОЛЬЗОВАТЕЛИ ==============
@app.route('/admin/users/add', methods=['POST'])
@admin_required
def admin_add_user():
    org = _current_org()
    count = User.query.filter_by(org_id=org.id).count()
    if org.plan == 'free' and count >= Config.FREE_MAX_USERS:
        flash(f'Лимит тарифа FREE: максимум {Config.FREE_MAX_USERS} пользователей. Обновите тариф.', 'error')
        return redirect('/admin#tab-users')

    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    email = (request.form.get('email') or '').strip().lower() or None
    if not username:
        flash('Введите логин.', 'error')
        return redirect('/admin#tab-users')
    if User.query.filter_by(org_id=org.id, username=username).first():
        flash('Пользователь с таким логином уже существует.', 'error')
        return redirect('/admin#tab-users')

    generated = None
    if not password:
        password = secrets.token_urlsafe(8)
        generated = password

    try:
        u = User(org_id=org.id, username=username, email=email, role='operator')
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash(f'Оператор «{username}» создан.', 'success')
        # Всегда показываем реквизиты входа (админ должен передать их оператору)
        return redirect(url_for('admin_panel', new_user=username, new_pwd=password) + '#tab-users')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-users')


@app.route('/admin/users/delete/<int:uid>', methods=['POST'])
@admin_required
def admin_delete_user(uid):
    org = _current_org()
    if uid == current_user.raw_id:
        flash('Нельзя удалить самого себя.', 'error')
        return redirect('/admin#tab-users')
    u = User.query.filter_by(id=uid, org_id=org.id).first()
    if not u:
        flash('Пользователь не найден.', 'error')
        return redirect('/admin#tab-users')
    try:
        from models import RevisionItem, Revision
        # Историю ревизий НЕ удаляем — просто обнуляем ссылку на этого юзера.
        # Это сохраняет завершённые отчёты в неизменном виде.
        RevisionItem.query.filter_by(added_by_user_id=u.id).update(
            {'added_by_user_id': None}, synchronize_session=False
        )
        Revision.query.filter_by(user_id=u.id).update(
            {'user_id': None}, synchronize_session=False
        )
        db.session.delete(u)
        db.session.commit()
        flash('Пользователь удалён. Его прошлые ревизии и подсчёты остались в истории.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-users')


# ============== АДМИН: НОРМЫ ==============
@app.route('/admin/norms/save', methods=['POST'])
@admin_required
def admin_save_norms():
    org = _current_org()
    data = request.get_json(silent=True) or {}
    try:
        product_ids = {p.id for p in Product.query.filter_by(org_id=org.id).all()}
        location_ids = {l.id for l in Location.query.filter_by(org_id=org.id).all()}
        saved = 0
        for pid_s, loc_map in data.items():
            try:
                pid = int(pid_s)
            except (TypeError, ValueError):
                continue
            if pid not in product_ids:
                continue
            for lid_s, qty in (loc_map or {}).items():
                try:
                    lid = int(lid_s)
                    q = float(qty) if qty not in (None, '', 'null') else 0.0
                except (TypeError, ValueError):
                    continue
                if lid not in location_ids:
                    continue
                n = ProductNorm.query.filter_by(product_id=pid, location_id=lid).first()
                if q <= 0:
                    if n:
                        db.session.delete(n)
                        saved += 1
                    continue
                if n:
                    if n.norm_qty != q:
                        n.norm_qty = q
                        saved += 1
                else:
                    db.session.add(ProductNorm(product_id=pid, location_id=lid, norm_qty=q))
                    saved += 1
        db.session.commit()
        return jsonify({'ok': True, 'saved': saved})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


# ============== АДМИН: ИМПОРТ ==============
IMPORT_COLUMNS = ['Категория', 'Название', 'Код', 'Ед. изм.']


@app.route('/admin/import/template')
@admin_required
def admin_import_template():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Товары'
    ws.append(IMPORT_COLUMNS)
    ws.append(['Напитки', 'Кола 0.5л', '', 'шт'])
    ws.append(['Напитки', 'Вода 0.5л', 'P-WATER-500', 'шт'])
    ws.append(['Булочки', 'Булочка бриошь', '', 'шт'])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name='import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# Сокращение "в асс" / "в ассорт." / "в ассортименте" — частая ошибка в импортах
# из iiko/1С: на слух читается двусмысленно. Чистим прямо во входных данных,
# чтобы такие имена никогда не попадали в БД, даже на короткое время до старта.
_IMPORT_ASS_RE = re.compile(
    r'\s*\(?\s*в\s+асс(?:\.|орт\.?|ортименте)?(?=\s|[,.)\]]|$)\s*\)?',
    re.IGNORECASE,
)


def _sanitize_product_name(name):
    if not name:
        return name
    new = _IMPORT_ASS_RE.sub(' ', name)
    new = re.sub(r'\s+', ' ', new).strip()
    new = re.sub(r'\s+([,.)\]])', r'\1', new)
    new = re.sub(r'\(\s+', '(', new)
    return new


def _parse_iiko_bank(file_bytes):
    """Разобрать «Бланк инвентаризации» из iiko.

    iiko выгружает бланк либо как настоящий .xlsx, либо как XML SpreadsheetML
    с расширением .xls. Поддерживаем оба. Структура: колонка B — код, C —
    наименование, F — единица измерения; строки-заголовки категорий содержат
    только текст в колонке B; верхние строки — метаинформация и шапка таблицы.

    Возвращает кортеж (rows, xlsx_bytes):
      • rows — список dict {Категория, Название, Код, Ед. изм.} для _import_rows;
      • xlsx_bytes — нормализованный .xlsx-бланк (та же раскладка), который
        сохраняется как шаблон отчёта и заполняется остатками при ревизии.
    """
    import io as _io
    import xml.etree.ElementTree as _ET
    from openpyxl import load_workbook as _lw, Workbook as _WB
    from openpyxl.styles import Font as _F, Alignment as _Al, Border as _Bd, Side as _Sd
    from openpyxl.utils import get_column_letter as _gcl

    _meta_prefixes = (
        'бланк', 'на дату', 'дата печати', 'подразделения', 'товарные группы',
        'валюта', 'способ группировки', 'ответственное', 'организация',
        'со склада', 'на склад', 'основание', '№ пф',
    )

    def _is_xml(b):
        head = b[:200].lstrip()
        return head[:5] == b'<?xml' or b'spreadsheet' in head.lower()

    # ── 1. Считываем строки как сетку значений ──
    grid = []  # list[list[str]]
    if _is_xml(file_bytes):
        NS = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
        root = _ET.fromstring(file_bytes)
        for ws in root.findall('ss:Worksheet', NS):
            table = ws.find('ss:Table', NS)
            if table is None:
                continue
            for row in table.findall('ss:Row', NS):
                cells, col = [], 1
                for cell in row.findall('ss:Cell', NS):
                    idx = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
                    if idx:
                        col = int(idx)
                    while len(cells) < col - 1:
                        cells.append('')
                    d = cell.find('ss:Data', NS)
                    cells.append(d.text if d is not None and d.text is not None else '')
                    col += 1
                grid.append(cells)
            break  # только первый лист
    else:
        wb = _lw(_io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            grid.append(['' if v is None else str(v) for v in row])

    def _g(r, c):  # 1-based колонка
        return (r[c - 1].strip() if c - 1 < len(r) and r[c - 1] is not None else '')

    # ── 2. Извлекаем товары по категориям ──
    rows = []
    current_cat = None
    for r in grid:
        b, c, f = _g(r, 2), _g(r, 3), _g(r, 6)
        if not b and not c:
            continue
        # товарная строка: есть наименование (C) и единица (F)
        if c and f:
            rows.append({
                'Категория': current_cat or '',
                'Название': c,
                'Код': b if not b.lower() in ('код',) else '',
                'Ед. изм.': f,
            })
            continue
        # заголовок категории: только текст в B, без C и F
        if b and not c and not f:
            bl = b.lower()
            if bl in ('товар', 'код'):
                continue
            if any(bl.startswith(p) for p in _meta_prefixes):
                continue
            current_cat = b

    # ── 3. Нормализуем бланк в .xlsx (для шаблона отчёта) ──
    if _is_xml(file_bytes):
        # XML SpreadsheetML openpyxl не читает — пересобираем в чистый .xlsx,
        # сохраняя ту же раскладку (B=код, C=наим., F=ед., G=остаток).
        wb2 = _WB()
        ws2 = wb2.active
        ws2.title = 'Page1'
        thin = _Sd(border_style='thin', color='CCCCCC')
        border = _Bd(left=thin, right=thin, top=thin, bottom=thin)
        for ri, r in enumerate(grid, start=1):
            for ci in range(1, 9):
                val = _g(r, ci)
                if val:
                    ws2.cell(row=ri, column=ci, value=val)
        widths = {1: 4, 2: 10, 3: 48, 6: 10, 7: 18, 8: 14}
        for col, w in widths.items():
            ws2.column_dimensions[_gcl(col)].width = w
        out = _io.BytesIO()
        wb2.save(out)
        xlsx_bytes = out.getvalue()
    else:
        xlsx_bytes = file_bytes  # уже .xlsx — используем как есть

    return rows, xlsx_bytes


def _import_rows(org_id, rows):
    """rows — iterable of dict-like {Категория, Название, Код, Ед. изм.}."""
    added = 0
    skipped = 0
    errors = []

    cat_cache = {c.name: c for c in Category.query.filter_by(org_id=org_id).all()}
    existing_codes = {p.code for p in Product.query.filter_by(org_id=org_id).all() if p.code}
    existing_names = {(p.category_id, p.name.lower()) for p in Product.query.filter_by(org_id=org_id).all()}

    for i, row in enumerate(rows, start=2):
        try:
            cat_name = (row.get('Категория') or '').strip()
            name = _sanitize_product_name((row.get('Название') or '').strip())
            code = (row.get('Код') or '').strip()
            unit = (row.get('Ед. изм.') or 'шт').strip() or 'шт'
            if not name:
                skipped += 1
                errors.append(f'Строка {i}: пустое название')
                continue

            cat_id = None
            if cat_name:
                c = cat_cache.get(cat_name)
                if not c:
                    max_order = db.session.query(db.func.max(Category.sort_order)).filter_by(org_id=org_id).scalar() or 0
                    c = Category(org_id=org_id, name=cat_name, sort_order=max_order + 1)
                    db.session.add(c)
                    db.session.flush()
                    cat_cache[cat_name] = c
                cat_id = c.id

            key = (cat_id, name.lower())
            if key in existing_names:
                skipped += 1
                errors.append(f'Строка {i}: дубликат «{name}»')
                continue

            if code:
                if code in existing_codes:
                    skipped += 1
                    errors.append(f'Строка {i}: код «{code}» занят')
                    continue
            else:
                code = _generate_product_code(org_id)
                while code in existing_codes:
                    # на случай конкуренции
                    num = int(code.rsplit('-', 1)[-1]) + 1
                    code = f'P{org_id}-{num:03d}'

            p = Product(org_id=org_id, name=name, category_id=cat_id, unit=unit, code=code)
            db.session.add(p)
            existing_codes.add(code)
            existing_names.add(key)
            added += 1
        except Exception as e:
            skipped += 1
            errors.append(f'Строка {i}: {e}')

    db.session.commit()
    return added, skipped, errors


@app.route('/admin/import/upload', methods=['POST'])
@admin_required
def admin_import_upload():
    org = _current_org()
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Файл не выбран'}), 400

    filename = f.filename.lower()
    rows = []
    try:
        if filename.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = None
            for r in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c).strip() if c is not None else '' for c in r]
                    continue
                if all(c is None or str(c).strip() == '' for c in r):
                    continue
                row = {}
                for idx, col in enumerate(headers):
                    if idx < len(r):
                        val = r[idx]
                        row[col] = '' if val is None else str(val)
                rows.append(row)
        elif filename.endswith('.csv'):
            data = f.read().decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(data))
            for row in reader:
                rows.append({k.strip() if k else '': (v or '') for k, v in row.items()})
        else:
            return jsonify({'ok': False, 'error': 'Поддерживаются только .xlsx и .csv'}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Ошибка чтения файла: {e}'}), 400

    try:
        added, skipped, errors = _import_rows(org.id, rows)
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400

    return jsonify({'ok': True, 'added': added, 'skipped': skipped, 'errors': errors[:20]})


# Предустановленные наборы товаров
PRESETS = {
    'burger': {
        'name': 'Бургерная',
        'items': [
            ('Напитки', 'Кола 0.5л', 'шт'), ('Напитки', 'Спрайт 0.5л', 'шт'),
            ('Напитки', 'Фанта 0.5л', 'шт'), ('Напитки', 'Вода 0.5л', 'шт'),
            ('Напитки', 'Сок яблочный', 'шт'), ('Напитки', 'Сок апельсиновый', 'шт'),
            ('Булочки', 'Булочка бриошь', 'шт'), ('Булочки', 'Булочка классик', 'шт'),
            ('Булочки', 'Булочка с кунжутом', 'шт'),
            ('Мясо', 'Котлета говяжья', 'шт'), ('Мясо', 'Котлета куриная', 'шт'),
            ('Мясо', 'Бекон', 'г'),
            ('Соусы', 'Кетчуп', 'г'), ('Соусы', 'Майонез', 'г'),
            ('Соусы', 'Горчица', 'г'), ('Соусы', 'Соус барбекю', 'г'),
            ('Соусы', 'Соус сырный', 'г'),
            ('Овощи', 'Лук', 'г'), ('Овощи', 'Помидор', 'г'),
            ('Овощи', 'Огурец', 'г'), ('Овощи', 'Салат', 'г'),
            ('Овощи', 'Сыр чеддер', 'шт'), ('Овощи', 'Сыр моцарелла', 'шт'),
            ('Картофель', 'Фри', 'г'), ('Картофель', 'Наггетсы', 'шт'),
        ],
    },
    'coffee': {
        'name': 'Кофейня',
        'items': [
            ('Кофе', 'Эспрессо', 'г'), ('Кофе', 'Арабика зерно', 'г'),
            ('Кофе', 'Робуста зерно', 'г'),
            ('Молоко', 'Молоко 3.2%', 'мл'), ('Молоко', 'Молоко растительное', 'мл'),
            ('Молоко', 'Сливки', 'мл'),
            ('Сиропы', 'Сироп ваниль', 'мл'), ('Сиропы', 'Сироп карамель', 'мл'),
            ('Сиропы', 'Сироп орех', 'мл'),
            ('Напитки', 'Чай черный', 'шт'), ('Напитки', 'Чай зеленый', 'шт'),
            ('Напитки', 'Какао', 'г'),
            ('Десерты', 'Круассан', 'шт'), ('Десерты', 'Маффин', 'шт'),
            ('Десерты', 'Чизкейк', 'шт'),
            ('Расходники', 'Стакан 200мл', 'шт'), ('Расходники', 'Стакан 300мл', 'шт'),
            ('Расходники', 'Крышка', 'шт'), ('Расходники', 'Трубочка', 'шт'),
        ],
    },
    'bakery': {
        'name': 'Пекарня',
        'items': [
            ('Мука', 'Мука пшеничная', 'кг'), ('Мука', 'Мука ржаная', 'кг'),
            ('Мука', 'Мука цельнозерновая', 'кг'),
            ('Хлеб', 'Батон', 'шт'), ('Хлеб', 'Хлеб белый', 'шт'),
            ('Хлеб', 'Хлеб ржаной', 'шт'), ('Хлеб', 'Багет', 'шт'),
            ('Выпечка', 'Круассан', 'шт'), ('Выпечка', 'Булочка с корицей', 'шт'),
            ('Выпечка', 'Слойка с яблоком', 'шт'), ('Выпечка', 'Плюшка', 'шт'),
            ('Торты', 'Наполеон', 'шт'), ('Торты', 'Медовик', 'шт'),
            ('Торты', 'Чизкейк', 'шт'),
            ('Ингредиенты', 'Масло сливочное', 'г'), ('Ингредиенты', 'Сахар', 'кг'),
            ('Ингредиенты', 'Дрожжи', 'г'), ('Ингредиенты', 'Яйца', 'шт'),
        ],
    },
    'sushi': {
        'name': 'Суши',
        'items': [
            ('Рис', 'Рис суши', 'кг'), ('Рис', 'Уксус рисовый', 'мл'),
            ('Рыба', 'Лосось', 'г'), ('Рыба', 'Тунец', 'г'),
            ('Рыба', 'Угорь', 'г'), ('Рыба', 'Креветки', 'шт'),
            ('Овощи', 'Огурец', 'шт'), ('Овощи', 'Авокадо', 'шт'),
            ('Овощи', 'Зелень', 'г'),
            ('Нори', 'Нори', 'лист'),
            ('Соусы', 'Соевый соус', 'мл'), ('Соусы', 'Васаби', 'г'),
            ('Соусы', 'Имбирь', 'г'), ('Соусы', 'Унаги', 'мл'),
            ('Упаковка', 'Контейнер', 'шт'), ('Упаковка', 'Палочки', 'пара'),
        ],
    },
}


@app.route('/admin/import/preset/<ptype>', methods=['POST'])
@admin_required
def admin_import_preset(ptype):
    preset = PRESETS.get(ptype)
    if not preset:
        flash('Неизвестный тип заведения.', 'error')
        return redirect('/admin#tab-import')
    org = _current_org()
    rows = [{'Категория': c, 'Название': n, 'Код': '', 'Ед. изм.': u} for c, n, u in preset['items']]
    try:
        added, skipped, _ = _import_rows(org.id, rows)
        flash(f'Шаблон «{preset["name"]}» применён: добавлено {added}, пропущено {skipped}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-import')


@app.route('/admin/iiko/upload', methods=['POST'])
@admin_required
def admin_iiko_upload():
    """Загрузка «Бланка инвентаризации» из iiko одним файлом:
    1) парсим товары и категории и импортируем их;
    2) сохраняем нормализованный бланк как шаблон отчёта — при завершении
       ревизии остатки будут выгружаться в привычный для клиента вид.
    """
    org = _current_org()
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Файл не выбран.', 'error')
        return redirect('/admin#tab-import')
    fn = f.filename.lower()
    if not (fn.endswith('.xls') or fn.endswith('.xlsx')):
        flash('Нужен файл бланка инвентаризации из iiko (.xls или .xlsx).', 'error')
        return redirect('/admin#tab-import')
    try:
        data = f.read()
        rows, xlsx_bytes = _parse_iiko_bank(data)
        if not rows:
            flash('Не удалось распознать товары. Убедитесь, что это «Бланк инвентаризации» из iiko (колонки: код, наименование, ед. изм.).', 'error')
            return redirect('/admin#tab-import')
        added, skipped, _errs = _import_rows(org.id, rows)
        org.excel_template = xlsx_bytes
        if not org.pos_system:
            org.pos_system = 'iiko'
        db.session.commit()
        flash(
            f'Бланк iiko загружен: добавлено товаров {added}, пропущено {skipped}. '
            f'Отчёт теперь выгружается в вашем формате.',
            'success',
        )
    except Exception as e:
        db.session.rollback()
        flash(f'Не удалось обработать файл: {e}', 'error')
    return redirect('/admin#tab-import')


# ============== РЕВИЗИЯ (ОПЕРАТОР/АДМИН) ==============
@app.route('/revision')
@login_required_user
def revision():
    org = _current_org()
    if not org:
        return redirect('/login')

    locations = Location.query.filter_by(org_id=org.id).order_by(Location.sort_order, Location.name).all()
    categories = Category.query.filter_by(org_id=org.id).order_by(Category.sort_order, Category.name).all()
    all_products = Product.query.filter_by(org_id=org.id, is_active=True).order_by(Product.name).all()

    # Выбранная локация (по ?location=Name) — иначе первая
    selected_name = request.args.get('location')
    selected_loc = None
    if selected_name:
        selected_loc = next((l for l in locations if l.name == selected_name), None)
    if not selected_loc and locations:
        selected_loc = locations[0]

    # Ассортимент локации: показываем только товары, явно привязанные к ней.
    # Бэк-компат: если у компании ни на одной локации не настроен ассортимент,
    # один раз заполняем «всё × все», чтобы старые клиенты не получили пустые экраны.
    _seed_location_assortment_if_empty(org.id)
    if selected_loc:
        allowed_ids = {lp.product_id for lp in LocationProduct.query.filter_by(
            location_id=selected_loc.id
        ).all()}
        products = [p for p in all_products if p.id in allowed_ids]
    else:
        products = []

    # Нормы для всех товаров этой локации
    norms_map = {}  # product_id -> norm_qty
    if selected_loc and products:
        for n in ProductNorm.query.filter_by(location_id=selected_loc.id).all():
            norms_map[n.product_id] = n.norm_qty

    # Активная ревизия локации (не создаём — только читаем)
    # Показываем и in_progress, и pending (пока ждём решения администратора)
    current_rev = None
    qty_map = {}  # product_id -> сумма quantity в активной ревизии на этой локации
    if selected_loc:
        current_rev = Revision.query.filter(
            Revision.org_id == org.id,
            Revision.location_id == selected_loc.id,
            Revision.status.in_(['in_progress', 'pending']),
        ).order_by(Revision.created_at.desc()).first()
        if current_rev:
            items = RevisionItem.query.filter_by(
                revision_id=current_rev.id, location_id=selected_loc.id
            ).all()
            for it in items:
                qty_map[it.product_id] = qty_map.get(it.product_id, 0) + (it.quantity or 0)
            qty_map = {k: round(v, 3) for k, v in qty_map.items()}

    # Группировка товаров по категориям (с учётом "без категории")
    cat_map = {c.id: c for c in categories}
    grouped = {}
    for p in products:
        grouped.setdefault(p.category_id, []).append(p)
    # Упорядоченный список (cat_obj или None, [products...])
    grouped_ordered = []
    for c in categories:
        items = grouped.get(c.id)
        if items:
            grouped_ordered.append((c, items))
    if grouped.get(None):
        grouped_ordered.append((None, grouped[None]))

    is_admin = (current_user.user and current_user.user.role == 'admin')

    total_products = sum(len(items) for _, items in grouped_ordered)
    counted_products = sum(1 for pid, q in qty_map.items() if q and q > 0)
    progress_pct = int((counted_products / total_products) * 100) if total_products else 0

    return render_template_string(
        revision_html,
        org=org,
        username=current_user.username,
        locations=locations,
        selected=selected_loc,
        grouped=grouped_ordered,
        qty_map=qty_map,
        norms_map=norms_map,
        is_admin=is_admin,
        rev_status=(current_rev.status if current_rev else None),
        total_products=total_products,
        counted_products=counted_products,
        progress_pct=progress_pct,
    )


@app.route('/revision/add', methods=['POST'])
@login_required_user
def revision_add():
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'error': 'no org'}), 400
    try:
        location_id = int(request.form.get('location_id') or 0)
        product_id = int(request.form.get('product_id') or 0)
        raw_count = str(request.form.get('count') or '0').replace(',', '.')
        count = round(float(raw_count), 3)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'bad params'}), 400

    if count <= 0:
        return jsonify({'ok': False, 'error': 'count must be > 0'}), 400

    loc = Location.query.filter_by(id=location_id, org_id=org.id).first()
    prod = Product.query.filter_by(id=product_id, org_id=org.id).first()
    if not loc or not prod:
        return jsonify({'ok': False, 'error': 'not found'}), 404

    # Нельзя добавлять позиции, пока ревизия ожидает подтверждения администратором
    pending_rev = Revision.query.filter_by(
        org_id=org.id, location_id=loc.id, status='pending'
    ).first()
    if pending_rev:
        return jsonify({'ok': False, 'error': 'Ревизия ожидает подтверждения администратором'}), 400

    try:
        rev = _get_or_create_active_revision(org.id, loc.id, current_user.raw_id)
        item = RevisionItem(
            revision_id=rev.id,
            product_id=prod.id,
            location_id=loc.id,
            quantity=count,
            added_by_user_id=current_user.raw_id,
        )
        db.session.add(item)
        db.session.commit()
        # Пересчитать суммарное количество
        total = db.session.query(db.func.coalesce(db.func.sum(RevisionItem.quantity), 0)).filter_by(
            revision_id=rev.id, product_id=prod.id, location_id=loc.id
        ).scalar() or 0
        return jsonify({'ok': True, 'total': float(total), 'revision_id': rev.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/revision/finish', methods=['POST'])
@login_required_user
def revision_finish():
    """Завершить всю ревизию ресторана одним действием: переводим в pending
    ВСЕ открытые (in_progress) ревизии компании, в которых есть хотя бы одна
    посчитанная позиция. Локации без подсчётов не трогаем. Так оператор одной
    кнопкой отправляет на проверку все посчитанные зоны, а админ получает
    единый запрос «Ревизия ресторана»."""
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'error': 'no org'}), 400

    open_revs = Revision.query.filter_by(org_id=org.id, status='in_progress').all()
    if not open_revs:
        return jsonify({'ok': False, 'error': 'Нет активной ревизии'}), 400

    # Считаем позиции пачкой, чтобы не делать N запросов
    from sqlalchemy import func as _sa_f
    _ids = [r.id for r in open_revs]
    _counts = dict(
        db.session.query(RevisionItem.revision_id, _sa_f.count(RevisionItem.id))
        .filter(RevisionItem.revision_id.in_(_ids))
        .group_by(RevisionItem.revision_id)
        .all()
    )

    to_finish = [r for r in open_revs if _counts.get(r.id, 0) > 0]
    if not to_finish:
        return jsonify({'ok': False, 'error': 'Ревизия пуста — посчитайте хотя бы одну позицию'}), 400

    try:
        now = datetime.utcnow()
        for rev in to_finish:
            rev.status = 'pending'
            rev.finished_at = now
        db.session.commit()
        return jsonify({'ok': True, 'locations_finished': len(to_finish)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/revision/edit_item/<int:item_id>', methods=['POST'])
@login_required_user
def revision_edit_item(item_id):
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'error': 'no org'}), 400
    item = RevisionItem.query.get(item_id)
    if not item:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    rev = db.session.get(Revision, item.revision_id)
    if not rev or rev.org_id != org.id:
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    if rev.status in ('completed', 'cancelled'):
        return jsonify({'ok': False, 'error': 'Ревизия уже завершена — править её позиции нельзя.'}), 400
    is_admin = (current_user.user and current_user.user.role == 'admin')
    if not is_admin and item.added_by_user_id != current_user.raw_id:
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    try:
        raw = request.form.get('quantity') or request.form.get('count') or '0'
        new_qty = float(str(raw).replace(',', '.'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'bad quantity'}), 400
    if new_qty <= 0:
        return jsonify({'ok': False, 'error': 'quantity must be > 0'}), 400

    try:
        item.quantity = round(new_qty, 3)
        db.session.commit()
        total = db.session.query(db.func.coalesce(db.func.sum(RevisionItem.quantity), 0)).filter_by(
            revision_id=item.revision_id, product_id=item.product_id, location_id=item.location_id,
        ).scalar() or 0
        return jsonify({'ok': True, 'total': float(total)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/revision/delete_item/<int:item_id>', methods=['POST'])
@login_required_user
def revision_delete_item(item_id):
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'error': 'no org'}), 400
    item = RevisionItem.query.get(item_id)
    if not item:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    # Проверка, что запись принадлежит нашей организации
    rev = db.session.get(Revision, item.revision_id)
    if not rev or rev.org_id != org.id:
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    if rev.status in ('completed', 'cancelled'):
        return jsonify({'ok': False, 'error': 'Ревизия уже завершена — удалять её позиции нельзя.'}), 400
    # Только свою запись — или админ
    is_admin = (current_user.user and current_user.user.role == 'admin')
    if not is_admin and item.added_by_user_id != current_user.raw_id:
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    product_id = item.product_id
    location_id = item.location_id
    revision_id = item.revision_id
    try:
        db.session.delete(item)
        db.session.commit()
        total = db.session.query(db.func.coalesce(db.func.sum(RevisionItem.quantity), 0)).filter_by(
            revision_id=revision_id, product_id=product_id, location_id=location_id,
        ).scalar() or 0
        return jsonify({'ok': True, 'total': float(total)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/revision/product_history/<int:product_id>')
@login_required_user
def revision_product_history(product_id):
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'items': []}), 400
    try:
        location_id = int(request.args.get('location_id') or 0)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'items': []}), 400

    prod = Product.query.filter_by(id=product_id, org_id=org.id).first()
    loc = Location.query.filter_by(id=location_id, org_id=org.id).first()
    if not prod or not loc:
        return jsonify({'ok': True, 'items': [], 'total': 0})

    # Берём как in_progress, так и pending — чтобы оператор и админ видели
    # историю позиции, пока ревизия ждёт подтверждения. Самую свежую
    # (по created_at) на случай, если их несколько.
    rev = Revision.query.filter(
        Revision.org_id == org.id,
        Revision.location_id == loc.id,
        Revision.status.in_(['in_progress', 'pending']),
    ).order_by(Revision.created_at.desc()).first()
    if not rev:
        return jsonify({'ok': True, 'items': [], 'total': 0})

    items = (
        RevisionItem.query.filter_by(
            revision_id=rev.id, product_id=prod.id, location_id=loc.id
        ).order_by(RevisionItem.added_at.asc()).all()
    )
    user_ids = {i.added_by_user_id for i in items if i.added_by_user_id}
    users = {}
    if user_ids:
        for u in User.query.filter(User.id.in_(user_ids)).all():
            users[u.id] = u.username

    current_uid = current_user.raw_id
    is_admin = (current_user.user and current_user.user.role == 'admin')

    out = []
    total = 0.0
    for it in items:
        uname = users.get(it.added_by_user_id, '—')
        can_delete = is_admin or (it.added_by_user_id == current_uid)
        ts = it.added_at.strftime('%d.%m %H:%M') if it.added_at else ''
        qty_rounded = round(it.quantity or 0, 3)
        qty_str = f'{qty_rounded:g}'.replace('.', ',')
        out.append({
            'id': it.id,
            'quantity': qty_rounded,
            'user': uname,
            'timestamp': ts,
            'can_delete': can_delete,
            'text': f'{ts}: {uname} добавил {qty_str}',
        })
        total += qty_rounded
    return jsonify({'ok': True, 'items': out, 'total': round(total, 3)})


# ============== АДМИН: РЕВИЗИИ (ЗАПРОСЫ/ИСТОРИЯ) ==============
def _build_revision_xlsx(revs):
    """Сгенерировать xlsx с итогами ревизии.

    `revs` — одна ревизия ИЛИ список ревизий. Если список — формируется общий
    отчёт по ресторану: остаток каждого товара суммируется по всем локациям,
    список товаров — объединение ассортиментов этих локаций.
    Если у компании есть кастомный шаблон (excel_template) — заполняем его,
    иначе генерируем стандартный отчёт.
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not isinstance(revs, (list, tuple)):
        revs = [revs]
    revs = [r for r in revs if r]
    primary = revs[0]

    org = db.session.get(Organization, primary.org_id)
    user = db.session.get(User, primary.user_id) if primary.user_id else None

    loc_ids = [r.location_id for r in revs if r.location_id]
    uniq_loc_ids = list(dict.fromkeys(loc_ids))
    locs = {l.id: l for l in Location.query.filter(Location.id.in_(uniq_loc_ids)).all()} if uniq_loc_ids else {}
    if len(uniq_loc_ids) > 1:
        loc_label = 'Все локации'
    elif uniq_loc_ids:
        _l = locs.get(uniq_loc_ids[0])
        loc_label = _l.name if _l else '—'
    else:
        loc_label = '—'

    rev_ids = [r.id for r in revs]
    items = RevisionItem.query.filter(RevisionItem.revision_id.in_(rev_ids)).all()

    # Суммируем фактическое количество по товарам — по всем переданным ревизиям
    # (т.е. по всем локациям сразу, если их несколько).
    qty_by_pid = {}
    for it in items:
        qty_by_pid[it.product_id] = qty_by_pid.get(it.product_id, 0) + (it.quantity or 0)

    # Список товаров отчёта — объединение ассортиментов всех локаций ревизий.
    # Fallback на фактически посчитанные, если ассортимент нигде не настроен.
    report_pids = set()
    if uniq_loc_ids:
        report_pids = {lp.product_id for lp in LocationProduct.query.filter(
            LocationProduct.location_id.in_(uniq_loc_ids)
        ).all()}
    if not report_pids:
        report_pids = set(qty_by_pid.keys())
    products = {p.id: p for p in Product.query.filter(
        Product.id.in_(report_pids), Product.is_active == True
    ).all()} if report_pids else {}

    # Если у компании есть кастомный Excel-шаблон — заполняем его
    if org and org.excel_template:
        # Два индекса: по коду и по имени (lower). По коду — основной матч;
        # по имени — фолбэк для импортных позиций без кода (полуфабрикаты).
        # Если в ассортименте оказались два товара с одинаковым кодом / именем
        # (теоретически возможно — нет уникального ограничения в БД),
        # суммируем количества, а не теряем второе.
        agg_by_code = {}
        agg_by_name = {}
        for pid, p in products.items():
            qty = qty_by_pid.get(pid, 0)
            if p.code:
                agg_by_code[p.code] = agg_by_code.get(p.code, 0) + qty
            if p.name:
                key = p.name.strip().lower()
                agg_by_name[key] = agg_by_name.get(key, 0) + qty

        wb = openpyxl.load_workbook(BytesIO(org.excel_template))
        ws = wb.active

        today_str = _fmt_dt(primary.finished_at or datetime.utcnow(), org, fmt='%d.%m.%Y')

        for row in range(1, ws.max_row + 1):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if isinstance(val, str):
                    if 'Дата печати' in val:
                        cell.value = f'Дата печати: {today_str}'
                    elif 'НА ДАТУ:' in val.upper() or val.startswith('На дату'):
                        cell.value = f'На дату: {today_str}'

            # Сначала пытаемся сопоставить по коду из колонки B (2). Если кода нет
            # или он не находится — фолбэк на имя товара из колонки C (3). Это нужно
            # для полуфабрикатов и любых других позиций, которые в шаблоне без кода.
            qty = None
            code_cell = ws.cell(row=row, column=2).value
            if code_cell is not None:
                code_str = str(code_cell).strip()
                if code_str in agg_by_code:
                    qty = agg_by_code[code_str]
            if qty is None:
                name_cell = ws.cell(row=row, column=3).value
                if name_cell is not None:
                    name_str = str(name_cell).strip().lower()
                    if name_str in agg_by_name:
                        qty = agg_by_name[name_str]
            # Пишем остаток только если реально считали (qty > 0). Непосчитанные
            # позиции оставляем пустыми — как в исходном iiko-бланке, где пустая
            # ячейка означает «не считали». Нули в отчёте больше не появляются.
            if qty:
                try:
                    ws.cell(row=row, column=7, value=qty)
                except AttributeError:
                    # Объединённая ячейка в этом месте — пропускаем тихо
                    pass

        # Постобработка бланка:
        #  • убираем строки товаров не из ассортимента локации;
        #  • убираем целые категории, в которых ни один товар не посчитан;
        #  • снимаем фоновую заливку с заголовков категорий.
        allowed_codes = {p.code for p in products.values() if p.code}
        allowed_names = {p.name.strip().lower() for p in products.values() if p.name}
        _meta_prefixes = (
            'бланк', 'на дату', 'дата печати', 'подразделения', 'товарные группы',
            'валюта', 'способ группировки', 'ответственное', 'организация',
        )

        def _cell_s(row, col):
            v = ws.cell(row=row, column=col).value
            return str(v).strip() if v is not None else ''

        def _is_category_row(row):
            b, c, f = _cell_s(row, 2), _cell_s(row, 3), _cell_s(row, 6)
            if not b or c or f:
                return False
            bl = b.lower()
            if bl in ('товар', 'код'):
                return False
            return not any(bl.startswith(p) for p in _meta_prefixes)

        def _is_product_row(row):
            return bool(_cell_s(row, 3) and _cell_s(row, 6))

        def _row_qty(row):
            code_s = _cell_s(row, 2)
            name_s = _cell_s(row, 3).lower()
            return agg_by_code.get(code_s) or agg_by_name.get(name_s) or 0

        def _in_assortment(row):
            return _cell_s(row, 2) in allowed_codes or _cell_s(row, 3).lower() in allowed_names

        max_r = ws.max_row
        cat_rows = [r for r in range(1, max_r + 1) if _is_category_row(r)]
        rows_to_delete = set()
        for idx, crow in enumerate(cat_rows):
            next_cat = cat_rows[idx + 1] if idx + 1 < len(cat_rows) else max_r + 1
            block_products = [r for r in range(crow + 1, next_cat) if _is_product_row(r)]
            counted_any = False
            for pr in block_products:
                if not _in_assortment(pr):
                    rows_to_delete.add(pr)
                elif _row_qty(pr) > 0:
                    counted_any = True
            # Категория без единого посчитанного товара — убираем целиком.
            if not counted_any:
                rows_to_delete.add(crow)
                for pr in block_products:
                    rows_to_delete.add(pr)

        # Снимаем заливку с оставшихся заголовков категорий (бежевый фон).
        _no_fill = PatternFill(fill_type=None)
        for crow in cat_rows:
            if crow not in rows_to_delete:
                for col in range(1, 9):
                    ws.cell(row=crow, column=col).fill = _no_fill

        for row in sorted(rows_to_delete, reverse=True):
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row == row and rng.max_row == row:
                    ws.unmerge_cells(str(rng))
            ws.delete_rows(row, 1)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # Группировка идёт по всему ассортименту локации (а не только по
    # фактически посчитанным позициям). Не посчитанные товары попадают в
    # отчёт с нулём — оператору видно, что позиция была, но осталась пустой.
    cat_ids = {p.category_id for p in products.values() if p.category_id}
    cats = {c.id: c for c in Category.query.filter(Category.id.in_(cat_ids)).all()} if cat_ids else {}

    grouped = {}
    for pid, p in products.items():
        c = cats.get(p.category_id) if p.category_id else None
        cat_name = c.name if c else 'Без категории'
        grouped.setdefault(cat_name, {})
        grouped[cat_name][pid] = {
            'code': p.code or '', 'name': p.name, 'unit': p.unit,
            'qty': qty_by_pid.get(pid, 0),
        }

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ревизия'

    bold = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='7C6CF0')
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Метаинформация
    ws['A1'] = 'Бланк инвентаризации'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    date_str = _fmt_dt(primary.finished_at or primary.created_at or datetime.utcnow(), org)
    ws['A3'] = 'Дата:'
    ws['B3'] = date_str
    ws['A4'] = 'Компания:'
    ws['B4'] = org.name if org else ''
    ws['A5'] = 'Локация:'
    ws['B5'] = loc_label
    ws['A6'] = 'Оператор:'
    ws['B6'] = user.username if user else '—'

    header_row = 8
    headers = ['Код', 'Наименование', 'Ед. изм.', 'Категория', 'Остаток фактический']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = border

    r = header_row + 1
    for cat_name in sorted(grouped.keys()):
        cat_items = grouped[cat_name]
        # Категории без единого посчитанного товара в отчёт не выводим.
        if not any(i['qty'] for i in cat_items.values()):
            continue
        # Строка-заголовок категории (без фоновой заливки)
        cc = ws.cell(row=r, column=1, value=cat_name)
        cc.font = bold
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for _pid, info in sorted(cat_items.items(), key=lambda kv: kv[1]['name']):
            ws.cell(row=r, column=1, value=info['code']).border = border
            ws.cell(row=r, column=2, value=info['name']).border = border
            ws.cell(row=r, column=3, value=info['unit']).border = border
            ws.cell(row=r, column=4, value=cat_name).border = border
            # Непосчитанные позиции — пустая ячейка, а не 0 (нули убираем из отчёта).
            ws.cell(row=r, column=5, value=(info['qty'] if info['qty'] else None)).border = border
            r += 1

    # Автоширина — упрощённо. Используем get_column_letter напрямую, потому что
    # ws.cell(row=1, col) для col∈2..5 вернёт MergedCell (A1:E1 объединена),
    # а у MergedCell нет атрибута column_letter и валится с AttributeError.
    widths = {1: 14, 2: 48, 3: 12, 4: 22, 5: 22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route('/admin/revisions/confirm/<int:rev_id>', methods=['POST'])
@admin_required
def admin_revision_confirm(rev_id):
    org = _current_org()
    rev = Revision.query.filter_by(id=rev_id, org_id=org.id).first()
    if not rev:
        flash('Ревизия не найдена.', 'error')
        return redirect('/admin#tab-requests')
    if rev.status != 'pending':
        flash('Ревизия не в статусе ожидания.', 'error')
        return redirect('/admin#tab-requests')
    try:
        # Подтверждение закрывает ревизию всего ресторана: берём ВСЕ ожидающие
        # (pending) ревизии компании, формируем ОДИН общий отчёт с суммированием
        # остатков по всем локациям и помечаем их все completed. Так админ одним
        # нажатием получает единый бланк по ресторану, а не файл на каждую точку.
        pending_revs = Revision.query.filter_by(
            org_id=org.id, status='pending'
        ).all()
        # На всякий случай гарантируем, что нажатая ревизия в наборе.
        if rev.id not in {r.id for r in pending_revs}:
            pending_revs.append(rev)

        # Схлопываем дубли по локации (наследие гонок): если на одной локации
        # несколько открытых ревизий — сливаем их позиции в одну, чтобы ничьи
        # подсчёты не потерялись и в отчёте не было двойного счёта.
        by_loc = {}
        for r in pending_revs:
            by_loc.setdefault(r.location_id, []).append(r)
        confirmed = []
        for _loc_id, _group in by_loc.items():
            _keep = _group[0]
            for _extra in _group[1:]:
                RevisionItem.query.filter_by(revision_id=_extra.id).update(
                    {'revision_id': _keep.id}, synchronize_session=False
                )
                _extra.status = 'cancelled'
                _extra.finished_at = datetime.utcnow()
            confirmed.append(_keep)
        db.session.flush()

        buf = _build_revision_xlsx(confirmed)
        now = datetime.utcnow()
        for _r in confirmed:
            _r.status = 'completed'
            if not _r.finished_at:
                _r.finished_at = now
        db.session.commit()

        date_str = now.strftime('%Y%m%d_%H%M')
        org_slug = (org.name or 'revision').replace(' ', '_').replace('"', '')
        fname = f'revision_{org_slug}_{date_str}.xlsx'
        return send_file(
            buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
        return redirect('/admin#tab-requests')


@app.route('/admin/revisions/reject/<int:rev_id>', methods=['POST'])
@admin_required
def admin_revision_reject(rev_id):
    org = _current_org()
    rev = Revision.query.filter_by(id=rev_id, org_id=org.id).first()
    if not rev:
        flash('Ревизия не найдена.', 'error')
        return redirect('/admin#tab-requests')
    if rev.status != 'pending':
        flash('Ревизия не в статусе ожидания.', 'error')
        return redirect('/admin#tab-requests')
    try:
        rev.status = 'in_progress'
        rev.finished_at = None
        db.session.commit()
        flash('Ревизия возвращена на пересчёт. Оператор может продолжить подсчёт.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-requests')


@app.route('/admin/settings/timezone', methods=['POST'])
@admin_required
def admin_set_timezone():
    org = _current_org()
    tz = (request.form.get('timezone') or '').strip()
    if tz not in _TZ_VALID:
        flash('Неизвестный часовой пояс.', 'error')
        return redirect('/admin#tab-locations')
    try:
        org.timezone = tz
        db.session.commit()
        flash('Часовой пояс сохранён. Время ревизий теперь показывается в нём.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-locations')


@app.route('/admin/revisions/reject_all', methods=['POST'])
@admin_required
def admin_revision_reject_all():
    """Вернуть на пересчёт ВСЮ ревизию ресторана — все ожидающие локации сразу.
    Запросы схлопнуты в один, поэтому и возврат единый."""
    org = _current_org()
    try:
        pending = Revision.query.filter_by(org_id=org.id, status='pending').all()
        for _r in pending:
            _r.status = 'in_progress'
            _r.finished_at = None
        db.session.commit()
        if pending:
            flash(f'Ревизия возвращена на пересчёт ({len(pending)} локаций). Операторы могут продолжить.', 'success')
        else:
            flash('Нет ожидающих ревизий.', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/admin#tab-requests')


@app.route('/admin/revisions/<int:rev_id>/download')
@admin_required
def admin_revision_download(rev_id):
    org = _current_org()
    rev = Revision.query.filter_by(id=rev_id, org_id=org.id).first()
    if not rev:
        abort(404)
    buf = _build_revision_xlsx(rev)
    loc = db.session.get(Location, rev.location_id) if rev.location_id else None
    loc_name = (loc.name if loc else 'location').replace(' ', '_')
    date_str = (rev.finished_at or rev.created_at or datetime.utcnow()).strftime('%Y%m%d_%H%M')
    fname = f'revision_{loc_name}_{date_str}.xlsx'
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/admin/revisions/<int:rev_id>/download_combined')
@admin_required
def admin_revision_download_combined(rev_id):
    """Скачать общий отчёт по ревизии ресторана из истории: все локации,
    подтверждённые в той же волне (тот же finished_at), суммируются в один
    бланк — так же, как при первичном подтверждении."""
    org = _current_org()
    rev = Revision.query.filter_by(id=rev_id, org_id=org.id, status='completed').first()
    if not rev:
        abort(404)
    if rev.finished_at:
        group = Revision.query.filter_by(
            org_id=org.id, status='completed', finished_at=rev.finished_at
        ).all()
    else:
        group = [rev]
    buf = _build_revision_xlsx(group)
    date_str = (rev.finished_at or datetime.utcnow()).strftime('%Y%m%d_%H%M')
    org_slug = (org.name or 'revision').replace(' ', '_').replace('"', '')
    fname = f'revision_{org_slug}_{date_str}.xlsx'
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ============== CLI КОМАНДЫ ==============
@app.cli.command('seed-demo')
def seed_demo():
    """Создать демо-организацию."""
    email = 'demo@revisi.ru'
    if Organization.query.filter_by(owner_email=email).first():
        click.echo('Демо-организация уже существует.')
        return

    org = Organization(
        name='Демо Бургерная',
        owner_email=email,
        email_verified=True,
        plan='trial',
        trial_ends_at=datetime.utcnow() + timedelta(days=30),
    )
    db.session.add(org)
    db.session.flush()

    admin = User(org_id=org.id, username='admin', email=email, role='admin')
    admin.set_password('demo123')
    db.session.add(admin)

    for order, loc_name in enumerate(('Склад', 'Кухня', 'Островок')):
        db.session.add(Location(org_id=org.id, name=loc_name, sort_order=order))

    db.session.commit()
    click.echo('✔ Демо-организация создана.')
    click.echo(f'  Email:    {email}')
    click.echo('  Username: admin')
    click.echo('  Password: demo123')


@app.cli.command('create-owner')
def create_owner():
    """Создать супер-админа (владельца системы)."""
    email = click.prompt('Email').strip().lower()
    password = click.prompt('Password', hide_input=True, confirmation_prompt=True)

    if OwnerUser.query.filter_by(email=email).first():
        click.echo('Владелец с таким email уже существует.')
        return

    owner = OwnerUser(email=email)
    owner.set_password(password)
    db.session.add(owner)
    db.session.commit()
    click.echo(f'✔ OwnerUser создан: {email}')


# ============== ПОДДЕРЖКА (КЛИЕНТСКАЯ СТОРОНА) ==============
@app.route('/support', methods=['GET', 'POST'])
@login_required_user
def support_page():
    org = _current_org()
    if not org:
        return redirect('/login')
    user = current_user.user
    if request.method == 'POST':
        is_xhr = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        subject = (request.form.get('subject') or '').strip()
        body = (request.form.get('body') or '').strip()
        if not subject or not body:
            if is_xhr:
                return jsonify({'ok': False, 'error': 'Заполните тему и текст.'}), 400
            flash('Заполните тему и текст.', 'error')
            return redirect('/support')
        t = SupportTicket(
            org_id=org.id, user_id=user.id if user else None,
            subject=subject[:300], body=body, status='open',
        )
        db.session.add(t)
        db.session.commit()
        if is_xhr:
            return jsonify({'ok': True, 'ticket_id': t.id})
        flash('Тикет создан. Мы ответим в ближайшее время.', 'success')
        return redirect(f'/support/{t.id}')

    tickets = SupportTicket.query.filter_by(org_id=org.id).order_by(
        SupportTicket.updated_at.desc()
    ).limit(50).all()
    items = []
    for t in tickets:
        items.append({
            'id': t.id, 'subject': t.subject, 'status': t.status,
            'when': t.updated_at.strftime('%d.%m.%Y %H:%M') if t.updated_at else '—',
        })
    return render_template_string(support_list_html, org=org, items=items)


@app.route('/support/<int:ticket_id>', methods=['GET', 'POST'])
@login_required_user
def support_view(ticket_id):
    org = _current_org()
    if not org:
        return redirect('/login')
    t = db.session.get(SupportTicket, ticket_id)
    if not t or t.org_id != org.id:
        flash('Тикет не найден.', 'error')
        return redirect('/support')
    if request.method == 'POST':
        body = (request.form.get('body') or '').strip()
        if body and t.status != 'closed':
            user = current_user.user
            r = SupportTicketReply(
                ticket_id=t.id, author_type='user',
                author_label=user.username if user else 'Клиент',
                body=body,
            )
            db.session.add(r)
            t.status = 'open'
            t.updated_at = datetime.utcnow()
            db.session.commit()
        return redirect(f'/support/{ticket_id}')

    replies = SupportTicketReply.query.filter_by(ticket_id=t.id).order_by(
        SupportTicketReply.created_at
    ).all()
    return render_template_string(support_view_html, org=org, t=t, replies=replies)


# ============== SUPPORT API (для встроенного чат-виджета) ==============
@app.route('/support/api/tickets', methods=['GET'])
@login_required_user
def support_api_list():
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    tickets = SupportTicket.query.filter_by(org_id=org.id).order_by(
        SupportTicket.updated_at.desc()
    ).limit(20).all()
    return jsonify({'ok': True, 'tickets': [
        {
            'id': t.id, 'subject': t.subject, 'status': t.status,
            'updated_at': t.updated_at.strftime('%d.%m.%Y %H:%M') if t.updated_at else '',
        } for t in tickets
    ]})


@app.route('/support/api/ticket/<int:ticket_id>', methods=['GET'])
@login_required_user
def support_api_view(ticket_id):
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    t = db.session.get(SupportTicket, ticket_id)
    if not t or t.org_id != org.id:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    replies = SupportTicketReply.query.filter_by(ticket_id=t.id).order_by(
        SupportTicketReply.created_at
    ).all()
    messages = [{
        'author_type': 'user',
        'author_label': 'Вы',
        'body': t.body,
        'created_at': t.created_at.strftime('%d.%m.%Y %H:%M') if t.created_at else '',
    }]
    for r in replies:
        messages.append({
            'author_type': r.author_type,
            'author_label': r.author_label or ('Вы' if r.author_type == 'user' else 'Поддержка'),
            'body': r.body,
            'created_at': r.created_at.strftime('%d.%m.%Y %H:%M') if r.created_at else '',
        })
    return jsonify({
        'ok': True,
        'ticket': {
            'id': t.id, 'subject': t.subject, 'status': t.status,
            'messages': messages,
        },
    })


@app.route('/support/api/ticket/<int:ticket_id>/reply', methods=['POST'])
@login_required_user
def support_api_reply(ticket_id):
    org = _current_org()
    if not org:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    t = db.session.get(SupportTicket, ticket_id)
    if not t or t.org_id != org.id:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    if t.status == 'closed':
        return jsonify({'ok': False, 'error': 'Тикет закрыт.'}), 400
    body = (request.values.get('body') or '').strip()
    if not body:
        return jsonify({'ok': False, 'error': 'Введите сообщение.'}), 400
    try:
        user = current_user.user
        r = SupportTicketReply(
            ticket_id=t.id, author_type='user',
            author_label=user.username if user else 'Клиент',
            body=body,
        )
        db.session.add(r)
        t.status = 'open'
        t.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({
            'ok': True,
            'message': {
                'author_type': 'user',
                'author_label': 'Вы',
                'body': r.body,
                'created_at': r.created_at.strftime('%d.%m.%Y %H:%M'),
            },
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


# ============== OWNER PANEL ==============

def owner_required(fn):
    """Decorator: only authenticated OwnerUser may access this route.
    Idle-timeout: после Config.OWNER_SESSION_TIMEOUT_MINUTES без активности — logout.
    """
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or current_user.user_type != 'owner':
            return redirect('/owner/login')

        login_at = session.get('owner_login_at', 0)
        idle_seconds = datetime.utcnow().timestamp() - login_at
        if idle_seconds > Config.OWNER_SESSION_TIMEOUT_MINUTES * 60:
            logout_user()
            session.pop('owner_login_at', None)
            flash('Сессия владельца истекла. Войдите заново.', 'error')
            return redirect('/owner/login')

        # Rolling timeout: каждый запрос продлевает таймер на N минут
        session['owner_login_at'] = datetime.utcnow().timestamp()
        return fn(*args, **kwargs)
    return wrapper


@app.route('/owner/login', methods=['GET', 'POST'])
def owner_login():
    error = None
    if current_user.is_authenticated and current_user.user_type == 'owner':
        return redirect('/owner')

    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        password = request.form.get('password') or ''
        ip = request.remote_addr or 'unknown'

        # Защита от брутфорса — для owner порог жёстче, чем для обычных юзеров
        if count_recent_failed_attempts(ip) >= Config.OWNER_MAX_LOGIN_ATTEMPTS:
            error = (
                f'Слишком много неудачных попыток. '
                f'Попробуйте через {Config.LOGIN_LOCKOUT_MINUTES} минут.'
            )
            return render_template_string(owner_login_html, error=error)

        owner = OwnerUser.query.filter_by(email=email).first()
        if not owner or not owner.check_password(password):
            log_attempt(ip, email, False)
            error = 'Неверный email или пароль.'
        else:
            log_attempt(ip, email, True)
            # Если у owner включена 2FA — не логиним сразу, отправляем на /owner/2fa/verify
            emergency_bypass = (
                os.environ.get('OWNER_2FA_EMERGENCY_DISABLE', '').lower() == 'true'
            )
            if owner.totp_enabled and not emergency_bypass:
                session['_pending_2fa_owner_id'] = owner.id
                session['_pending_2fa_at'] = datetime.utcnow().timestamp()
                return redirect('/owner/2fa/verify')
            if owner.totp_enabled and emergency_bypass:
                log_owner_action(
                    '2fa_emergency_bypass',
                    details=f'ip={ip} email={owner.email}',
                )
            login_user(AuthUser(owner, 'owner'), remember=False)  # без remember — короткая сессия
            session['owner_login_at'] = datetime.utcnow().timestamp()
            return redirect('/owner')

    return render_template_string(owner_login_html, error=error)


@app.route('/owner/logout')
def owner_logout():
    logout_user()
    return redirect('/owner/login')


# ============== 2FA (TOTP) для OWNER ==============

def _generate_recovery_codes(count=10):
    """Генерируем человекочитаемые recovery-коды формата XXXX-XXXX."""
    codes = []
    for _ in range(count):
        raw = secrets.token_hex(4).upper()  # 8 hex
        codes.append(f'{raw[:4]}-{raw[4:]}')
    return codes


def _hash_recovery_codes(codes):
    """Сериализуем recovery-коды как JSON-список захешированных значений."""
    norm = [c.replace('-', '').replace(' ', '').lower() for c in codes]
    return json.dumps([generate_password_hash(c) for c in norm])


def _verify_totp(owner, code):
    """Проверка TOTP-кода с окном ±1 (30 сек) для clock drift."""
    if not owner or not owner.totp_secret:
        return False
    code = (code or '').strip().replace(' ', '')
    if not code.isdigit() or len(code) != 6:
        return False
    import pyotp
    return pyotp.TOTP(owner.totp_secret).verify(code, valid_window=1)


def _verify_recovery_code(owner, code):
    """Проверка recovery-кода. Если совпал — удаляем из списка (одноразовый)."""
    if not owner or not owner.recovery_codes_json:
        return False
    code_norm = (code or '').replace('-', '').replace(' ', '').lower()
    if not code_norm:
        return False
    try:
        hashed_list = json.loads(owner.recovery_codes_json)
    except Exception:
        return False
    for i, h in enumerate(hashed_list):
        if check_password_hash(h, code_norm):
            hashed_list.pop(i)
            owner.recovery_codes_json = json.dumps(hashed_list)
            db.session.commit()
            return True
    return False


def _make_qr_svg(uri):
    """Сгенерировать QR-код как SVG-строку (без зависимости от Pillow)."""
    import qrcode
    import qrcode.image.svg
    factory = qrcode.image.svg.SvgPathImage
    img = qrcode.make(uri, image_factory=factory, box_size=10, border=2)
    buf = BytesIO()
    img.save(buf)
    return buf.getvalue().decode('utf-8')


@app.route('/owner/security')
@owner_required
def owner_security():
    owner = OwnerUser.query.get(current_user.record.id)
    remaining = 0
    if owner.recovery_codes_json:
        try:
            remaining = len(json.loads(owner.recovery_codes_json))
        except Exception:
            remaining = 0
    return render_template_string(
        owner_security_html,
        owner_email=owner.email,
        totp_enabled=owner.totp_enabled,
        remaining_codes=remaining,
        owner_pwa_head=_OWNER_PWA_HEAD,
    )


@app.route('/owner/2fa/setup', methods=['GET', 'POST'])
@owner_required
def owner_2fa_setup():
    owner = OwnerUser.query.get(current_user.record.id)
    if owner.totp_enabled:
        flash('2FA уже включена. Сначала отключите её.', 'error')
        return redirect('/owner/security')

    import pyotp
    secret = session.get('_2fa_setup_secret')
    if not secret:
        secret = pyotp.random_base32()
        session['_2fa_setup_secret'] = secret

    error = None
    if request.method == 'POST':
        code = (request.form.get('code') or '').strip().replace(' ', '')
        if not code.isdigit() or len(code) != 6:
            error = 'Введите 6-значный код из приложения.'
        elif not pyotp.TOTP(secret).verify(code, valid_window=1):
            error = 'Неверный код. Проверьте, что время на телефоне точное.'
        else:
            owner.totp_secret = secret
            owner.totp_enabled = True
            recovery = _generate_recovery_codes(10)
            owner.recovery_codes_json = _hash_recovery_codes(recovery)
            db.session.commit()
            session.pop('_2fa_setup_secret', None)
            log_owner_action('enable_2fa')
            return render_template_string(
                owner_2fa_codes_html,
                codes=recovery,
                owner_pwa_head=_OWNER_PWA_HEAD,
            )

    uri = pyotp.TOTP(secret).provisioning_uri(
        name=owner.email, issuer_name='Revisi Owner'
    )
    qr_svg = _make_qr_svg(uri)
    return render_template_string(
        owner_2fa_setup_html,
        qr_svg=qr_svg,
        secret=secret,
        error=error,
        owner_pwa_head=_OWNER_PWA_HEAD,
    )


@app.route('/owner/2fa/verify', methods=['GET', 'POST'])
def owner_2fa_verify():
    pending_id = session.get('_pending_2fa_owner_id')
    pending_at = session.get('_pending_2fa_at', 0)
    # 5-минутное окно между паролем и 2FA-кодом
    if not pending_id or datetime.utcnow().timestamp() - pending_at > 300:
        session.pop('_pending_2fa_owner_id', None)
        session.pop('_pending_2fa_at', None)
        return redirect('/owner/login')

    owner = OwnerUser.query.get(pending_id)
    if not owner or not owner.totp_enabled:
        session.pop('_pending_2fa_owner_id', None)
        session.pop('_pending_2fa_at', None)
        return redirect('/owner/login')

    error = None
    if request.method == 'POST':
        ip = request.remote_addr or 'unknown'
        if count_recent_failed_attempts(ip) >= Config.OWNER_MAX_LOGIN_ATTEMPTS:
            error = (
                f'Слишком много неудачных попыток. '
                f'Через {Config.LOGIN_LOCKOUT_MINUTES} минут.'
            )
        else:
            code = (request.form.get('code') or '').strip()
            if _verify_totp(owner, code) or _verify_recovery_code(owner, code):
                session.pop('_pending_2fa_owner_id', None)
                session.pop('_pending_2fa_at', None)
                log_attempt(ip, owner.email + ' [2fa]', True)
                login_user(AuthUser(owner, 'owner'), remember=False)
                session['owner_login_at'] = datetime.utcnow().timestamp()
                return redirect('/owner')
            log_attempt(ip, owner.email + ' [2fa]', False)
            error = 'Неверный код'

    return render_template_string(
        owner_2fa_verify_html,
        error=error,
        owner_pwa_head=_OWNER_PWA_HEAD,
    )


@app.route('/owner/2fa/disable', methods=['POST'])
@owner_required
def owner_2fa_disable():
    owner = OwnerUser.query.get(current_user.record.id)
    password = request.form.get('password') or ''
    code = (request.form.get('code') or '').strip()

    if not owner.check_password(password):
        flash('Неверный пароль', 'error')
        return redirect('/owner/security')
    if not (_verify_totp(owner, code) or _verify_recovery_code(owner, code)):
        flash('Неверный код', 'error')
        return redirect('/owner/security')

    owner.totp_secret = None
    owner.totp_enabled = False
    owner.recovery_codes_json = None
    db.session.commit()
    log_owner_action('disable_2fa')
    flash('2FA отключена', 'success')
    return redirect('/owner/security')


@app.route('/owner')
@owner_required
def owner_dashboard():
    now = datetime.utcnow()
    seven_days_ago = now - timedelta(days=7)

    total_orgs = Organization.query.count()

    # Активные: есть пользователь с last_login_at за последние 7 дней
    active_org_ids = db.session.query(User.org_id).filter(
        User.last_login_at >= seven_days_ago
    ).distinct().subquery()
    active_orgs = db.session.query(db.func.count()).select_from(
        Organization
    ).filter(Organization.id.in_(active_org_ids)).scalar() or 0

    # На trial
    trial_orgs = Organization.query.filter(
        Organization.plan == 'trial',
        Organization.trial_ends_at > now,
    ).count()

    # Платящих
    paying_orgs = Organization.query.filter(
        Organization.plan.in_(('pro', 'business')),
        Organization.subscription_ends_at > now,
    ).count()

    # Trial заканчивается в ближайшие 3 дня
    in_3_days = now + timedelta(days=3)
    expiring_soon = Organization.query.filter(
        Organization.plan == 'trial',
        Organization.trial_ends_at > now,
        Organization.trial_ends_at <= in_3_days,
    ).all()

    expiring_list = []
    for org in expiring_soon:
        delta = org.trial_ends_at - now
        days_left = max(0, delta.days)
        expiring_list.append({
            'id': org.id,
            'name': org.name,
            'email': org.owner_email,
            'days_left': days_left,
        })

    # Последние 5 регистраций
    recent_orgs = Organization.query.order_by(Organization.created_at.desc()).limit(5).all()
    recent_list = []
    for org in recent_orgs:
        recent_list.append({
            'name': org.name,
            'email': org.owner_email,
            'plan': org.plan,
            'created_at': org.created_at.strftime('%d.%m.%Y') if org.created_at else '—',
        })

    # Активность сегодня
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    activity_today = RevisionItem.query.filter(RevisionItem.added_at >= today_start).count()

    # Тоталы по системе
    total_users = User.query.count()
    total_products = Product.query.count()
    total_revisions = Revision.query.filter(Revision.status == 'completed').count()

    # MRR — сумма monthly_price для платящих
    mrr_q = db.session.query(db.func.coalesce(db.func.sum(Organization.monthly_price), 0)).filter(
        Organization.plan.in_(('pro', 'business')),
        Organization.subscription_ends_at > now,
    ).scalar() or 0
    mrr = int(mrr_q)

    # Конверсия trial → paid
    finished_trials_total = Organization.query.filter(
        (Organization.plan != 'trial') | ((Organization.trial_ends_at != None) & (Organization.trial_ends_at <= now))
    ).filter(Organization.created_at < now - timedelta(days=14)).count()
    converted = Organization.query.filter(
        Organization.plan.in_(('pro', 'business'))
    ).count()
    conv_pct = int((converted / finished_trials_total) * 100) if finished_trials_total else 0

    # 30-дневный график регистраций
    thirty_days_ago = now - timedelta(days=30)
    signups_q = Organization.query.filter(Organization.created_at >= thirty_days_ago).all()
    signups_per_day = [0] * 30
    for o in signups_q:
        if not o.created_at:
            continue
        days_back = (now.date() - o.created_at.date()).days
        if 0 <= days_back < 30:
            signups_per_day[29 - days_back] += 1

    # 30-дневный график активности (revision items)
    activity_q = db.session.query(RevisionItem.added_at).filter(
        RevisionItem.added_at >= thirty_days_ago
    ).all()
    activity_per_day = [0] * 30
    for (dt,) in activity_q:
        if not dt:
            continue
        days_back = (now.date() - dt.date()).days
        if 0 <= days_back < 30:
            activity_per_day[29 - days_back] += 1

    # Алерты: неактивные >14 дней
    fourteen_days_ago = now - timedelta(days=14)
    inactive_org_ids = set()
    for o in Organization.query.all():
        users = User.query.filter_by(org_id=o.id).all()
        if not users:
            continue
        last_login = max((u.last_login_at for u in users if u.last_login_at), default=None)
        if not last_login or last_login < fourteen_days_ago:
            inactive_org_ids.add(o.id)

    inactive_count = len(inactive_org_ids)

    # Открытых тикетов
    open_tickets = SupportTicket.query.filter_by(status='open').count()

    return render_template_string(
        owner_dashboard_html,
        owner_email=current_user.username,
        total_orgs=total_orgs,
        active_orgs=active_orgs,
        trial_orgs=trial_orgs,
        paying_orgs=paying_orgs,
        expiring_list=expiring_list,
        recent_list=recent_list,
        activity_today=activity_today,
        total_users=total_users,
        total_products=total_products,
        total_revisions=total_revisions,
        mrr=mrr,
        conv_pct=conv_pct,
        signups_per_day=signups_per_day,
        activity_per_day=activity_per_day,
        inactive_count=inactive_count,
        open_tickets=open_tickets,
    )


@app.route('/owner/orgs')
@owner_required
def owner_orgs():
    orgs = Organization.query.order_by(Organization.created_at.desc()).all()
    now = datetime.utcnow()

    org_rows = []
    for org in orgs:
        users = User.query.filter_by(org_id=org.id).all()
        user_count = len(users)
        last_login = None
        for u in users:
            if u.last_login_at:
                if last_login is None or u.last_login_at > last_login:
                    last_login = u.last_login_at

        if org.plan == 'trial' and org.trial_ends_at:
            ends_str = org.trial_ends_at.strftime('%d.%m.%Y')
        elif org.plan in ('pro', 'business') and org.subscription_ends_at:
            ends_str = org.subscription_ends_at.strftime('%d.%m.%Y')
        else:
            ends_str = '—'

        org_rows.append({
            'id': org.id,
            'name': org.name,
            'email': org.owner_email,
            'plan': org.plan,
            'ends': ends_str,
            'user_count': user_count,
            'last_activity': last_login.strftime('%d.%m.%Y %H:%M') if last_login else 'никогда',
            'is_blocked': org.is_blocked,
        })

    return render_template_string(
        owner_orgs_html,
        owner_email=current_user.username,
        org_rows=org_rows,
    )


@app.route('/owner/orgs/<int:org_id>/extend_trial', methods=['POST'])
@owner_required
def owner_extend_trial(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    now = datetime.utcnow()
    if org.trial_ends_at and org.trial_ends_at > now:
        org.trial_ends_at = org.trial_ends_at + timedelta(days=7)
    else:
        org.trial_ends_at = now + timedelta(days=7)
    org.plan = 'trial'
    try:
        db.session.commit()
        flash(f'Trial компании «{org.name}» продлён на 7 дней.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/owner/orgs')


@app.route('/owner/orgs/<int:org_id>/set_plan', methods=['POST'])
@owner_required
def owner_set_plan(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    plan = (request.form.get('plan') or '').lower()
    if plan not in ('free', 'trial', 'pro', 'business'):
        flash('Неизвестный тариф.', 'error')
        return redirect('/owner/orgs')
    org.plan = plan
    if plan in ('pro', 'business'):
        org.subscription_ends_at = datetime.utcnow() + timedelta(days=30)
    try:
        db.session.commit()
        flash(f'Тариф компании «{org.name}» изменён на {plan.upper()}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/owner/orgs')


@app.route('/owner/orgs/<int:org_id>/toggle_block', methods=['POST'])
@owner_required
def owner_toggle_block(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    org.is_blocked = not org.is_blocked
    action = 'заблокирована' if org.is_blocked else 'разблокирована'
    try:
        db.session.commit()
        flash(f'Компания «{org.name}» {action}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect('/owner/orgs')


@app.route('/owner/orgs/<int:org_id>/delete', methods=['POST'])
@owner_required
def owner_delete_org(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    name = org.name
    try:
        log_owner_action('delete_org', target_org_id=org.id, details=f'name={name}, email={org.owner_email}')
        # Audit-лог сохраняем, но обнуляем ссылку на удаляемую org, иначе ForeignKey блокирует delete
        OwnerAuditLog.query.filter_by(target_org_id=org.id).update({'target_org_id': None}, synchronize_session=False)

        # Полностью ручное удаление в строгом порядке «листья → корень».
        # Не полагаемся на ORM-cascade: он не покрывает LocationProduct и не
        # гарантирует порядок, из-за чего FK (revision_items_location_id_fkey,
        # location_products_*, и т.п.) блокируют удаление.
        org_id_val = org.id
        loc_ids = [l.id for l in Location.query.filter_by(org_id=org_id_val).all()]
        prod_ids = [p.id for p in Product.query.filter_by(org_id=org_id_val).all()]
        ticket_ids = [t.id for t in SupportTicket.query.filter_by(org_id=org_id_val).all()]

        D = lambda q: q.delete(synchronize_session=False)
        # тикеты поддержки
        if ticket_ids:
            D(SupportTicketReply.query.filter(SupportTicketReply.ticket_id.in_(ticket_ids)))
        D(SupportTicket.query.filter_by(org_id=org_id_val))
        # позиции ревизий → ревизии (по org покрывает все позиции этих ревизий)
        rev_ids = [r.id for r in Revision.query.filter_by(org_id=org_id_val).all()]
        if rev_ids:
            D(RevisionItem.query.filter(RevisionItem.revision_id.in_(rev_ids)))
        D(Revision.query.filter_by(org_id=org_id_val))
        # ассортимент и нормы (ссылаются на locations и products)
        if loc_ids:
            D(LocationProduct.query.filter(LocationProduct.location_id.in_(loc_ids)))
            D(ProductNorm.query.filter(ProductNorm.location_id.in_(loc_ids)))
        if prod_ids:
            D(LocationProduct.query.filter(LocationProduct.product_id.in_(prod_ids)))
            D(ProductNorm.query.filter(ProductNorm.product_id.in_(prod_ids)))
        # справочники компании
        D(Product.query.filter_by(org_id=org_id_val))
        D(Category.query.filter_by(org_id=org_id_val))
        D(Location.query.filter_by(org_id=org_id_val))
        D(User.query.filter_by(org_id=org_id_val))
        db.session.flush()
        # очищаем identity map, чтобы ORM не пытался каскадить уже удалённое
        db.session.expire_all()
        D(Organization.query.filter_by(id=org_id_val))
        db.session.commit()
        flash(f'Компания «{name}» и все её данные удалены.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка удаления: {e}', 'error')
    return redirect('/owner/orgs')


# ============== OWNER: ДЕТАЛИ КОМПАНИИ ==============
@app.route('/owner/orgs/<int:org_id>')
@owner_required
def owner_org_detail(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')

    now = datetime.utcnow()
    users = User.query.filter_by(org_id=org.id).order_by(User.role.desc(), User.username).all()
    user_rows = []
    for u in users:
        user_rows.append({
            'id': u.id, 'username': u.username, 'email': u.email or '—',
            'role': u.role,
            'last_login': u.last_login_at.strftime('%d.%m.%Y %H:%M') if u.last_login_at else 'никогда',
        })

    locations_count = Location.query.filter_by(org_id=org.id).count()
    products_count = Product.query.filter_by(org_id=org.id).count()
    revisions_count = Revision.query.filter_by(org_id=org.id).count()

    seven_days_ago = now - timedelta(days=7)
    items_7d = db.session.query(db.func.count(RevisionItem.id)).join(Revision).filter(
        Revision.org_id == org.id, RevisionItem.added_at >= seven_days_ago
    ).scalar() or 0

    # Sparkline активности 14 дней
    fourteen_days_ago = now - timedelta(days=14)
    spark_q = db.session.query(RevisionItem.added_at).join(Revision).filter(
        Revision.org_id == org.id, RevisionItem.added_at >= fourteen_days_ago
    ).all()
    spark = [0] * 14
    for (dt,) in spark_q:
        if dt:
            d = (now.date() - dt.date()).days
            if 0 <= d < 14:
                spark[13 - d] += 1

    if org.plan == 'trial' and org.trial_ends_at:
        ends_str = org.trial_ends_at.strftime('%d.%m.%Y')
    elif org.plan in ('pro', 'business') and org.subscription_ends_at:
        ends_str = org.subscription_ends_at.strftime('%d.%m.%Y')
    else:
        ends_str = '—'

    features = org.features or {}
    has_template = bool(org.excel_template)
    template_size = len(org.excel_template) if org.excel_template else 0
    return render_template_string(
        owner_org_detail_html,
        owner_email=current_user.username,
        org=org, ends_str=ends_str,
        user_rows=user_rows,
        locations_count=locations_count,
        products_count=products_count,
        revisions_count=revisions_count,
        items_7d=items_7d, spark=spark,
        features=features,
        has_template=has_template, template_size=template_size,
    )


@app.route('/owner/orgs/<int:org_id>/features', methods=['POST'])
@owner_required
def owner_set_features(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    feats = {
        'excel_export': request.form.get('excel_export') == 'on',
        'multi_location': request.form.get('multi_location') == 'on',
        'history': request.form.get('history') == 'on',
        'unlimited_users': request.form.get('unlimited_users') == 'on',
    }
    org.features = feats
    try:
        log_owner_action('set_features', target_org_id=org.id, details=str(feats))
        db.session.commit()
        flash('Feature-флаги обновлены.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect(f'/owner/orgs/{org_id}')


@app.route('/owner/orgs/<int:org_id>/price', methods=['POST'])
@owner_required
def owner_set_price(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    try:
        price = int(request.form.get('monthly_price') or 0)
    except (TypeError, ValueError):
        price = 0
    org.monthly_price = max(0, price)
    try:
        log_owner_action('set_price', target_org_id=org.id, details=f'price={price}')
        db.session.commit()
        flash(f'Цена подписки: {price}₽/мес.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect(f'/owner/orgs/{org_id}')


# ============== OWNER: ИМПЕРСONATION ==============
@app.route('/owner/orgs/<int:org_id>/impersonate', methods=['POST'])
@owner_required
def owner_impersonate(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    admin_user = User.query.filter_by(org_id=org.id, role='admin').first()
    if not admin_user:
        flash('У компании нет админа — нельзя войти.', 'error')
        return redirect(f'/owner/orgs/{org_id}')

    log_owner_action(
        'impersonate_start', target_org_id=org.id, target_user_id=admin_user.id,
        details=f'as={admin_user.username}'
    )
    owner_id_remember = current_user.raw_id
    logout_user()
    session['impersonating_owner_id'] = owner_id_remember
    session['impersonating_org_id'] = org.id
    login_user(AuthUser(admin_user, 'user'), remember=True)
    flash(f'Вы вошли как admin компании «{org.name}». Нажмите «Вернуться» в шапке, чтобы выйти.', 'success')
    return redirect('/admin')


@app.route('/owner/orgs/<int:org_id>/template/upload', methods=['POST'])
@owner_required
def owner_org_upload_template(org_id):
    """Принять .xlsx-шаблон отчёта по ревизии и привязать его к компании.
    После этого _build_revision_xlsx будет заполнять остатки в этом шаблоне.
    """
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Файл не выбран.', 'error')
        return redirect(f'/owner/orgs/{org_id}')
    if not f.filename.lower().endswith('.xlsx'):
        flash('Нужен файл .xlsx (Excel 2007+). Старый .xls и .xml не подходят — пересохраните в .xlsx.', 'error')
        return redirect(f'/owner/orgs/{org_id}')
    try:
        data = f.read()
        # Проверим, что openpyxl может это прочитать — иначе при ревизии будет крах
        import openpyxl as _opx
        _opx.load_workbook(BytesIO(data))
        org.excel_template = data
        db.session.commit()
        log_owner_action('upload_template', target_org_id=org.id,
                         details=f'size={len(data)}')
        flash(f'Шаблон отчёта обновлён ({len(data)} байт).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Не удалось прочитать файл как .xlsx: {e}', 'error')
    return redirect(f'/owner/orgs/{org_id}')


@app.route('/owner/orgs/<int:org_id>/template/remove', methods=['POST'])
@owner_required
def owner_org_remove_template(org_id):
    org = db.session.get(Organization, org_id)
    if not org:
        flash('Компания не найдена.', 'error')
        return redirect('/owner/orgs')
    org.excel_template = None
    db.session.commit()
    log_owner_action('remove_template', target_org_id=org.id)
    flash('Шаблон отчёта снят. Отчёт будет генериться в стандартном виде.', 'success')
    return redirect(f'/owner/orgs/{org_id}')


@app.route('/owner/stop_impersonate')
def owner_stop_impersonate():
    owner_id = session.pop('impersonating_owner_id', None)
    org_id = session.pop('impersonating_org_id', None)
    if not owner_id:
        return redirect('/')
    owner = db.session.get(OwnerUser, owner_id)
    if not owner:
        return redirect('/owner/login')
    log_owner_action('impersonate_stop', target_org_id=org_id, details=f'owner_id={owner_id}')
    logout_user()
    login_user(AuthUser(owner, 'owner'), remember=True)
    flash('Вы вернулись в панель владельца.', 'success')
    return redirect('/owner')


# ============== OWNER: АУДИТ ==============
@app.route('/owner/audit')
@owner_required
def owner_audit():
    rows = OwnerAuditLog.query.order_by(OwnerAuditLog.created_at.desc()).limit(200).all()
    org_names = {o.id: o.name for o in Organization.query.all()}
    log_rows = []
    for r in rows:
        log_rows.append({
            'when': r.created_at.strftime('%d.%m.%Y %H:%M:%S') if r.created_at else '—',
            'action': r.action,
            'org': org_names.get(r.target_org_id, '—') if r.target_org_id else '—',
            'details': (r.details or '')[:200],
            'ip': r.ip_address or '—',
        })
    return render_template_string(
        owner_audit_html,
        owner_email=current_user.username,
        log_rows=log_rows,
    )


# ============== OWNER: АЛЕРТЫ ==============
@app.route('/owner/alerts')
@owner_required
def owner_alerts():
    now = datetime.utcnow()
    in_3 = now + timedelta(days=3)
    in_7 = now + timedelta(days=7)
    fourteen_ago = now - timedelta(days=14)
    thirty_ago = now - timedelta(days=30)

    expiring_3 = []
    expiring_7 = []
    for org in Organization.query.filter(Organization.plan == 'trial', Organization.trial_ends_at > now).all():
        days = (org.trial_ends_at - now).days
        item = {'id': org.id, 'name': org.name, 'email': org.owner_email, 'days': max(0, days)}
        if org.trial_ends_at <= in_3:
            expiring_3.append(item)
        elif org.trial_ends_at <= in_7:
            expiring_7.append(item)

    inactive = []
    for org in Organization.query.all():
        users = User.query.filter_by(org_id=org.id).all()
        if not users:
            continue
        last = max((u.last_login_at for u in users if u.last_login_at), default=None)
        if not last:
            inactive.append({'id': org.id, 'name': org.name, 'email': org.owner_email, 'last': 'никогда'})
        elif last < thirty_ago:
            inactive.append({'id': org.id, 'name': org.name, 'email': org.owner_email,
                             'last': last.strftime('%d.%m.%Y'), 'churn_risk': True})
        elif last < fourteen_ago:
            inactive.append({'id': org.id, 'name': org.name, 'email': org.owner_email,
                             'last': last.strftime('%d.%m.%Y'), 'churn_risk': False})

    paid_expiring = []
    for org in Organization.query.filter(
        Organization.plan.in_(('pro', 'business')),
        Organization.subscription_ends_at != None,
        Organization.subscription_ends_at > now,
        Organization.subscription_ends_at <= in_7,
    ).all():
        days = (org.subscription_ends_at - now).days
        paid_expiring.append({'id': org.id, 'name': org.name, 'email': org.owner_email, 'days': max(0, days)})

    return render_template_string(
        owner_alerts_html, owner_email=current_user.username,
        expiring_3=expiring_3, expiring_7=expiring_7,
        inactive=inactive, paid_expiring=paid_expiring,
    )


# ============== OWNER: РАССЫЛКА ==============
@app.route('/owner/broadcast', methods=['GET', 'POST'])
@owner_required
def owner_broadcast():
    sent = None
    if request.method == 'POST':
        subject = (request.form.get('subject') or '').strip()
        body = (request.form.get('body') or '').strip()
        audience = request.form.get('audience') or 'all'
        if not subject or not body:
            flash('Тема и текст обязательны.', 'error')
            return redirect('/owner/broadcast')

        now = datetime.utcnow()
        q = Organization.query
        if audience == 'trial':
            q = q.filter(Organization.plan == 'trial', Organization.trial_ends_at > now)
        elif audience == 'paid':
            q = q.filter(Organization.plan.in_(('pro', 'business')))
        elif audience == 'inactive14':
            fourteen_ago = now - timedelta(days=14)
            inactive_ids = []
            for org in Organization.query.all():
                last = db.session.query(db.func.max(User.last_login_at)).filter(User.org_id == org.id).scalar()
                if not last or last < fourteen_ago:
                    inactive_ids.append(org.id)
            q = q.filter(Organization.id.in_(inactive_ids))
        recipients = q.all()

        log_owner_action('broadcast', details=f'audience={audience} count={len(recipients)} subject={subject[:80]}')
        sent = len(recipients)
        flash(f'Сообщение поставлено в очередь для {sent} компаний. (Реальная отправка email подключается отдельно.)', 'success')
        return redirect('/owner/broadcast')
    return render_template_string(owner_broadcast_html, owner_email=current_user.username)


# ============== OWNER: BACKUP ==============
@app.route('/owner/backup')
@owner_required
def owner_backup():
    return render_template_string(owner_backup_html, owner_email=current_user.username)


@app.route('/owner/backup/download')
@owner_required
def owner_backup_download():
    import json, zipfile
    log_owner_action('backup_download')
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for model_cls, fname in [
            (Organization, 'organizations.json'), (User, 'users.json'),
            (Location, 'locations.json'), (Category, 'categories.json'),
            (Product, 'products.json'), (ProductNorm, 'product_norms.json'),
            (Revision, 'revisions.json'), (RevisionItem, 'revision_items.json'),
            (CatalogProduct, 'catalog_products.json'),
            (SupportTicket, 'support_tickets.json'),
            (OwnerAuditLog, 'owner_audit_log.json'),
        ]:
            rows = []
            for r in model_cls.query.all():
                d = {}
                for col in r.__table__.columns:
                    val = getattr(r, col.name)
                    if isinstance(val, datetime):
                        val = val.isoformat()
                    elif isinstance(val, (bytes, memoryview)):
                        val = f'<binary {len(bytes(val))} bytes>'
                    d[col.name] = val
                rows.append(d)
            zf.writestr(fname, json.dumps(rows, ensure_ascii=False, indent=2, default=str))
    buf.seek(0)
    fname = f'revisi_backup_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.zip'
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/zip')


# ============== OWNER: ТИКЕТЫ ==============
@app.route('/owner/tickets')
@owner_required
def owner_tickets():
    rows = SupportTicket.query.order_by(
        SupportTicket.status.asc(), SupportTicket.updated_at.desc()
    ).limit(200).all()
    org_names = {o.id: o.name for o in Organization.query.all()}
    items = []
    for t in rows:
        items.append({
            'id': t.id, 'subject': t.subject, 'status': t.status,
            'org_name': org_names.get(t.org_id, '—'),
            'when': t.updated_at.strftime('%d.%m.%Y %H:%M') if t.updated_at else '—',
        })
    return render_template_string(owner_tickets_html, owner_email=current_user.username, items=items)


@app.route('/owner/tickets/<int:ticket_id>', methods=['GET', 'POST'])
@owner_required
def owner_ticket_view(ticket_id):
    t = db.session.get(SupportTicket, ticket_id)
    if not t:
        flash('Тикет не найден.', 'error')
        return redirect('/owner/tickets')
    if request.method == 'POST':
        body = (request.form.get('body') or '').strip()
        action = request.form.get('action') or 'reply'
        if body:
            r = SupportTicketReply(
                ticket_id=t.id, author_type='owner',
                author_label='Владелец', body=body,
            )
            db.session.add(r)
            t.status = 'answered'
            t.updated_at = datetime.utcnow()
            log_owner_action('ticket_reply', target_org_id=t.org_id, details=f'ticket={t.id}')
            db.session.commit()
        if action == 'close':
            t.status = 'closed'
            t.updated_at = datetime.utcnow()
            log_owner_action('ticket_close', target_org_id=t.org_id, details=f'ticket={t.id}')
            db.session.commit()
        return redirect(f'/owner/tickets/{ticket_id}')

    org = db.session.get(Organization, t.org_id)
    replies = SupportTicketReply.query.filter_by(ticket_id=t.id).order_by(SupportTicketReply.created_at).all()
    return render_template_string(
        owner_ticket_view_html, owner_email=current_user.username,
        t=t, org=org, replies=replies,
    )


# ============== OWNER: КАТАЛОГ ==============
@app.route('/owner/catalog', methods=['GET', 'POST'])
@owner_required
def owner_catalog():
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        unit = (request.form.get('unit') or 'шт').strip()
        category = (request.form.get('category') or '').strip()
        code = (request.form.get('code') or '').strip()
        if name:
            cp = CatalogProduct(name=name, unit=unit, category_name=category or None, code=code or None)
            db.session.add(cp)
            log_owner_action('catalog_add', details=f'name={name}')
            db.session.commit()
            flash(f'Добавлено: {name}', 'success')
        return redirect('/owner/catalog')

    rows = CatalogProduct.query.order_by(CatalogProduct.category_name, CatalogProduct.name).all()
    return render_template_string(owner_catalog_html, owner_email=current_user.username, rows=rows)


@app.route('/owner/catalog/<int:cp_id>/delete', methods=['POST'])
@owner_required
def owner_catalog_delete(cp_id):
    cp = db.session.get(CatalogProduct, cp_id)
    if cp:
        log_owner_action('catalog_delete', details=f'name={cp.name}')
        db.session.delete(cp)
        db.session.commit()
    return redirect('/owner/catalog')


# ============== OWNER: HEALTH ==============
@app.route('/owner/health')
@owner_required
def owner_health():
    import time as _time
    health = {'db_ok': False, 'db_latency_ms': None, 'db_size': '—',
              'orgs': 0, 'users': 0, 'products': 0, 'revisions': 0, 'items': 0,
              'app_started': _APP_STARTED_AT.strftime('%d.%m.%Y %H:%M:%S'),
              'uptime': str(datetime.utcnow() - _APP_STARTED_AT).split('.')[0]}
    try:
        t0 = _time.time()
        db.session.execute(db.text('SELECT 1')).scalar()
        health['db_latency_ms'] = round((_time.time() - t0) * 1000, 1)
        health['db_ok'] = True
    except Exception:
        pass
    try:
        size = db.session.execute(db.text("SELECT pg_size_pretty(pg_database_size(current_database()))")).scalar()
        health['db_size'] = size or '—'
    except Exception:
        pass
    health['orgs'] = Organization.query.count()
    health['users'] = User.query.count()
    health['products'] = Product.query.count()
    health['revisions'] = Revision.query.count()
    health['items'] = RevisionItem.query.count()
    return render_template_string(owner_health_html, owner_email=current_user.username, h=health)


# ============== OWNER: SQL КОНСОЛЬ ==============
@app.route('/owner/sql', methods=['GET', 'POST'])
@owner_required
def owner_sql():
    result_cols, result_rows, error, query, executed = [], [], None, '', False
    if request.method == 'POST':
        query = (request.form.get('query') or '').strip()
        executed = True
        # Защита: разрешаем только SELECT (default). Для DDL/DML — явное подтверждение.
        allow_write = request.form.get('allow_write') == 'on'
        first_word = query.lstrip().split(None, 1)[0].lower() if query else ''
        if not allow_write and first_word and first_word != 'select':
            error = 'Разрешены только SELECT-запросы. Для других включите чекбокс «Разрешить запись».'
        elif query:
            try:
                log_owner_action('sql_exec', details=f'write={allow_write} query={query[:300]}')
                with db.engine.connect() as conn:
                    res = conn.execute(db.text(query))
                    if res.returns_rows:
                        result_cols = list(res.keys())
                        result_rows = [list(r) for r in res.fetchmany(500)]
                    else:
                        if allow_write:
                            conn.commit()
                        result_cols = ['result']
                        result_rows = [[f'OK (rowcount={res.rowcount})']]
            except Exception as e:
                error = str(e)
    return render_template_string(
        owner_sql_html, owner_email=current_user.username,
        query=query, result_cols=result_cols, result_rows=result_rows,
        error=error, executed=executed,
    )


# ============== HTML ШАБЛОНЫ ==============

_BASE_CSS = '''
* { box-sizing: border-box; -webkit-tap-highlight-color: transparent; margin: 0; padding: 0; }
body {
    font-family: 'Outfit', sans-serif;
    background: linear-gradient(135deg, #13111C 0%, #1d1635 50%, #231b50 100%);
    background-attachment: fixed;
    min-height: 100vh;
    color: white;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 20px;
    position: relative;
    overflow-x: hidden;
}
.blob { position: fixed; border-radius: 50%; filter: blur(80px); opacity: 0.35; pointer-events: none; z-index: 0; }
.blob-1 { width: 400px; height: 400px; background: radial-gradient(circle, #7c6cf0, #a855f7); top: -100px; left: -100px; }
.blob-2 { width: 350px; height: 350px; background: radial-gradient(circle, #a855f7, #6d28d9); bottom: -80px; right: -80px; }
.blob-3 { width: 250px; height: 250px; background: radial-gradient(circle, #818cf8, #7c6cf0); top: 50%; left: 50%; transform: translate(-50%, -50%); }
.card {
    background: rgba(255,255,255,0.07);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 24px;
    box-shadow: 0 20px 50px rgba(0,0,0,0.4);
    padding: 40px 32px;
    max-width: 440px;
    width: 100%;
    position: relative;
    z-index: 1;
    animation: fadeIn 0.5s ease-out;
}
@keyframes fadeIn { from { opacity: 0; transform: translateY(24px); } to { opacity: 1; transform: translateY(0); } }
.title { text-align: center; color: white; font-weight: 700; font-size: 26px; margin-bottom: 6px; }
.subtitle { text-align: center; color: rgba(255,255,255,0.5); font-size: 13px; margin-bottom: 28px; }
.form-group { margin-bottom: 16px; }
.form-group label {
    display: block; margin-bottom: 6px; font-weight: 600;
    color: rgba(255,255,255,0.55); font-size: 12px;
    text-transform: uppercase; letter-spacing: 0.8px;
}
.input {
    width: 100%;
    background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 14px;
    color: white;
    padding: 14px 16px;
    font-family: 'Outfit', sans-serif;
    font-size: 15px;
    transition: all 0.2s;
}
.input::placeholder { color: rgba(255,255,255,0.3); }
.input:focus {
    border-color: rgba(124,108,240,0.7);
    background: rgba(124,108,240,0.12);
    outline: none;
    box-shadow: 0 0 0 3px rgba(124,108,240,0.2);
}
select.input option { background: #1d1635; color: white; }
.hint { font-size: 11px; color: rgba(255,255,255,0.4); margin-top: 4px; padding-left: 2px; }
.btn-primary {
    width: 100%;
    background: linear-gradient(135deg, #7c6cf0, #a855f7);
    color: white;
    border: none;
    padding: 15px;
    border-radius: 14px;
    font-weight: 600;
    font-size: 16px;
    font-family: 'Outfit', sans-serif;
    cursor: pointer;
    box-shadow: 0 6px 24px rgba(124,108,240,0.4);
    margin-top: 6px;
    transition: transform 0.1s;
}
.btn-primary:active { transform: scale(0.98); }
.error {
    color: #fca5a5;
    background: rgba(239,68,68,0.12);
    border: 1px solid rgba(239,68,68,0.3);
    padding: 12px 16px;
    border-radius: 12px;
    margin-bottom: 18px;
    text-align: center;
    font-size: 14px;
}
.link { color: #a78bfa; text-decoration: none; font-weight: 600; }
.link:hover { color: #c4b5fd; }
.bottom-link {
    text-align: center;
    margin-top: 20px;
    font-size: 14px;
    color: rgba(255,255,255,0.55);
}
.checkbox-row {
    display: flex; align-items: flex-start; gap: 10px;
    color: rgba(255,255,255,0.7); font-size: 13px;
    margin: 4px 0 14px; cursor: pointer;
}
.checkbox-row input { margin-top: 3px; accent-color: #7c6cf0; }
.icon-box {
    width: 64px; height: 64px;
    margin: 0 auto 18px;
    background: rgba(255,255,255,0.07);
    border: 1px solid rgba(124,108,240,0.5);
    border-radius: 18px;
    display: flex; align-items: center; justify-content: center;
    font-size: 30px;
    box-shadow: 0 0 24px rgba(124,108,240,0.3);
}
'''


register_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Регистрация компании — Revisi</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="blob blob-3"></div>
<div class="card">
  <div class="icon-box">{{ icon('shop', 30)|safe }}</div>
  <div class="title">Регистрация компании</div>
  <div class="subtitle">14 дней бесплатно, без карты</div>
  {% if error %}<div class="error">{{ error }}</div>{% endif %}
  <form method="post">
    <div class="form-group">
      <label>Название компании</label>
      <input class="input" type="text" name="company" required placeholder="ООО Ромашка">
    </div>
    <div class="form-group">
      <label>Учётная система</label>
      <select class="input" name="pos_system">
        <option value="">— выберите (необязательно) —</option>
        <option value="iiko">iiko</option>
        <option value="rkeeper">R-Keeper / StoreHouse</option>
        <option value="1c">1С</option>
        <option value="excel">Только Excel / без системы</option>
        <option value="other">Другая</option>
      </select>
      <div class="hint">Для iiko товары и бланк отчёта загрузятся автоматически из вашей выгрузки</div>
    </div>
    <div class="form-group">
      <label>Email владельца</label>
      <input class="input" type="email" name="email" required placeholder="owner@example.com">
    </div>
    <div class="form-group">
      <label>Ваш логин</label>
      <input class="input" type="text" name="username" required placeholder="admin">
      <div class="hint">Этот логин вы будете использовать для входа</div>
    </div>
    <div class="form-group">
      <label>Пароль</label>
      <input class="input" type="password" name="password" required minlength="6" placeholder="Минимум 6 символов">
    </div>
    <div class="form-group">
      <label>Повторите пароль</label>
      <input class="input" type="password" name="password2" required minlength="6" placeholder="Ещё раз">
    </div>
    <label class="checkbox-row">
      <input type="checkbox" name="terms" value="1" required>
      <span>Согласен с условиями использования</span>
    </label>
    <button class="btn-primary" type="submit">Создать аккаунт →</button>
  </form>
  <div class="bottom-link">
    Уже есть аккаунт? <a class="link" href="/login">Войти</a>
  </div>
</div>
</body>
</html>'''


register_success_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Проверьте email — Revisi</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="blob blob-3"></div>
<div class="card">
  <div class="icon-box">{{ icon('mail', 30)|safe }}</div>
  <div class="title">Проверьте email</div>
  <div class="subtitle">Мы отправили ссылку для подтверждения на указанный email</div>
  {% if verify_url %}
  <div style="background: rgba(124,108,240,0.12); border: 1px solid rgba(124,108,240,0.3);
              border-radius: 14px; padding: 16px; margin: 16px 0; font-size: 13px;
              color: rgba(255,255,255,0.75); text-align: center;">
    <div style="margin-bottom:10px;color:rgba(255,255,255,0.5);">Dev-режим: SMTP ещё не настроен.</div>
    <a class="link" href="{{ verify_url }}">Подтвердить сейчас →</a>
  </div>
  {% endif %}
  <div class="bottom-link">
    <a class="link" href="/login">Вернуться ко входу</a>
  </div>
</div>
</body>
</html>'''


message_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ title }} — Revisi</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="blob blob-3"></div>
<div class="card">
  <div class="icon-box">{{ icon('alert', 30)|safe }}</div>
  <div class="title">{{ title }}</div>
  <div class="subtitle">{{ text }}</div>
  <div class="bottom-link"><a class="link" href="/login">На главную</a></div>
</div>
</body>
</html>'''


placeholder_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ title }} — Revisi</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''
.info-row {
    display: flex; justify-content: space-between;
    padding: 10px 0; border-bottom: 1px solid rgba(255,255,255,0.07);
    font-size: 14px;
}
.info-row:last-child { border-bottom: none; }
.info-row .k { color: rgba(255,255,255,0.5); }
.info-row .v { color: white; font-weight: 600; }
.logout-btn {
    display: inline-block; margin-top: 18px;
    background: rgba(239,68,68,0.15);
    color: #fca5a5;
    border: 1px solid rgba(239,68,68,0.25);
    padding: 12px 24px;
    border-radius: 12px;
    text-decoration: none;
    font-weight: 600;
    font-size: 14px;
}
.logout-btn:hover { background: rgba(239,68,68,0.25); }
</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="blob blob-3"></div>
<div class="card" style="text-align:center;">
  <div class="icon-box">{{ icon('alert', 30)|safe }}</div>
  <div class="title">{{ title }}</div>
  <div class="subtitle">Раздел в разработке. Вернитесь позже.</div>
  <div style="text-align:left; margin: 18px 0;">
    <div class="info-row"><span class="k">Пользователь</span><span class="v">{{ username }}</span></div>
    <div class="info-row"><span class="k">Компания</span><span class="v">{{ org_name }}</span></div>
    <div class="info-row"><span class="k">Trial дней осталось</span><span class="v">{{ days_left }}</span></div>
  </div>
  <a class="logout-btn" href="/logout" onclick="return confirm('Выйти из аккаунта?');">Выйти</a>
</div>
</body>
</html>'''


login_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<title>Вход — Revisi</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''
.version { text-align: center; color: rgba(255,255,255,0.2); font-size: 11px; margin-top: 20px; }
</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="blob blob-3"></div>
<div class="card">
  <div class="icon-box">{{ icon('shop', 30)|safe }}</div>
  <div class="title">Инвентаризация</div>
  <div class="subtitle">Система учёта товаров</div>
  {% if error %}<div class="error">{{ error }}</div>{% endif %}
  <form method="post">
    <div class="form-group">
      <label>Email компании</label>
      <input class="input" type="email" name="email" required autocomplete="email" placeholder="company@example.com">
      <div class="hint">Общий email компании (один на всех сотрудников). Администратор указывает свой email, оператор — тот же email, который дал ему администратор.</div>
    </div>
    <div class="form-group">
      <label>Логин</label>
      <input class="input" type="text" name="username" required autocomplete="username" placeholder="Введите логин">
    </div>
    <div class="form-group">
      <label>Пароль</label>
      <input class="input" type="password" name="password" required autocomplete="current-password" placeholder="Введите пароль">
    </div>
    <button class="btn-primary" type="submit">Войти →</button>
  </form>
  <div class="bottom-link">
    Нет аккаунта? <a class="link" href="/register">Зарегистрировать компанию</a>
  </div>
  <div class="version">Версия: {{ now }}</div>
</div>
</body>
</html>'''


# ============== АДМИН ПАНЕЛЬ ==============
admin_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Админ панель — {{ org.name }}</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>
* { box-sizing: border-box; -webkit-tap-highlight-color: transparent; margin: 0; padding: 0; }
* { box-sizing: border-box; }
html, body { max-width: 100%; overflow-x: hidden; }
body {
  font-family: 'Outfit', sans-serif;
  background: linear-gradient(135deg, #13111C 0%, #1d1635 50%, #231b50 100%);
  background-attachment: fixed;
  min-height: 100vh;
  color: white;
  padding-bottom: 60px;
  margin: 0;
}
img, svg, video, canvas { max-width: 100%; height: auto; }
.blob { position: fixed; border-radius: 50%; filter: blur(80px); opacity: 0.25; pointer-events: none; z-index: 0; }
.blob-1 { width: 400px; height: 400px; background: radial-gradient(circle, #7c6cf0, #a855f7); top: -100px; left: -100px; }
.blob-2 { width: 350px; height: 350px; background: radial-gradient(circle, #a855f7, #6d28d9); bottom: -80px; right: -80px; }
.header {
  position: sticky; top: 0; z-index: 20;
  background: rgba(19,17,28,0.75);
  backdrop-filter: blur(20px);
  -webkit-backdrop-filter: blur(20px);
  border-bottom: 1px solid rgba(255,255,255,0.08);
  padding: 14px 20px;
  display: flex; align-items: center; gap: 12px; justify-content: space-between;
}
.header .brand { display: flex; flex-direction: column; min-width: 0; }
.header .brand .name { font-weight: 700; font-size: 17px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.header .brand .sub { font-size: 11px; color: rgba(255,255,255,0.5); text-transform: uppercase; letter-spacing: 0.6px; }
.header .actions { display: flex; gap: 8px; }
.btn {
  display: inline-flex; align-items: center; gap: 6px;
  background: rgba(255,255,255,0.07);
  border: 1px solid rgba(255,255,255,0.12);
  color: white;
  padding: 9px 14px;
  border-radius: 12px;
  font-family: 'Outfit', sans-serif;
  font-size: 13px;
  font-weight: 600;
  text-decoration: none;
  cursor: pointer;
  transition: all 0.15s;
}
.btn:hover { background: rgba(255,255,255,0.12); }
.btn-primary {
  background: linear-gradient(135deg, #7c6cf0, #a855f7);
  border-color: transparent;
  box-shadow: 0 6px 18px rgba(124,108,240,0.3);
}
.btn-primary:hover { filter: brightness(1.08); }
.btn-danger {
  background: rgba(239,68,68,0.12);
  border-color: rgba(239,68,68,0.3);
  color: #fca5a5;
}
.btn-danger:hover { background: rgba(239,68,68,0.22); }
.btn-small { padding: 6px 10px; font-size: 12px; }
.tabs-wrap {
  position: sticky; top: 60px; z-index: 15;
  background: rgba(19,17,28,0.65);
  backdrop-filter: blur(18px);
  border-bottom: 1px solid rgba(255,255,255,0.06);
  padding: 10px 16px;
}
.tabs {
  display: flex; gap: 8px;
  overflow-x: auto; scrollbar-width: none;
  padding-bottom: 4px;
}
.tabs::-webkit-scrollbar { display: none; }
.tab-pill {
  display: inline-flex; align-items: center; gap: 7px;
  white-space: nowrap;
  padding: 9px 16px;
  border-radius: 20px;
  background: rgba(255,255,255,0.06);
  border: 1px solid rgba(255,255,255,0.1);
  color: rgba(255,255,255,0.75);
  font-size: 13px;
  font-weight: 600;
  cursor: pointer;
  transition: all 0.15s;
}
.tab-pill svg { opacity: 0.85; }
.tab-pill.active svg { opacity: 1; }
.tab-pill.active {
  background: linear-gradient(135deg, #7c6cf0, #a855f7);
  color: white;
  border-color: transparent;
  box-shadow: 0 4px 12px rgba(124,108,240,0.35);
}
.container { max-width: 980px; width: 100%; margin: 20px auto; padding: 0 16px; position: relative; z-index: 1; }
.card {
  background: rgba(255,255,255,0.07);
  backdrop-filter: blur(20px);
  -webkit-backdrop-filter: blur(20px);
  border: 1px solid rgba(255,255,255,0.12);
  border-radius: 20px;
  padding: 20px;
  margin-bottom: 16px;
  box-shadow: 0 10px 30px rgba(0,0,0,0.3);
}
.card h2 { font-size: 17px; margin-bottom: 14px; font-weight: 700; }
.card h3 { font-size: 14px; margin-bottom: 10px; font-weight: 600; color: rgba(255,255,255,0.85); }
.tab-content { display: none; }
.tab-content.active { display: block; animation: fadeIn 0.25s ease-out; }
@keyframes fadeIn { from { opacity: 0; transform: translateY(6px); } to { opacity: 1; transform: translateY(0); } }
.flash-box {
  padding: 12px 14px; border-radius: 12px; margin-bottom: 12px;
  font-size: 13px; border: 1px solid rgba(255,255,255,0.12);
}
.flash-success { background: rgba(34,197,94,0.14); border-color: rgba(34,197,94,0.3); color: #86efac; }
.flash-error { background: rgba(239,68,68,0.14); border-color: rgba(239,68,68,0.3); color: #fca5a5; }
.input, select.input {
  width: 100%;
  background: rgba(255,255,255,0.07);
  border: 1px solid rgba(255,255,255,0.1);
  border-radius: 12px;
  color: white;
  padding: 12px 14px;
  font-family: 'Outfit', sans-serif;
  font-size: 14px;
}
.input::placeholder { color: rgba(255,255,255,0.3); }
.input:focus {
  outline: none;
  border-color: rgba(124,108,240,0.7);
  background: rgba(124,108,240,0.12);
}
select.input option { background: #1d1635; color: white; }
.row {
  display: flex; align-items: center; gap: 12px;
  padding: 12px 14px;
  background: rgba(255,255,255,0.04);
  border: 1px solid rgba(255,255,255,0.08);
  border-radius: 14px;
  margin-bottom: 8px;
}
.row .name { flex: 1; font-weight: 600; font-size: 14px; }
.row .meta { font-size: 12px; color: rgba(255,255,255,0.5); }
.inline-form { display: flex; gap: 8px; align-items: center; margin-top: 12px; flex-wrap: wrap; }
.inline-form .input { flex: 1; min-width: 160px; }
.empty {
  padding: 32px 20px; text-align: center; color: rgba(255,255,255,0.4); font-size: 13px;
  border: 1px dashed rgba(255,255,255,0.12); border-radius: 14px;
}
.empty-icon { font-size: 42px; margin-bottom: 10px; opacity: 0.6; }
.empty-title { color: rgba(255,255,255,0.8); font-size: 15px; font-weight: 600; margin-bottom: 4px; }
.empty-hint { font-size: 12px; color: rgba(255,255,255,0.45); }
.badge {
  display: inline-block; padding: 3px 10px; border-radius: 999px;
  font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;
}
.badge-admin { background: rgba(168,85,247,0.2); color: #d8b4fe; border: 1px solid rgba(168,85,247,0.4); }
.badge-operator { background: rgba(59,130,246,0.15); color: #93c5fd; border: 1px solid rgba(59,130,246,0.3); }
.cat-section { margin-bottom: 18px; }
.cat-title {
  font-size: 12px; text-transform: uppercase; letter-spacing: 0.8px;
  color: rgba(255,255,255,0.5); font-weight: 700; margin-bottom: 8px; padding-left: 4px;
}
.count-pill {
  display: inline-block; padding: 4px 12px; border-radius: 999px;
  background: rgba(124,108,240,0.15); border: 1px solid rgba(124,108,240,0.3);
  color: #c4b5fd; font-size: 12px; font-weight: 600;
}
.search-bar { margin-bottom: 14px; }
/* Bottom-sheet modal */
.modal-backdrop {
  display: none;
  position: fixed; inset: 0; background: rgba(0,0,0,0.55);
  z-index: 50; backdrop-filter: blur(4px);
  align-items: flex-end; justify-content: center;
}
.modal-backdrop.open { display: flex; animation: fadeIn 0.2s; }
.modal {
  background: rgba(29,22,53,0.98);
  backdrop-filter: blur(24px);
  border-top: 1px solid rgba(255,255,255,0.12);
  border-radius: 24px 24px 0 0;
  width: 100%; max-width: 540px;
  padding: 22px 22px 28px;
  max-height: 88vh; overflow-y: auto;
  animation: slideUp 0.25s ease-out;
}
@keyframes slideUp { from { transform: translateY(100%); } to { transform: translateY(0); } }
.modal h3 { margin-bottom: 14px; font-size: 18px; }
.form-group { margin-bottom: 12px; }
.form-group label {
  display: block; margin-bottom: 6px; font-weight: 600;
  color: rgba(255,255,255,0.55); font-size: 11px;
  text-transform: uppercase; letter-spacing: 0.8px;
}
.norms-table-wrap { overflow-x: auto; border-radius: 14px; }
.norms-table { width: 100%; border-collapse: separate; border-spacing: 0; min-width: 500px; }
.norms-table th, .norms-table td {
  padding: 10px 12px; text-align: left; font-size: 13px;
  border-bottom: 1px solid rgba(255,255,255,0.07);
}
.norms-table th {
  background: rgba(124,108,240,0.12);
  color: rgba(255,255,255,0.8); font-weight: 600;
  white-space: nowrap; position: sticky; top: 0;
}
.norms-table input {
  width: 72px;
  background: rgba(255,255,255,0.05);
  border: 1px solid rgba(255,255,255,0.1);
  color: white;
  padding: 6px 8px;
  border-radius: 8px;
  font-family: inherit;
  font-size: 13px;
  text-align: center;
}
.norms-table input:focus { outline: none; border-color: rgba(124,108,240,0.7); }
.tip {
  background: rgba(124,108,240,0.1);
  border: 1px solid rgba(124,108,240,0.25);
  border-radius: 12px;
  padding: 10px 14px;
  font-size: 13px;
  color: rgba(255,255,255,0.75);
  margin-bottom: 14px;
}
.preset-grid {
  display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
  gap: 10px; margin-top: 10px;
}
.syswrap { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
@media (max-width: 720px) { .syswrap { grid-template-columns: 1fr; } }
.syscol {
  background: rgba(255,255,255,0.03);
  border: 1px solid rgba(255,255,255,0.08);
  border-radius: 14px; padding: 16px;
}
.syscol.active { border-color: rgba(124,108,240,0.4); background: rgba(124,108,240,0.06); }
.sysname { font-weight: 700; font-size: 15px; }
.preset-btn {
  padding: 16px 12px; border-radius: 16px;
  background: rgba(255,255,255,0.05);
  border: 1px solid rgba(255,255,255,0.1);
  color: white; font-weight: 600; font-size: 14px;
  cursor: pointer; text-align: center;
  transition: all 0.15s;
}
.preset-btn:hover {
  background: rgba(124,108,240,0.2);
  border-color: rgba(124,108,240,0.5);
  transform: translateY(-2px);
}
.drop-zone {
  border: 2px dashed rgba(255,255,255,0.18);
  border-radius: 14px;
  padding: 20px;
  text-align: center;
  color: rgba(255,255,255,0.6);
  font-size: 13px;
  margin-top: 10px;
}
.import-result {
  margin-top: 12px; padding: 12px 14px;
  border-radius: 12px; font-size: 13px;
  background: rgba(34,197,94,0.1); border: 1px solid rgba(34,197,94,0.25);
  color: #86efac;
}
.import-result.err { background: rgba(239,68,68,0.1); border-color: rgba(239,68,68,0.25); color: #fca5a5; }
.pwd-box {
  background: rgba(250,204,21,0.1);
  border: 1px solid rgba(250,204,21,0.3);
  color: #fde68a;
  padding: 14px;
  border-radius: 12px;
  margin-bottom: 14px;
  font-size: 13px;
}
.pwd-box code { background: rgba(0,0,0,0.3); padding: 3px 8px; border-radius: 6px; font-size: 14px; }
.cred-row { display: flex; align-items: center; gap: 8px; margin: 6px 0; flex-wrap: wrap; }
.cred-label { color: rgba(255,255,255,0.6); font-size: 12px; min-width: 60px; }
.copy-btn {
  background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.2);
  color: #fde68a; padding: 4px 10px; border-radius: 8px; cursor: pointer;
  font-size: 13px; transition: all .15s ease;
}
.copy-btn:hover { background: rgba(255,255,255,0.2); transform: scale(1.05); }
.copy-btn:active { transform: scale(0.95); }
.row-actions { display: flex; gap: 6px; }
hr.soft { border: 0; border-top: 1px solid rgba(255,255,255,0.08); margin: 14px 0; }

/* === Ассортимент локаций === */
.assort-wrap { display: grid; grid-template-columns: 260px minmax(0, 1fr); gap: 16px; align-items: start; }
.assort-cols { min-width: 0; }
.assort-loc-list { display: flex; flex-direction: column; gap: 6px; min-width: 0; }
.assort-loc-item {
  display: flex; justify-content: space-between; align-items: center;
  padding: 10px 12px; border-radius: 10px; cursor: pointer;
  background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.06);
  font-size: 14px; transition: all 0.15s; gap: 8px;
}
.assort-loc-item:hover { background: rgba(255,255,255,0.07); }
.assort-loc-item.active { background: linear-gradient(135deg, #7c6cf0, #a855f7); border-color: transparent; }
.assort-loc-name { flex: 1; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.assort-loc-actions { display: flex; align-items: center; gap: 6px; flex-shrink: 0; }
.assort-loc-count {
  font-size: 11px; font-weight: 600; padding: 2px 8px;
  background: rgba(0,0,0,0.3); border-radius: 999px;
}
.assort-loc-del {
  padding: 2px 8px !important; font-size: 14px !important; line-height: 1;
  opacity: 0.5; transition: opacity 0.15s;
}
.assort-loc-item:hover .assort-loc-del,
.assort-loc-item.active .assort-loc-del { opacity: 1; }
.assort-loc-add {
  display: flex; gap: 6px; margin-top: 8px;
  padding: 8px; background: rgba(255,255,255,0.02); border-radius: 10px;
  border: 1px dashed rgba(255,255,255,0.12);
}
.assort-loc-add .input { flex: 1; min-width: 0; font-size: 13px; padding: 7px 10px; }
.assort-loc-add .btn { font-size: 12px; padding: 7px 10px; white-space: nowrap; }

/* === Плавающий виджет поддержки === */
.support-fab {
  position: fixed; right: 22px; bottom: 22px; z-index: 9000;
  width: 56px; height: 56px; border-radius: 50%; border: none;
  background: linear-gradient(135deg, #7c6cf0, #a855f7);
  color: #fff; font-size: 24px; cursor: pointer;
  box-shadow: 0 8px 24px rgba(124,108,240,0.45), 0 2px 6px rgba(0,0,0,0.25);
  display: flex; align-items: center; justify-content: center;
  transition: transform 0.15s, box-shadow 0.15s;
}
.support-fab:hover { transform: translateY(-2px); box-shadow: 0 12px 30px rgba(124,108,240,0.55), 0 4px 8px rgba(0,0,0,0.3); }
.support-fab:active { transform: translateY(0); }
.support-fab.is-open { background: linear-gradient(135deg, #ef4444, #f97316); }
.support-panel {
  position: fixed; right: 22px; bottom: 92px; z-index: 9000;
  width: 380px; max-width: calc(100vw - 32px); height: 560px; max-height: calc(100vh - 120px);
  background: rgba(20, 18, 38, 0.96); backdrop-filter: blur(14px);
  border: 1px solid rgba(255,255,255,0.1); border-radius: 18px;
  box-shadow: 0 24px 60px rgba(0,0,0,0.55);
  display: none; flex-direction: column; overflow: hidden;
}
.support-panel.is-open { display: flex; animation: spIn 0.18s ease-out; }
@keyframes spIn { from { transform: translateY(10px); opacity: 0; } to { transform: translateY(0); opacity: 1; } }
.sp-head {
  padding: 14px 16px; display: flex; justify-content: space-between; align-items: flex-start;
  background: linear-gradient(135deg, rgba(124,108,240,0.18), rgba(168,85,247,0.12));
  border-bottom: 1px solid rgba(255,255,255,0.08);
}
.sp-title { font-size: 15px; font-weight: 700; color: #fff; }
.sp-sub { font-size: 11px; color: rgba(255,255,255,0.55); margin-top: 2px; }
.sp-close {
  background: rgba(255,255,255,0.08); border: none; color: #fff;
  width: 28px; height: 28px; border-radius: 8px; cursor: pointer; font-size: 13px;
}
.sp-close:hover { background: rgba(255,255,255,0.15); }
.sp-body { padding: 14px 16px; display: flex; flex-direction: column; gap: 8px; overflow-y: auto; }
.sp-label { font-size: 11px; font-weight: 600; color: rgba(255,255,255,0.6); text-transform: uppercase; letter-spacing: 0.04em; }
.sp-input, .sp-textarea { font-size: 13px; }
.sp-textarea { resize: vertical; min-height: 90px; font-family: inherit; }
.sp-submit { margin-top: 4px; }
.sp-link {
  font-size: 12px; color: #a78bfa; text-decoration: none; text-align: center;
  margin-top: 8px; padding: 8px; border-radius: 8px; transition: background 0.15s;
}
.sp-link:hover { background: rgba(124,108,240,0.1); }
.sp-back {
  background: rgba(255,255,255,0.08); border: none; color: #fff;
  width: 28px; height: 28px; border-radius: 8px; cursor: pointer; font-size: 14px;
  margin-right: 8px; flex-shrink: 0;
}
.sp-back:hover { background: rgba(255,255,255,0.15); }
.sp-view { display: none; flex-direction: column; flex: 1; min-height: 0; }
.sp-view.active { display: flex; }
.sp-list { flex: 1; overflow-y: auto; padding: 12px; display: flex; flex-direction: column; gap: 6px; }
.sp-empty { text-align: center; padding: 24px 12px; color: rgba(255,255,255,0.45); font-size: 13px; }
.sp-ticket {
  padding: 12px; border-radius: 12px; cursor: pointer;
  background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.06);
  transition: all 0.15s;
}
.sp-ticket:hover { background: rgba(124,108,240,0.12); border-color: rgba(124,108,240,0.3); }
.sp-ticket-head { display: flex; justify-content: space-between; align-items: center; gap: 8px; margin-bottom: 4px; }
.sp-ticket-subj { font-weight: 600; font-size: 13px; flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.sp-status {
  font-size: 10px; font-weight: 700; padding: 2px 8px; border-radius: 999px;
  text-transform: uppercase; letter-spacing: 0.04em; flex-shrink: 0;
}
.sp-status-open { background: rgba(245,158,11,0.2); color: #fbbf24; }
.sp-status-answered { background: rgba(34,197,94,0.2); color: #86efac; }
.sp-status-closed { background: rgba(148,163,184,0.2); color: #cbd5e1; }
.sp-ticket-date { font-size: 11px; color: rgba(255,255,255,0.45); }
.sp-new-btn { margin: 12px; flex-shrink: 0; }
.sp-messages {
  flex: 1; overflow-y: auto; padding: 14px; display: flex; flex-direction: column; gap: 10px;
}
.sp-msg { display: flex; flex-direction: column; max-width: 85%; }
.sp-msg-user { align-self: flex-end; }
.sp-msg-owner { align-self: flex-start; }
.sp-msg-bubble {
  padding: 9px 12px; border-radius: 14px; font-size: 13px; line-height: 1.4;
  word-wrap: break-word; white-space: pre-wrap;
}
.sp-msg-user .sp-msg-bubble {
  background: linear-gradient(135deg, #7c6cf0, #a855f7); color: #fff;
  border-bottom-right-radius: 4px;
}
.sp-msg-owner .sp-msg-bubble {
  background: rgba(255,255,255,0.07); color: #fff;
  border: 1px solid rgba(255,255,255,0.08); border-bottom-left-radius: 4px;
}
.sp-msg-meta {
  font-size: 10px; color: rgba(255,255,255,0.4); margin-top: 3px; padding: 0 4px;
}
.sp-msg-user .sp-msg-meta { text-align: right; }
.sp-reply-form {
  display: flex; gap: 8px; padding: 10px 12px;
  border-top: 1px solid rgba(255,255,255,0.08); background: rgba(0,0,0,0.2);
}
.sp-reply-input {
  flex: 1; resize: none; min-height: 38px; max-height: 100px; padding: 8px 12px !important;
  font-size: 13px !important;
}
.sp-reply-btn {
  width: 40px; height: 40px; padding: 0 !important; font-size: 16px; flex-shrink: 0;
  display: flex; align-items: center; justify-content: center;
}
.sp-closed-note {
  padding: 10px; text-align: center; font-size: 11px;
  color: rgba(255,255,255,0.5); background: rgba(0,0,0,0.2);
  border-top: 1px solid rgba(255,255,255,0.08);
}
@media (max-width: 600px) {
  .support-fab { right: 14px; bottom: 14px; width: 52px; height: 52px; font-size: 22px; }
  .support-panel { right: 8px; left: 8px; bottom: 76px; width: auto; max-width: none; }
}
.assort-grid { display: grid; grid-template-columns: minmax(0, 1fr) minmax(0, 1fr); gap: 14px; }
.assort-col {
  background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.06);
  border-radius: 12px; padding: 10px; min-height: 320px; max-height: 60vh;
  display: flex; flex-direction: column; min-width: 0;
}
.assort-col-head {
  display: flex; justify-content: space-between; align-items: center; gap: 8px;
  font-size: 13px; font-weight: 600; color: rgba(255,255,255,0.85);
  margin-bottom: 8px; padding: 0 4px; flex-wrap: wrap;
}
.assort-list { flex: 1; min-width: 0; overflow-y: auto; overflow-x: hidden; display: flex; flex-direction: column; gap: 4px; }
.assort-item {
  display: flex; justify-content: space-between; align-items: center;
  padding: 8px 10px; border-radius: 8px; font-size: 13px;
  background: rgba(255,255,255,0.03); cursor: pointer; transition: all 0.1s;
}
.assort-item:hover { background: rgba(124,108,240,0.15); }
.assort-item .pname { flex: 1; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.assort-item .pcode { font-size: 11px; color: rgba(255,255,255,0.45); margin-left: 8px; }
.assort-item .ai-btn {
  flex-shrink: 0; width: 26px; height: 26px; border-radius: 6px; border: none;
  font-size: 14px; cursor: pointer; display: flex; align-items: center; justify-content: center;
}
.assort-item .ai-btn-add { background: #10b981; color: white; }
.assort-item .ai-btn-rm { background: #ef4444; color: white; }
.assort-item.muted { opacity: 0.4; pointer-events: none; }
.assort-empty { text-align: center; padding: 24px 12px; color: rgba(255,255,255,0.4); font-size: 12px; }
@media (max-width: 800px) {
  .assort-wrap { grid-template-columns: 1fr; }
  .assort-grid { grid-template-columns: minmax(0, 1fr); }
  .assort-col { max-height: 50vh; }
}

/* === Бургер + drawer === */
.burger-btn {
  display: inline-flex; flex-direction: column; justify-content: center; gap: 4px;
  width: 40px; height: 40px; border-radius: 10px; border: none;
  background: rgba(255,255,255,0.08); cursor: pointer; padding: 0 10px;
  transition: background 0.15s;
}
.burger-btn:hover { background: rgba(255,255,255,0.14); }
.burger-btn span {
  display: block; width: 20px; height: 2px; background: #fff;
  border-radius: 2px; transition: transform 0.2s, opacity 0.2s;
}
.burger-btn.is-open span:nth-child(1) { transform: translateY(6px) rotate(45deg); }
.burger-btn.is-open span:nth-child(2) { opacity: 0; }
.burger-btn.is-open span:nth-child(3) { transform: translateY(-6px) rotate(-45deg); }

/* На десктопе показываем вкладки, бургер прячем; на мобиле — наоборот */
.burger-btn { display: none; }
@media (max-width: 800px) {
  .burger-btn { display: inline-flex; }
  .tabs-wrap { display: none !important; }
}

.drawer-backdrop {
  position: fixed; inset: 0; background: rgba(0,0,0,0.6); backdrop-filter: blur(4px);
  z-index: 8500; opacity: 0; pointer-events: none; transition: opacity 0.2s;
}
.drawer-backdrop.is-open { opacity: 1; pointer-events: auto; }
.drawer {
  position: fixed; top: 0; left: 0; bottom: 0; z-index: 8600;
  width: 300px; max-width: 86vw; background: rgba(20, 18, 38, 0.98);
  backdrop-filter: blur(14px); border-right: 1px solid rgba(255,255,255,0.08);
  transform: translateX(-100%); transition: transform 0.25s ease-out;
  display: flex; flex-direction: column; overflow: hidden;
  box-shadow: 18px 0 50px rgba(0,0,0,0.5);
}
.drawer.is-open { transform: translateX(0); }
.drawer-head {
  padding: 16px 18px; display: flex; justify-content: space-between; align-items: flex-start;
  background: linear-gradient(135deg, rgba(124,108,240,0.22), rgba(168,85,247,0.14));
  border-bottom: 1px solid rgba(255,255,255,0.08);
}
.drawer-title { font-size: 16px; font-weight: 700; }
.drawer-org { font-size: 12px; color: rgba(255,255,255,0.55); margin-top: 2px; max-width: 200px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.drawer-nav { flex: 1; overflow-y: auto; padding: 12px; display: flex; flex-direction: column; gap: 4px; }
.drawer-link {
  display: flex; align-items: center; gap: 10px;
  padding: 12px 14px; border-radius: 10px; cursor: pointer;
  font-size: 14px; color: rgba(255,255,255,0.85); text-decoration: none;
  background: rgba(255,255,255,0.02); border: 1px solid transparent;
  transition: all 0.15s;
}
.drawer-link svg { opacity: 0.78; flex-shrink: 0; }
.drawer-link.active svg { opacity: 1; }
.drawer-link:hover { background: rgba(255,255,255,0.06); }
.drawer-link.active {
  background: linear-gradient(135deg, rgba(124,108,240,0.3), rgba(168,85,247,0.2));
  border-color: rgba(124,108,240,0.45); color: #fff;
}
.drawer-badge {
  margin-left: auto; background: #ef4444; color: white;
  padding: 1px 7px; border-radius: 999px; font-size: 10px; font-weight: 700;
}
.drawer-footer {
  padding: 12px; border-top: 1px solid rgba(255,255,255,0.08);
  display: flex; flex-direction: column; gap: 6px;
}
.drawer-foot-link {
  display: block; padding: 10px 12px; border-radius: 10px; text-decoration: none;
  font-size: 13px; color: rgba(255,255,255,0.8); background: rgba(255,255,255,0.04);
  text-align: center; transition: background 0.15s;
}
.drawer-foot-link:hover { background: rgba(255,255,255,0.08); }
.drawer-foot-danger { color: #fca5a5; }
.drawer-foot-danger:hover { background: rgba(239,68,68,0.16); }

@media (max-width: 800px) {
  /* На мобиле все действия живут в drawer-меню — шапку чистим, чтобы не толкаться */
  .header .actions { display: none; }
}
@media (max-width: 600px) {
  .header { padding: 12px 14px; gap: 8px; }
  .header .brand .name { font-size: 15px; max-width: none; }
  .header .brand .sub { font-size: 11px; }
  .burger-btn { width: 38px; height: 38px; }
  .container { padding: 12px; }
  .card { padding: 14px; }
  .card h2 { font-size: 16px; }
  .modal { width: calc(100vw - 24px); max-height: calc(100vh - 60px); }
  .assort-col { max-height: 60vh; min-height: 240px; }
  .assort-item { font-size: 14px; padding: 10px 12px; }
  .assort-item .ai-btn { width: 32px; height: 32px; font-size: 16px; }
  .assort-loc-add { flex-wrap: wrap; }
  .assort-loc-add .btn { flex-basis: 100%; }
}
</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>

{% if is_impersonating %}
<div style="background:linear-gradient(135deg,#a855f7,#7c6cf0);padding:10px 14px;display:flex;align-items:center;justify-content:space-between;font-size:12px;font-weight:600;gap:10px;flex-wrap:wrap;">
  <span style="flex:1;min-width:200px;line-height:1.35;display:inline-flex;align-items:center;gap:6px;">{{ icon('mask', 15)|safe }} Вы вошли как admin <b>{{ org.name }}</b> от имени владельца. Действия пишутся в аудит.</span>
  <a href="/owner/stop_impersonate" style="background:rgba(255,255,255,0.22);color:white;padding:6px 12px;border-radius:8px;text-decoration:none;font-weight:700;white-space:nowrap;font-size:12px;">← В Owner</a>
</div>
{% endif %}

<div class="header">
  <button class="burger-btn" onclick="toggleDrawer()" aria-label="Меню" type="button">
    <span></span><span></span><span></span>
  </button>
  <div class="brand">
    <div class="name" style="display:flex;align-items:center;gap:8px;">{{ icon('shop', 18)|safe }} {{ org.name }}</div>
    <div class="sub" id="header-sub">Админ панель</div>
  </div>
  <div class="actions">
    <a class="btn" href="/revision" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('chart', 16)|safe }} Ревизия</a>
    <a class="btn btn-danger" href="/logout" onclick="return confirm('Выйти из аккаунта?');" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('logout', 16)|safe }} Выйти</a>
  </div>
</div>

<!-- Drawer / Burger menu -->
<div class="drawer-backdrop" id="drawer-backdrop" onclick="toggleDrawer()"></div>
<aside class="drawer" id="drawer">
  <div class="drawer-head">
    <div>
      <div class="drawer-title">Разделы</div>
      <div class="drawer-org">{{ org.name }}</div>
    </div>
    <button class="sp-close" onclick="toggleDrawer()" aria-label="Закрыть">✕</button>
  </div>
  <nav class="drawer-nav">
    <a class="drawer-link active" data-tab="tab-locations" onclick="switchTab('tab-locations', true)">{{ icon('mappin', 18)|safe }} Локации и ассортимент</a>
    <a class="drawer-link" data-tab="tab-categories" onclick="switchTab('tab-categories', true)">{{ icon('folder', 18)|safe }} Категории</a>
    <a class="drawer-link" data-tab="tab-products" onclick="switchTab('tab-products', true)">{{ icon('box', 18)|safe }} Товары</a>
    <a class="drawer-link" data-tab="tab-users" onclick="switchTab('tab-users', true)">{{ icon('users', 18)|safe }} Пользователи</a>
    <a class="drawer-link" data-tab="tab-norms" onclick="switchTab('tab-norms', true)">{{ icon('ruler', 18)|safe }} Нормы</a>
    <a class="drawer-link" data-tab="tab-import" onclick="switchTab('tab-import', true)">{{ icon('download', 18)|safe }} Импорт</a>
    <a class="drawer-link" data-tab="tab-requests" onclick="switchTab('tab-requests', true)">{{ icon('inbox', 18)|safe }} Запросы{% if pending_group %} <span class="drawer-badge">1</span>{% endif %}</a>
    <a class="drawer-link" data-tab="tab-history" onclick="switchTab('tab-history', true)">{{ icon('archive', 18)|safe }} История</a>
  </nav>
  <div class="drawer-footer">
    <a href="/revision" class="drawer-foot-link" style="display:flex;align-items:center;justify-content:center;gap:8px;">{{ icon('chart', 16)|safe }} К ревизии</a>
    <a href="/logout" class="drawer-foot-link drawer-foot-danger" onclick="return confirm('Выйти из аккаунта?');">↩ Выйти</a>
  </div>
</aside>

<!-- Floating support widget -->
<button class="support-fab" id="support-fab" onclick="toggleSupportPanel()" title="Поддержка" aria-label="Открыть чат поддержки">
  {{ icon('chat', 24)|safe }}
</button>
<div class="support-panel" id="support-panel">
  <div class="sp-head">
    <button class="sp-back" id="sp-back" onclick="supportShowList()" style="display:none;" aria-label="Назад">←</button>
    <div style="flex:1;min-width:0;">
      <div class="sp-title" id="sp-title">Поддержка Revisi</div>
      <div class="sp-sub" id="sp-sub">Ответим в течение рабочего дня</div>
    </div>
    <button class="sp-close" onclick="toggleSupportPanel()" aria-label="Закрыть">✕</button>
  </div>

  <!-- View 1: Tickets list -->
  <div class="sp-view sp-view-list active" id="sp-view-list">
    <div class="sp-list" id="sp-tickets-list">
      <div class="sp-empty">Загрузка…</div>
    </div>
    <button class="btn btn-primary sp-new-btn" onclick="supportShowNew()">+ Новое обращение</button>
  </div>

  <!-- View 2: Single ticket chat -->
  <div class="sp-view" id="sp-view-chat">
    <div class="sp-messages" id="sp-messages"></div>
    <form class="sp-reply-form" id="sp-reply-form" onsubmit="supportSendReply(event)">
      <textarea class="input sp-reply-input" name="body" rows="2" maxlength="3000" placeholder="Ваш ответ…" required></textarea>
      <button class="btn btn-primary sp-reply-btn" type="submit" aria-label="Отправить">➤</button>
    </form>
  </div>

  <!-- View 3: New ticket form -->
  <div class="sp-view" id="sp-view-new">
    <form class="sp-body" id="support-quick-form" onsubmit="submitSupportQuick(event)">
      <label class="sp-label">Тема</label>
      <input class="input sp-input" name="subject" maxlength="200" placeholder="Кратко о проблеме" required>
      <label class="sp-label">Сообщение</label>
      <textarea class="input sp-textarea" name="body" rows="5" maxlength="3000" placeholder="Опишите, что произошло — мы поможем" required></textarea>
      <button class="btn btn-primary sp-submit" type="submit">Отправить</button>
    </form>
  </div>
</div>

<div class="tabs-wrap">
  <div class="tabs" id="tabs">
    <div class="tab-pill active" data-tab="tab-locations" onclick="switchTab('tab-locations')">{{ icon('mappin', 16)|safe }} Локации и ассортимент</div>
    <div class="tab-pill" data-tab="tab-categories" onclick="switchTab('tab-categories')">{{ icon('folder', 16)|safe }} Категории</div>
    <div class="tab-pill" data-tab="tab-products" onclick="switchTab('tab-products')">{{ icon('box', 16)|safe }} Товары</div>
    <div class="tab-pill" data-tab="tab-users" onclick="switchTab('tab-users')">{{ icon('users', 16)|safe }} Пользователи</div>
    <div class="tab-pill" data-tab="tab-norms" onclick="switchTab('tab-norms')">{{ icon('ruler', 16)|safe }} Нормы</div>
    <div class="tab-pill" data-tab="tab-import" onclick="switchTab('tab-import')">{{ icon('download', 16)|safe }} Импорт</div>
    <div class="tab-pill" data-tab="tab-requests" onclick="switchTab('tab-requests')">{{ icon('inbox', 16)|safe }} Запросы{% if pending_group %} <span style="background:#ef4444;color:white;padding:1px 7px;border-radius:999px;font-size:10px;margin-left:4px;">1</span>{% endif %}</div>
    <div class="tab-pill" data-tab="tab-history" onclick="switchTab('tab-history')">{{ icon('archive', 16)|safe }} История</div>
  </div>
</div>

<div class="container">

  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for category, msg in messages %}
      <div class="flash-box {% if category == 'error' %}flash-error{% else %}flash-success{% endif %}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}

  {% if new_user_pwd %}
  <div class="pwd-box">
    <span style="display:inline-flex;align-items:center;gap:6px;">{{ icon('key', 16)|safe }} <b>Данные для входа оператора «{{ new_user_name }}»</b></span><br>
    <div class="cred-row" style="display:flex;align-items:center;gap:6px;">{{ icon('mail', 15)|safe }} <span class="cred-label">Email:</span> <code id="cred-email">{{ org.owner_email }}</code> <button type="button" class="copy-btn" onclick="copyCred('cred-email')">{{ icon('clipboard', 14)|safe }}</button></div>
    <div class="cred-row" style="display:flex;align-items:center;gap:6px;">{{ icon('user', 15)|safe }} <span class="cred-label">Логин:</span> <code id="cred-login">{{ new_user_name }}</code> <button type="button" class="copy-btn" onclick="copyCred('cred-login')">{{ icon('clipboard', 14)|safe }}</button></div>
    <div class="cred-row" style="display:flex;align-items:center;gap:6px;">{{ icon('lock', 15)|safe }} <span class="cred-label">Пароль:</span> <code id="cred-pwd">{{ new_user_pwd }}</code> <button type="button" class="copy-btn" onclick="copyCred('cred-pwd')">{{ icon('clipboard', 14)|safe }}</button></div>
    <button type="button" class="btn btn-primary btn-small" onclick="copyAllCreds('{{ org.owner_email }}', '{{ new_user_name }}', '{{ new_user_pwd }}')" style="margin-top:8px;display:inline-flex;align-items:center;gap:6px;">{{ icon('clipboard', 14)|safe }} Скопировать всё</button>
    <div style="font-size:11px;color:rgba(255,255,255,0.5);margin-top:6px;">Сохраните сейчас — повторно пароль показан не будет.</div>
  </div>
  {% endif %}

  <!-- Tab: Локации + ассортимент -->
  <div class="tab-content active" id="tab-locations">
    <div class="card">
      <h3 style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">{{ icon('clock', 16)|safe }} Часовой пояс</h3>
      <div style="font-size:12px;color:rgba(255,255,255,0.55);margin-bottom:10px;">Время ревизий и дата в отчётах показываются в этом поясе.</div>
      <form method="post" action="/admin/settings/timezone" class="inline-form">
        <select class="input" name="timezone" style="flex:1;min-width:200px;">
          {% for val, label in tz_choices %}
          <option value="{{ val }}" {% if org.timezone == val %}selected{% endif %}>{{ label }}</option>
          {% endfor %}
        </select>
        <button class="btn btn-primary btn-small" type="submit" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('save', 14)|safe }} Сохранить</button>
      </form>
    </div>
    <div class="card">
      <h2>Локации и ассортимент <span class="count-pill">{{ locations|length }}</span></h2>
      <div style="font-size:13px;color:rgba(255,255,255,0.6);margin-bottom:14px;">
        Создайте локацию и сразу добавьте в неё товары. Товары, не привязанные ни к одной локации, не видны операторам и не попадают в отчёт.
      </div>
      {% if not locations %}
        <div class="empty">
          <div class="empty-icon">{{ icon('mappin', 32)|safe }}</div>
          <div class="empty-title">Пока нет локаций</div>
          <div class="empty-hint">Добавьте первую локацию ниже — потом выберите её и заполните ассортиментом.</div>
        </div>
        <hr class="soft">
        <form method="post" action="/admin/locations/add" class="inline-form">
          <input class="input" type="text" name="name" placeholder="Название локации (склад, кухня, бар…)" required maxlength="100">
          <button class="btn btn-primary" type="submit">+ Добавить локацию</button>
        </form>
      {% else %}
      <div class="assort-wrap">
        <div class="assort-loc-list">
          {% for loc in locations %}
          <div class="assort-loc-item{% if loop.first %} active{% endif %}" data-loc-id="{{ loc.id }}" onclick="selectAssortLoc({{ loc.id }})">
            <span class="assort-loc-name" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('mappin', 15)|safe }} {{ loc.name }}</span>
            <span class="assort-loc-actions">
              <span class="assort-loc-count" id="assort-count-{{ loc.id }}">{{ assortment_map.get(loc.id, [])|length }}</span>
              <form method="post" action="/admin/locations/delete/{{ loc.id }}" onsubmit="event.stopPropagation();return confirm('Удалить локацию «{{ loc.name }}»? Все её ревизии и привязки товаров будут удалены.');" style="display:inline;margin:0;">
                <button class="btn btn-danger btn-small assort-loc-del" type="submit" title="Удалить локацию" onclick="event.stopPropagation()">×</button>
              </form>
            </span>
          </div>
          {% endfor %}
          <form method="post" action="/admin/locations/add" class="inline-form assort-loc-add">
            <input class="input" type="text" name="name" placeholder="Новая локация" required maxlength="100">
            <button class="btn btn-primary btn-small" type="submit">+ Добавить</button>
          </form>
        </div>
        {% if not products %}
        <div class="assort-cols">
          <div class="empty">
            <div class="empty-icon">{{ icon('box', 32)|safe }}</div>
            <div class="empty-title">Сначала добавьте товары</div>
            <div class="empty-hint">Перейдите во вкладку «Товары» и создайте позиции, потом распределите их по локациям.</div>
          </div>
        </div>
        {% else %}
        <div class="assort-cols">
          <div class="search-bar" style="margin-bottom:10px;">
            <input class="input" id="assort-search" type="text" placeholder="Поиск по товарам…" oninput="filterAssort()">
          </div>
          <div class="assort-grid">
            <div class="assort-col">
              <div class="assort-col-head">
                <span>Не в ассортименте <span id="assort-avail-count" class="count-pill">0</span></span>
              </div>
              <div class="assort-list" id="assort-available"></div>
            </div>
            <div class="assort-col">
              <div class="assort-col-head">
                <span>В ассортименте <span id="assort-selected-count" class="count-pill">0</span></span>
                <button class="btn btn-danger btn-small" id="assort-clear-btn" onclick="clearAssortLoc()" style="padding:4px 10px;font-size:11px;">Убрать все</button>
              </div>
              <div class="assort-list" id="assort-selected"></div>
            </div>
          </div>
        </div>
        {% endif %}
      </div>
      {% endif %}
    </div>
  </div>

  <!-- Tab: Категории -->
  <div class="tab-content" id="tab-categories">
    <div class="card">
      <h2>Категории <span class="count-pill">{{ categories|length }}</span></h2>
      {% if categories %}
        {% for cat in categories %}
        <div class="row">
          <div class="name" style="display:flex;align-items:center;gap:6px;">{{ icon('folder', 15)|safe }} {{ cat.name }}</div>
          <form method="post" action="/admin/categories/delete/{{ cat.id }}" onsubmit="return confirm('Удалить категорию «{{ cat.name }}»?');">
            <button class="btn btn-danger btn-small" type="submit">Удалить</button>
          </form>
        </div>
        {% endfor %}
      {% else %}
        <div class="empty">
          <div class="empty-icon">{{ icon('folder', 32)|safe }}</div>
          <div class="empty-title">Пока нет категорий</div>
          <div class="empty-hint">Категории помогают группировать товары: «Молочка», «Овощи», «Мясо» и т.д.</div>
        </div>
      {% endif %}
      <hr class="soft">
      <form method="post" action="/admin/categories/add" class="inline-form">
        <input class="input" type="text" name="name" placeholder="Название категории" required maxlength="200">
        <button class="btn btn-primary" type="submit">+ Добавить</button>
      </form>
    </div>
  </div>

  <!-- Tab: Товары -->
  <div class="tab-content" id="tab-products">
    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap;">
        <h2 style="margin:0;">Товары <span class="count-pill">Всего: <span id="products-count">{{ products|length }}</span></span></h2>
        <button class="btn btn-primary" onclick="openModal('modal-add-product')">+ Добавить товар</button>
      </div>
      <div class="search-bar">
        <input class="input" id="product-search" type="text" placeholder="Поиск по названию или коду…" oninput="filterProducts()">
      </div>
      {% if products %}
        {% set uncat = grouped.get(None, []) %}
        {% for cat in categories %}
          {% set items = grouped.get(cat.id, []) %}
          {% if items %}
          <div class="cat-section" data-cat="{{ cat.name }}">
            <div class="cat-title">{{ cat.name }} · {{ items|length }}</div>
            {% for p in items %}
            <div class="row product-row" id="prod-row-{{ p.id }}" data-name="{{ p.name|lower }}" data-code="{{ (p.code or '')|lower }}">
              <div style="flex:1;">
                <div class="name">{{ p.name }}</div>
                <div class="meta">{{ p.unit }}</div>
              </div>
              <div class="row-actions">
                <button class="btn btn-small" title="Редактировать" onclick='openEditProduct({{ p.id }}, {{ p.name|tojson }}, {{ (p.code or "")|tojson }}, {{ p.unit|tojson }}, {{ (p.category_id or 0) }})'>{{ icon('pencil', 14)|safe }}</button>
                <button class="btn btn-danger btn-small" onclick='deleteProduct({{ p.id }}, {{ p.name|tojson }})'>×</button>
              </div>
            </div>
            {% endfor %}
          </div>
          {% endif %}
        {% endfor %}
        {% if uncat %}
        <div class="cat-section" data-cat="без категории">
          <div class="cat-title">Без категории · {{ uncat|length }}</div>
          {% for p in uncat %}
          <div class="row product-row" id="prod-row-{{ p.id }}" data-name="{{ p.name|lower }}" data-code="{{ (p.code or '')|lower }}">
            <div style="flex:1;">
              <div class="name">{{ p.name }}</div>
              <div class="meta">{{ p.unit }}</div>
            </div>
            <div class="row-actions">
              <button class="btn btn-small" title="Редактировать" onclick='openEditProduct({{ p.id }}, {{ p.name|tojson }}, {{ (p.code or "")|tojson }}, {{ p.unit|tojson }}, 0)'>{{ icon('pencil', 14)|safe }}</button>
              <button class="btn btn-danger btn-small" onclick='deleteProduct({{ p.id }}, {{ p.name|tojson }})'>×</button>
            </div>
          </div>
          {% endfor %}
        </div>
        {% endif %}
      {% else %}
        <div class="empty">
          <div class="empty-icon">{{ icon('box', 32)|safe }}</div>
          <div class="empty-title">Пока нет товаров</div>
          <div class="empty-hint">Нажмите «+ Добавить товар» или загрузите Excel во вкладке «Импорт».</div>
        </div>
      {% endif %}
    </div>
  </div>

  <!-- Tab: Пользователи -->
  <div class="tab-content" id="tab-users">
    <div class="card">
      <h2>Пользователи <span class="count-pill">{{ users|length }}{% if org.plan == 'free' %} / {{ free_max_users }}{% endif %}</span></h2>
      {% if user_limit_reached %}
      <div class="flash-box flash-error">
        Достигнут лимит пользователей тарифа FREE ({{ free_max_users }}). Обновите тариф, чтобы добавлять больше.
      </div>
      {% endif %}
      {% for u in users %}
      <div class="row">
        <div style="flex:1;">
          <div class="name">{{ u.username }}
            <span class="badge {% if u.role == 'admin' %}badge-admin{% else %}badge-operator{% endif %}">{{ u.role }}</span>
          </div>
          <div class="meta">{{ u.email or '—' }}</div>
        </div>
        {% if u.id != current_user_id %}
        <form method="post" action="/admin/users/delete/{{ u.id }}" onsubmit="return confirm('Удалить «{{ u.username }}»?');">
          <button class="btn btn-danger btn-small" type="submit">Удалить</button>
        </form>
        {% else %}
        <span class="meta">это вы</span>
        {% endif %}
      </div>
      {% endfor %}
      <hr class="soft">
      <h3>Добавить оператора</h3>
      <form method="post" action="/admin/users/add">
        <div class="form-group">
          <label>Логин*</label>
          <input class="input" type="text" name="username" required maxlength="100" placeholder="operator1">
        </div>
        <div class="form-group">
          <label>Email (опционально)</label>
          <input class="input" type="email" name="email" placeholder="worker@example.com">
        </div>
        <div class="form-group">
          <label>Пароль (опционально — сгенерируется автоматически)</label>
          <input class="input" type="text" name="password" placeholder="Оставьте пустым для автогенерации">
        </div>
        <button class="btn btn-primary" type="submit" {% if user_limit_reached %}disabled{% endif %}>Создать</button>
      </form>
    </div>
  </div>

  <!-- Tab: Нормы -->
  <div class="tab-content" id="tab-norms">
    <div class="card">
      <h2>Нормы</h2>
      <div class="tip" style="display:flex;align-items:flex-start;gap:6px;"><span style="flex-shrink:0;margin-top:1px;">{{ icon('bulb', 15)|safe }}</span> Нормы — сколько каждого товара должно быть на каждой локации. Используется для подсветки дефицита в ревизии.</div>
      {% if products and locations %}
      <div class="norms-table-wrap">
        <table class="norms-table" id="norms-table">
          <thead>
            <tr>
              <th>Товар</th>
              {% for loc in locations %}<th>{{ loc.name }}</th>{% endfor %}
            </tr>
          </thead>
          <tbody>
            {% for p in products %}
            <tr>
              <td><b>{{ p.name }}</b> <span class="meta" style="font-size:11px;color:rgba(255,255,255,0.4);">({{ p.unit }})</span></td>
              {% for loc in locations %}
              <td>
                <input type="text" inputmode="decimal" pattern="[0-9]*[.,]?[0-9]*"
                       data-pid="{{ p.id }}" data-lid="{{ loc.id }}"
                       value="{{ norms_map.get((p.id, loc.id), 0) }}">
              </td>
              {% endfor %}
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      <div style="margin-top:14px;display:flex;gap:10px;align-items:center;">
        <button class="btn btn-primary" onclick="saveNorms()" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('save', 15)|safe }} Сохранить нормы</button>
        <span id="norms-status" class="meta"></span>
      </div>
      {% else %}
      <div class="empty">
        <div class="empty-icon">{{ icon('ruler', 32)|safe }}</div>
        <div class="empty-title">Нет данных для норм</div>
        <div class="empty-hint">Сначала добавьте локации и товары — тогда здесь появится таблица для задания норм.</div>
      </div>
      {% endif %}
    </div>
  </div>

  <!-- Tab: Импорт -->
  <div class="tab-content" id="tab-import">
    <div class="card">
      <h2 style="display:flex;align-items:center;gap:8px;">{{ icon('box', 18)|safe }} Загрузка из вашей системы</h2>
      <p class="meta" style="margin-bottom:14px;">Загрузите выгрузку из своей учётной системы — товары добавятся автоматически, а отчёт по ревизии будет выгружаться в привычном вам виде.</p>

      <div class="syswrap">
        <div class="syscol{% if org.pos_system in (none, '', 'iiko') %} active{% endif %}">
          <div class="sysname">iiko</div>
          <p class="meta" style="margin:6px 0 10px;">Выгрузите из iiko отчёт <b>«Бланк инвентаризации»</b> (.xls или .xlsx) и загрузите сюда. Товары, коды, единицы и категории импортируются автоматически.</p>
          <form method="post" action="/admin/iiko/upload" enctype="multipart/form-data"
                onsubmit="this.querySelector('button').textContent='Обработка…';this.querySelector('button').disabled=true;">
            <div class="drop-zone">
              <input type="file" name="file" accept=".xls,.xlsx" required style="color:white;">
            </div>
            <button class="btn btn-primary" type="submit" style="margin-top:10px;display:inline-flex;align-items:center;gap:6px;">{{ icon('upload', 15)|safe }} Загрузить бланк iiko</button>
          </form>
        </div>

        <div class="syscol">
          <div class="sysname">R-Keeper / StoreHouse · 1С</div>
          <p class="meta" style="margin:6px 0 10px;">Прямая загрузка для этих систем скоро будет. Пока выгрузите номенклатуру в <b>Excel</b> и загрузите её ниже как универсальный файл — структура: Категория, Название, Код, Ед. изм.</p>
          <a class="btn btn-small" href="#import-universal" onclick="document.getElementById('import-universal').scrollIntoView({behavior:'smooth'});return false;" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('arrow-left', 14)|safe }} К загрузке Excel</a>
        </div>
      </div>
    </div>

    <div class="card" id="import-universal">
      <h2>Универсальный импорт (Excel / CSV)</h2>
      <h3>1. Скачайте шаблон</h3>
      <p class="meta" style="margin-bottom:10px;">Excel с колонками: Категория, Название, Код, Ед. изм.</p>
      <a class="btn" href="/admin/import/template" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('download', 15)|safe }} Скачать шаблон .xlsx</a>

      <hr class="soft">
      <h3>2. Загрузите заполненный файл</h3>
      <form id="import-form" onsubmit="return uploadImport(event)">
        <div class="drop-zone">
          <input type="file" name="file" accept=".xlsx,.csv" required style="color:white;">
        </div>
        <button class="btn btn-primary" type="submit" style="margin-top:12px;display:inline-flex;align-items:center;gap:6px;">{{ icon('upload', 15)|safe }} Загрузить</button>
      </form>
      <div id="import-result"></div>
    </div>

    <div class="card">
      <h2 style="display:flex;align-items:center;gap:8px;">{{ icon('target', 18)|safe }} Быстрый старт — готовые наборы</h2>
      <p class="meta" style="margin-bottom:10px;">Загрузите типовые товары одной кнопкой.</p>
      <div class="preset-grid">
        <form method="post" action="/admin/import/preset/burger" onsubmit="return confirm('Добавить товары набора «Бургерная»?');">
          <button class="preset-btn" type="submit" style="width:100%;display:inline-flex;align-items:center;justify-content:center;gap:6px;">{{ icon('utensils', 15)|safe }} Бургерная</button>
        </form>
        <form method="post" action="/admin/import/preset/coffee" onsubmit="return confirm('Добавить товары набора «Кофейня»?');">
          <button class="preset-btn" type="submit" style="width:100%;display:inline-flex;align-items:center;justify-content:center;gap:6px;">{{ icon('utensils', 15)|safe }} Кофейня</button>
        </form>
        <form method="post" action="/admin/import/preset/bakery" onsubmit="return confirm('Добавить товары набора «Пекарня»?');">
          <button class="preset-btn" type="submit" style="width:100%;display:inline-flex;align-items:center;justify-content:center;gap:6px;">{{ icon('utensils', 15)|safe }} Пекарня</button>
        </form>
        <form method="post" action="/admin/import/preset/sushi" onsubmit="return confirm('Добавить товары набора «Суши»?');">
          <button class="preset-btn" type="submit" style="width:100%;display:inline-flex;align-items:center;justify-content:center;gap:6px;">{{ icon('utensils', 15)|safe }} Суши</button>
        </form>
      </div>
    </div>
  </div>

  <!-- Tab: Запросы -->
  <div class="tab-content" id="tab-requests">
    <div class="card">
      <h2>Запросы на подтверждение <span class="count-pill">{{ 1 if pending_group else 0 }}</span></h2>
      <div class="tip">{{ icon('alert', 15)|safe }} Подтверждение закрывает ревизию всего ресторана: все посчитанные локации объединяются в один общий бланк (остатки суммируются по локациям), и xlsx скачивается сразу.</div>
      {% if pending_group %}
        <div class="row" style="flex-wrap:wrap;gap:10px;">
          <div style="flex:1;min-width:200px;">
            <div class="name" style="display:flex;align-items:center;gap:6px;">{{ icon('chart', 16)|safe }} Ревизия ресторана</div>
            <div class="meta">Локации ({{ pending_group.loc_count }}): {{ pending_group.locations_str }}</div>
            <div class="meta">Операторы: {{ pending_group.operators_str }} · {{ pending_group.created_at }} · всего позиций: {{ pending_group.total_items }}</div>
          </div>
          <div class="row-actions" style="gap:8px;">
            <button class="btn btn-primary btn-small" id="confirm-rev-btn" onclick="confirmRevision({{ pending_group.confirm_id }})" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('check', 15)|safe }} Подтвердить и скачать</button>
            <form method="post" action="/admin/revisions/reject_all" onsubmit="return confirm('Вернуть всю ревизию на пересчёт? Операторы смогут продолжить подсчёт по всем локациям.');" style="display:inline;">
              <button class="btn btn-small" style="background:#f59e0b;color:#fff;" type="submit">↩ На пересчёт</button>
            </form>
          </div>
        </div>
      {% else %}
        <div class="empty">
          <div class="empty-icon">{{ icon('inbox', 32)|safe }}</div>
          <div class="empty-title">Нет ожидающих запросов</div>
          <div class="empty-hint">Здесь появится единый запрос на завершение ревизии, когда операторы посчитают локации.</div>
        </div>
      {% endif %}
    </div>
  </div>

  <!-- Tab: История -->
  <div class="tab-content" id="tab-history">
    <div class="card">
      <h2>История ревизий <span class="count-pill">{{ completed_groups|length }}</span></h2>
      {% if completed_groups %}
        {% for g in completed_groups %}
        <div class="row" style="flex-wrap:wrap;gap:10px;">
          <div style="flex:1;min-width:200px;">
            <div class="name" style="display:flex;align-items:center;gap:6px;">{{ icon('chart', 15)|safe }} Ревизия ресторана</div>
            <div class="meta">Локации ({{ g.loc_count }}): {{ g.locations_str }}</div>
            <div class="meta">{{ g.operators_str }} · {{ g.finished_at }} · всего позиций: {{ g.total_items }}</div>
          </div>
          <div class="row-actions">
            <a class="btn btn-small" href="/admin/revisions/{{ g.download_id }}/download_combined" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('download', 14)|safe }} Скачать общий</a>
          </div>
        </div>
        {% endfor %}
      {% else %}
        <div class="empty">
          <div class="empty-icon">{{ icon('archive', 32)|safe }}</div>
          <div class="empty-title">Нет завершённых ревизий</div>
          <div class="empty-hint">После подтверждения первой ревизии она появится здесь с возможностью скачать Excel-отчёт.</div>
        </div>
      {% endif %}
    </div>
  </div>

</div>

<!-- Modal: Добавить товар -->
<div class="modal-backdrop" id="modal-add-product" onclick="if(event.target===this) closeModal('modal-add-product')">
  <div class="modal">
    <h3>Новый товар</h3>
    <form id="add-prod-form" onsubmit="submitAddProduct(event)">
      <div class="form-group">
        <label>Название*</label>
        <input class="input" name="name" required maxlength="300" placeholder="Кола 0.5л">
      </div>
      <div class="form-group">
        <label>Категория</label>
        <select class="input" name="category_id" id="add-prod-cat" onchange="toggleNewCat(this)">
          <option value="">— без категории —</option>
          {% for c in categories %}<option value="{{ c.id }}">{{ c.name }}</option>{% endfor %}
          <option value="__new__">+ Новая категория…</option>
        </select>
      </div>
      <div class="form-group" id="new-cat-wrap" style="display:none;">
        <label>Название новой категории</label>
        <input class="input" name="new_category" maxlength="200" placeholder="Например: Напитки">
      </div>
      <div class="form-group">
        <label>Единица измерения</label>
        <input class="input" name="unit" value="шт" maxlength="20">
      </div>
      <div class="form-group">
        <label>Код</label>
        <input class="input" name="code" maxlength="100" placeholder="оставьте пустым для автогенерации">
      </div>
      <div style="display:flex;gap:10px;margin-top:6px;">
        <button type="button" class="btn" style="flex:1;" onclick="closeModal('modal-add-product')">Отмена</button>
        <button class="btn btn-primary" type="submit" style="flex:1;">Добавить</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal: Редактировать товар -->
<div class="modal-backdrop" id="modal-edit-product" onclick="if(event.target===this) closeModal('modal-edit-product')">
  <div class="modal">
    <h3>Редактировать товар</h3>
    <form id="edit-prod-form" onsubmit="submitEditProduct(event)">
      <div class="form-group">
        <label>Название*</label>
        <input class="input" name="name" id="edit-name" required maxlength="300">
      </div>
      <div class="form-group">
        <label>Категория</label>
        <select class="input" name="category_id" id="edit-cat">
          <option value="">— без категории —</option>
          {% for c in categories %}<option value="{{ c.id }}">{{ c.name }}</option>{% endfor %}
        </select>
      </div>
      <div class="form-group">
        <label>Единица измерения</label>
        <input class="input" name="unit" id="edit-unit" maxlength="20">
      </div>
      <div class="form-group">
        <label>Код</label>
        <input class="input" name="code" id="edit-code" maxlength="100">
      </div>
      <div style="display:flex;gap:10px;margin-top:6px;">
        <button type="button" class="btn" style="flex:1;" onclick="closeModal('modal-edit-product')">Отмена</button>
        <button class="btn btn-primary" type="submit" style="flex:1;">Сохранить</button>
      </div>
    </form>
  </div>
</div>

<script>
const TAB_TITLES = {
  'tab-locations': 'Локации и ассортимент',
  'tab-categories': 'Категории',
  'tab-products': 'Товары',
  'tab-users': 'Пользователи',
  'tab-norms': 'Нормы',
  'tab-import': 'Импорт',
  'tab-requests': 'Запросы',
  'tab-history': 'История',
};
function switchTab(id, closeDrawer) {
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.tab-pill').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.drawer-link').forEach(el => el.classList.remove('active'));
  const content = document.getElementById(id);
  if (content) content.classList.add('active');
  const pill = document.querySelector('.tab-pill[data-tab="' + id + '"]');
  if (pill) pill.classList.add('active');
  const dlink = document.querySelector('.drawer-link[data-tab="' + id + '"]');
  if (dlink) dlink.classList.add('active');
  const sub = document.getElementById('header-sub');
  if (sub && TAB_TITLES[id]) sub.textContent = TAB_TITLES[id];
  if (history.replaceState) history.replaceState(null, '', '#' + id);
  if (closeDrawer) setTimeout(toggleDrawer, 80);
}
function toggleDrawer() {
  const drawer = document.getElementById('drawer');
  const backdrop = document.getElementById('drawer-backdrop');
  const burger = document.querySelector('.burger-btn');
  if (!drawer) return;
  const open = drawer.classList.toggle('is-open');
  backdrop.classList.toggle('is-open', open);
  burger.classList.toggle('is-open', open);
  document.body.style.overflow = open ? 'hidden' : '';
}
(function initActiveTab() {
  let id = 'tab-locations';
  if (window.location.hash) {
    const h = window.location.hash.substring(1);
    if (document.getElementById(h)) id = h;
  }
  switchTab(id);
})();

function openModal(id) { document.getElementById(id).classList.add('open'); }
function closeModal(id) { document.getElementById(id).classList.remove('open'); }

function toggleNewCat(sel) {
  document.getElementById('new-cat-wrap').style.display = (sel.value === '__new__') ? 'block' : 'none';
  if (sel.value === '__new__') sel.name = '_ignore'; else sel.name = 'category_id';
}

function openEditProduct(id, name, code, unit, catId) {
  document.getElementById('edit-prod-form').action = '/admin/products/edit/' + id;
  document.getElementById('edit-name').value = name;
  document.getElementById('edit-code').value = code;
  document.getElementById('edit-unit').value = unit;
  document.getElementById('edit-cat').value = catId ? String(catId) : '';
  openModal('modal-edit-product');
}

function filterProducts() {
  const q = (document.getElementById('product-search').value || '').toLowerCase().trim();
  document.querySelectorAll('.product-row').forEach(r => {
    const n = r.dataset.name || '';
    const c = r.dataset.code || '';
    r.style.display = (!q || n.includes(q) || c.includes(q)) ? '' : 'none';
  });
  document.querySelectorAll('.cat-section').forEach(s => {
    const visible = Array.from(s.querySelectorAll('.product-row')).some(r => r.style.display !== 'none');
    s.style.display = visible ? '' : 'none';
  });
}

// === Ассортимент локаций ===
const ASSORT_PRODUCTS = [
  {% for p in products %}
  { id: {{ p.id }}, name: {{ p.name|tojson }}, code: {{ (p.code or '')|tojson }} },
  {% endfor %}
];
const ASSORT_MAP = {
  {% for loc_id, pids in assortment_map.items() %}
  "{{ loc_id }}": new Set({{ pids|tojson }}),
  {% endfor %}
};
let ASSORT_CURRENT_LOC = {% if locations %}{{ locations[0].id }}{% else %}null{% endif %};

function selectAssortLoc(locId) {
  ASSORT_CURRENT_LOC = locId;
  document.querySelectorAll('.assort-loc-item').forEach(el => {
    el.classList.toggle('active', String(el.dataset.locId) === String(locId));
  });
  renderAssortLists();
}

function renderAssortLists() {
  const locId = ASSORT_CURRENT_LOC;
  if (!locId) return;
  if (!ASSORT_MAP[locId]) ASSORT_MAP[locId] = new Set();
  const selSet = ASSORT_MAP[locId];
  const q = (document.getElementById('assort-search')?.value || '').toLowerCase().trim();
  const matches = (p) => !q || p.name.toLowerCase().includes(q) || (p.code || '').toLowerCase().includes(q);

  const availEl = document.getElementById('assort-available');
  const selEl = document.getElementById('assort-selected');
  const availHtml = [];
  const selHtml = [];
  let availCount = 0, selCount = 0;
  for (const p of ASSORT_PRODUCTS) {
    if (!matches(p)) continue;
    if (selSet.has(p.id)) {
      selCount++;
      selHtml.push(`<div class="assort-item" data-pid="${p.id}"><span class="pname">${escapeHtml(p.name)}</span><button class="ai-btn ai-btn-rm" title="Убрать" onclick="toggleAssort(${p.id}, this)">✕</button></div>`);
    } else {
      availCount++;
      availHtml.push(`<div class="assort-item" data-pid="${p.id}"><span class="pname">${escapeHtml(p.name)}</span><button class="ai-btn ai-btn-add" title="Добавить" onclick="toggleAssort(${p.id}, this)">+</button></div>`);
    }
  }
  availEl.innerHTML = availHtml.length ? availHtml.join('') : '<div class="assort-empty">Ничего не найдено</div>';
  selEl.innerHTML = selHtml.length ? selHtml.join('') : '<div class="assort-empty">Пока ничего не выбрано</div>';
  document.getElementById('assort-avail-count').textContent = availCount;
  document.getElementById('assort-selected-count').textContent = selCount;
  const totalSel = selSet.size;
  const badge = document.getElementById('assort-count-' + locId);
  if (badge) badge.textContent = totalSel;
}

function escapeHtml(s) {
  return String(s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}

function filterAssort() { renderAssortLists(); }

let SP_CURRENT_TICKET = null;

const SP_FAB_OPEN_SVG  = `{{ icon('close', 24)|safe }}`;
const SP_FAB_CLOSE_SVG = `{{ icon('chat', 24)|safe }}`;
function toggleSupportPanel() {
  const fab = document.getElementById('support-fab');
  const panel = document.getElementById('support-panel');
  if (!panel) return;
  const open = panel.classList.toggle('is-open');
  fab.classList.toggle('is-open', open);
  fab.innerHTML = open ? SP_FAB_OPEN_SVG : SP_FAB_CLOSE_SVG;
  if (open) {
    supportShowList();
    supportLoadList();
  }
}

function supportSetView(name) {
  document.querySelectorAll('.sp-view').forEach(el => el.classList.remove('active'));
  const v = document.getElementById('sp-view-' + name);
  if (v) v.classList.add('active');
  const back = document.getElementById('sp-back');
  if (back) back.style.display = (name === 'list') ? 'none' : '';
}

function supportShowList() {
  SP_CURRENT_TICKET = null;
  supportSetView('list');
  document.getElementById('sp-title').textContent = 'Поддержка Revisi';
  document.getElementById('sp-sub').textContent = 'Ответим в течение рабочего дня';
}

function supportShowNew() {
  supportSetView('new');
  document.getElementById('sp-title').textContent = 'Новое обращение';
  document.getElementById('sp-sub').textContent = 'Опишите проблему — мы ответим';
  setTimeout(() => document.querySelector('#support-quick-form input[name=subject]')?.focus(), 100);
}

function supportEscape(s) {
  return String(s || '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}

async function supportLoadList() {
  const listEl = document.getElementById('sp-tickets-list');
  listEl.innerHTML = '<div class="sp-empty">Загрузка…</div>';
  try {
    const res = await fetch('/support/api/tickets', { headers: {'X-Requested-With': 'XMLHttpRequest'} });
    const data = await res.json();
    if (!data.ok) {
      listEl.innerHTML = '<div class="sp-empty">Не удалось загрузить обращения.</div>';
      return;
    }
    if (!data.tickets.length) {
      listEl.innerHTML = '<div class="sp-empty">Пока обращений нет.<br>Нажмите «+ Новое обращение», чтобы написать в поддержку.</div>';
      return;
    }
    const statusLabel = { open: 'открыт', answered: 'отвечен', closed: 'закрыт' };
    listEl.innerHTML = data.tickets.map(t => `
      <div class="sp-ticket" onclick="supportOpenTicket(${t.id})">
        <div class="sp-ticket-head">
          <span class="sp-ticket-subj">${supportEscape(t.subject)}</span>
          <span class="sp-status sp-status-${t.status}">${statusLabel[t.status] || t.status}</span>
        </div>
        <div class="sp-ticket-date">${supportEscape(t.updated_at)}</div>
      </div>
    `).join('');
  } catch(e) {
    listEl.innerHTML = '<div class="sp-empty">Ошибка: ' + supportEscape(e) + '</div>';
  }
}

async function supportOpenTicket(ticketId) {
  SP_CURRENT_TICKET = ticketId;
  supportSetView('chat');
  const msgEl = document.getElementById('sp-messages');
  msgEl.innerHTML = '<div class="sp-empty">Загрузка переписки…</div>';
  document.getElementById('sp-title').textContent = 'Обращение #' + ticketId;
  document.getElementById('sp-sub').textContent = 'Загрузка…';
  try {
    const res = await fetch('/support/api/ticket/' + ticketId, { headers: {'X-Requested-With': 'XMLHttpRequest'} });
    const data = await res.json();
    if (!data.ok) {
      msgEl.innerHTML = '<div class="sp-empty">Не удалось загрузить тикет.</div>';
      return;
    }
    const t = data.ticket;
    document.getElementById('sp-title').textContent = (t.subject.length > 30 ? t.subject.slice(0, 30) + '…' : t.subject);
    document.getElementById('sp-sub').textContent = 'Статус: ' + ({open:'открыт',answered:'отвечен',closed:'закрыт'}[t.status] || t.status);
    renderSupportMessages(t.messages);
    // Hide reply form if closed
    const form = document.getElementById('sp-reply-form');
    const existingNote = document.getElementById('sp-closed-note');
    if (existingNote) existingNote.remove();
    if (t.status === 'closed') {
      form.style.display = 'none';
      const note = document.createElement('div');
      note.id = 'sp-closed-note';
      note.className = 'sp-closed-note';
      note.textContent = 'Тикет закрыт — нельзя ответить. Создайте новое обращение, если нужно.';
      form.parentNode.appendChild(note);
    } else {
      form.style.display = '';
    }
  } catch(e) {
    msgEl.innerHTML = '<div class="sp-empty">Ошибка: ' + supportEscape(e) + '</div>';
  }
}

function renderSupportMessages(messages) {
  const msgEl = document.getElementById('sp-messages');
  msgEl.innerHTML = messages.map(m => `
    <div class="sp-msg sp-msg-${m.author_type === 'user' ? 'user' : 'owner'}">
      <div class="sp-msg-bubble">${supportEscape(m.body)}</div>
      <div class="sp-msg-meta">${supportEscape(m.author_label)} · ${supportEscape(m.created_at)}</div>
    </div>
  `).join('');
  msgEl.scrollTop = msgEl.scrollHeight;
}

async function supportSendReply(ev) {
  ev.preventDefault();
  if (!SP_CURRENT_TICKET) return;
  const form = document.getElementById('sp-reply-form');
  const input = form.querySelector('textarea[name=body]');
  const body = input.value.trim();
  if (!body) return;
  const btn = form.querySelector('.sp-reply-btn');
  btn.disabled = true;
  try {
    const fd = new FormData();
    fd.append('body', body);
    const res = await fetch('/support/api/ticket/' + SP_CURRENT_TICKET + '/reply', {
      method: 'POST', body: fd,
      headers: {'X-Requested-With': 'XMLHttpRequest'},
    });
    const data = await res.json();
    if (data.ok) {
      input.value = '';
      // Append message
      const msgEl = document.getElementById('sp-messages');
      const m = data.message;
      msgEl.insertAdjacentHTML('beforeend', `
        <div class="sp-msg sp-msg-user">
          <div class="sp-msg-bubble">${supportEscape(m.body)}</div>
          <div class="sp-msg-meta">${supportEscape(m.author_label)} · ${supportEscape(m.created_at)}</div>
        </div>
      `);
      msgEl.scrollTop = msgEl.scrollHeight;
    } else {
      adminToast('✖ ' + (data.error || 'Не удалось отправить'), 'error');
    }
  } catch(e) {
    adminToast('✖ ' + e, 'error');
  } finally {
    btn.disabled = false;
  }
}

async function submitSupportQuick(ev) {
  ev.preventDefault();
  const form = document.getElementById('support-quick-form');
  const btn = form.querySelector('.sp-submit');
  const origText = btn.textContent;
  btn.disabled = true;
  btn.textContent = 'Отправка…';
  try {
    const res = await fetch('/support', {
      method: 'POST',
      body: new FormData(form),
      headers: {'X-Requested-With': 'XMLHttpRequest'},
    });
    const data = await res.json();
    if (data.ok) {
      form.reset();
      adminToast('✓ Обращение отправлено', 'success');
      // Jump straight into chat for the new ticket
      if (data.ticket_id) supportOpenTicket(data.ticket_id);
      else { supportShowList(); supportLoadList(); }
    } else {
      adminToast('✖ ' + (data.error || 'Не удалось отправить'), 'error');
    }
  } catch(e) {
    adminToast('✖ ' + e, 'error');
  } finally {
    btn.disabled = false;
    btn.textContent = origText;
  }
}

async function clearAssortLoc() {
  const locId = ASSORT_CURRENT_LOC;
  if (!locId) return;
  const locEl = document.querySelector('.assort-loc-item[data-loc-id="' + locId + '"] span');
  const locName = locEl ? locEl.textContent.trim() : 'локации';
  const selSet = ASSORT_MAP[locId] || new Set();
  if (selSet.size === 0) {
    adminToast('В ассортименте этой локации уже ничего нет.', 'success');
    return;
  }
  if (!confirm('Убрать все ' + selSet.size + ' товаров из ассортимента ' + locName + '?\\nВНИМАНИЕ: позиции этих товаров в открытых ревизиях этой локации тоже будут удалены, чтобы отчёт отражал реальный ассортимент. Завершённые ревизии не затрагиваются.')) return;
  const btn = document.getElementById('assort-clear-btn');
  if (btn) btn.disabled = true;
  try {
    const res = await fetch('/admin/locations/' + locId + '/assortment/clear', {
      method: 'POST',
      headers: {'X-Requested-With': 'XMLHttpRequest'},
    });
    const data = await res.json();
    if (data.ok) {
      ASSORT_MAP[locId] = new Set();
      renderAssortLists();
      let msg = '✓ Из «' + (data.location || locName) + '» убрано товаров: ' + (data.removed || 0);
      if (data.removed_counts) msg += ' (позиций в открытых ревизиях: ' + data.removed_counts + ')';
      adminToast(msg, 'success');
    } else {
      adminToast('✖ ' + (data.error || 'Ошибка'), 'error');
    }
  } catch(e) {
    adminToast('✖ ' + e, 'error');
  } finally {
    if (btn) btn.disabled = false;
  }
}

async function toggleAssort(pid, btn) {
  const locId = ASSORT_CURRENT_LOC;
  if (!locId) return;
  btn.disabled = true;
  try {
    const res = await fetch('/admin/locations/' + locId + '/assortment/toggle/' + pid, {
      method: 'POST',
      headers: {'X-Requested-With': 'XMLHttpRequest'},
    });
    const data = await res.json();
    if (data.ok) {
      if (!ASSORT_MAP[locId]) ASSORT_MAP[locId] = new Set();
      if (data.selected) ASSORT_MAP[locId].add(pid);
      else ASSORT_MAP[locId].delete(pid);
      renderAssortLists();
    } else {
      adminToast('✖ ' + (data.error || 'Ошибка'), 'error');
      btn.disabled = false;
    }
  } catch(e) {
    adminToast('✖ ' + e, 'error');
    btn.disabled = false;
  }
}

document.addEventListener('DOMContentLoaded', () => {
  if (ASSORT_CURRENT_LOC) renderAssortLists();
});

async function saveNorms() {
  const status = document.getElementById('norms-status');
  status.textContent = 'Сохранение…';
  const payload = {};
  document.querySelectorAll('#norms-table input').forEach(inp => {
    const pid = inp.dataset.pid, lid = inp.dataset.lid;
    if (!payload[pid]) payload[pid] = {};
    // Принимаем и "0,354" и "0.354"
    const rawVal = (inp.value || '0').replace(',', '.');
    payload[pid][lid] = parseFloat(rawVal) || 0;
  });
  try {
    const res = await fetch('/admin/norms/save', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
    });
    const data = await res.json();
    if (data.ok) status.textContent = '✔ Сохранено (' + data.saved + ' изменений)';
    else status.textContent = '✖ ' + (data.error || 'ошибка');
  } catch (e) {
    status.textContent = '✖ ' + e;
  }
}

function adminToast(msg, type) {
  const t = document.createElement('div');
  t.textContent = msg;
  t.style.cssText = 'position:fixed;bottom:30px;left:50%;transform:translateX(-50%);' +
    'background:' + (type === 'error' ? '#ef4444' : '#22c55e') + ';color:#fff;' +
    'padding:12px 22px;border-radius:14px;font-weight:600;font-size:14px;' +
    'box-shadow:0 8px 24px rgba(0,0,0,0.4);z-index:9999;opacity:0;transition:opacity .2s ease;';
  document.body.appendChild(t);
  requestAnimationFrame(() => { t.style.opacity = '1'; });
  setTimeout(() => { t.style.opacity = '0'; setTimeout(() => t.remove(), 250); }, 1800);
}

async function confirmRevision(revId) {
  if (!confirm('Завершить ревизию ресторана? Все посчитанные локации войдут в один общий отчёт, остатки суммируются.')) return;
  const btn = document.getElementById('confirm-rev-btn');
  const orig = btn ? btn.innerHTML : '';
  if (btn) { btn.disabled = true; btn.textContent = 'Формирование…'; }
  try {
    const res = await fetch('/admin/revisions/confirm/' + revId, { method: 'POST' });
    const ct = res.headers.get('content-type') || '';
    if (res.ok && ct.indexOf('spreadsheet') !== -1) {
      // Скачиваем файл программно
      const blob = await res.blob();
      let fname = 'revision.xlsx';
      const cd = res.headers.get('content-disposition') || '';
      const m = cd.match(/filename\*?=(?:UTF-8'')?["']?([^"';]+)/i);
      if (m) fname = decodeURIComponent(m[1]);
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url; a.download = fname;
      document.body.appendChild(a); a.click(); a.remove();
      URL.revokeObjectURL(url);
      adminToast('✓ Отчёт сформирован и скачан', 'success');
      // Обновляем страницу, чтобы Запросы/История отразили изменения
      setTimeout(() => { window.location.href = '/admin#tab-history'; window.location.reload(); }, 600);
    } else {
      // Ошибка — перезагружаем, чтобы показать flash-сообщение
      window.location.href = '/admin#tab-requests';
      window.location.reload();
    }
  } catch (e) {
    adminToast('✖ ' + e, 'error');
    if (btn) { btn.disabled = false; btn.innerHTML = orig; }
  }
}

async function copyToClipboard(text) {
  try {
    if (navigator.clipboard && window.isSecureContext) {
      await navigator.clipboard.writeText(text);
    } else {
      const ta = document.createElement('textarea');
      ta.value = text; ta.style.position='fixed'; ta.style.left='-9999px';
      document.body.appendChild(ta); ta.select();
      document.execCommand('copy'); ta.remove();
    }
    return true;
  } catch (e) { return false; }
}

async function copyCred(id) {
  const el = document.getElementById(id);
  if (!el) return;
  const ok = await copyToClipboard(el.textContent.trim());
  adminToast(ok ? '✓ Скопировано' : 'Не удалось скопировать', ok ? 'success' : 'error');
}

async function copyAllCreds(email, login, pwd) {
  const text = 'Email: ' + email + '\\nЛогин: ' + login + '\\nПароль: ' + pwd;
  const ok = await copyToClipboard(text);
  adminToast(ok ? '✓ Все данные скопированы' : 'Не удалось скопировать', ok ? 'success' : 'error');
}

async function submitAddProduct(ev) {
  ev.preventDefault();
  const form = document.getElementById('add-prod-form');
  const btn = form.querySelector('button[type=submit]');
  const origText = btn.textContent;
  btn.disabled = true;
  btn.textContent = 'Сохранение…';
  try {
    const res = await fetch('/admin/products/add', {
      method: 'POST',
      body: new FormData(form),
      headers: {'X-Requested-With': 'XMLHttpRequest'},
    });
    const data = await res.json();
    if (data.ok) {
      closeModal('modal-add-product');
      form.reset();
      document.getElementById('new-cat-wrap').style.display = 'none';
      const catSel = document.getElementById('add-prod-cat');
      if (catSel) catSel.name = 'category_id';
      adminToast('✓ ' + data.msg, 'success');
      // Только смена hash не перезагружает страницу — нужен reload, иначе
      // новый товар не появится в списке.
      window.location.hash = 'tab-products';
      setTimeout(() => window.location.reload(), 600);
    } else {
      adminToast('✖ ' + (data.error || 'Ошибка'), 'error');
    }
  } catch(e) {
    adminToast('✖ ' + e, 'error');
  } finally {
    btn.disabled = false;
    btn.textContent = origText;
  }
}

async function deleteProduct(pid, name) {
  if (!confirm('Удалить «' + name + '»?')) return;
  try {
    const res = await fetch('/admin/products/delete/' + pid, {
      method: 'POST',
      headers: {'X-Requested-With': 'XMLHttpRequest'},
    });
    const data = await res.json();
    if (data.ok) {
      const row = document.getElementById('prod-row-' + pid);
      if (row) row.remove();
      // Обновляем счётчик «Всего»
      const cnt = document.getElementById('products-count');
      if (cnt) {
        const n = parseInt(cnt.textContent, 10);
        if (!isNaN(n) && n > 0) cnt.textContent = n - 1;
      }
      adminToast('✓ ' + (data.msg || ('Товар «' + name + '» удалён')), 'success');
    } else {
      adminToast('✖ ' + (data.error || 'Ошибка'), 'error');
    }
  } catch(e) {
    adminToast('✖ ' + e, 'error');
  }
}

async function submitEditProduct(ev) {
  ev.preventDefault();
  const form = document.getElementById('edit-prod-form');
  const btn = form.querySelector('button[type=submit]');
  const origText = btn.textContent;
  btn.disabled = true;
  btn.textContent = 'Сохранение…';
  try {
    const res = await fetch(form.action, {
      method: 'POST',
      body: new FormData(form),
      headers: {'X-Requested-With': 'XMLHttpRequest'},
    });
    const data = await res.json();
    if (data.ok) {
      closeModal('modal-edit-product');
      adminToast('✓ Товар обновлён', 'success');
      setTimeout(() => window.location.reload(), 800);
    } else {
      adminToast('✖ ' + (data.error || 'Ошибка'), 'error');
    }
  } catch(e) {
    adminToast('✖ ' + e, 'error');
  } finally {
    btn.disabled = false;
    btn.textContent = origText;
  }
}

async function uploadImport(ev) {
  ev.preventDefault();
  const form = document.getElementById('import-form');
  const fd = new FormData(form);
  const box = document.getElementById('import-result');
  box.className = '';
  box.textContent = 'Загрузка…';
  try {
    const res = await fetch('/admin/import/upload', { method: 'POST', body: fd });
    const data = await res.json();
    if (data.ok) {
      box.className = 'import-result';
      let html = 'Добавлено ' + data.added + ' товаров. Пропущено ' + data.skipped + '.';
      if (data.errors && data.errors.length) {
        html += '<br><small style="opacity:0.8">' + data.errors.slice(0, 8).join('<br>') + '</small>';
      }
      box.innerHTML = html;
      setTimeout(() => window.location.href = '/admin#tab-products', 1500);
    } else {
      box.className = 'import-result err';
      box.textContent = '✖ ' + (data.error || 'ошибка');
    }
  } catch (e) {
    box.className = 'import-result err';
    box.textContent = '✖ ' + e;
  }
  return false;
}
</script>
</body>
</html>'''


revision_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<title>Ревизия — {{ org.name }}</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent;}
body{font-family:'Outfit',sans-serif;background:linear-gradient(135deg,#13111C 0%,#1d1635 50%,#231b50 100%);background-attachment:fixed;min-height:100vh;color:#fff;padding-bottom:90px;}
header{background:rgba(255,255,255,0.05);backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);border-bottom:1px solid rgba(255,255,255,0.08);padding:14px 20px;display:flex;justify-content:space-between;align-items:center;position:sticky;top:0;z-index:100;}
header h1{font-size:18px;font-weight:700;background:linear-gradient(135deg,#7c6cf0,#a855f7);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
header p{font-size:12px;color:rgba(255,255,255,0.4);margin-top:2px;}
.hbtns{display:flex;gap:8px;}
.hbtn{padding:8px 12px;border:none;border-radius:10px;font-size:13px;font-weight:600;cursor:pointer;font-family:'Outfit',sans-serif;text-decoration:none;display:inline-block;}
.hbtn-outline{background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.12);color:rgba(255,255,255,0.8);}
.hbtn-primary{background:linear-gradient(135deg,#7c6cf0,#a855f7);color:#fff;box-shadow:0 4px 15px rgba(124,108,240,0.3);}
.tabs{display:flex;overflow-x:auto;padding:14px 20px;gap:10px;scrollbar-width:none;}
.tabs::-webkit-scrollbar{display:none;}
.tab{padding:8px 18px;border-radius:50px;font-weight:600;color:rgba(255,255,255,0.5);text-decoration:none;white-space:nowrap;font-size:14px;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.1);transition:all 0.2s;}
.tab.active{background:linear-gradient(135deg,rgba(124,108,240,0.5),rgba(168,85,247,0.4));border-color:rgba(124,108,240,0.5);color:#fff;box-shadow:0 4px 15px rgba(124,108,240,0.25);}
.container{padding:0 16px;}
.progress-wrap{margin:14px 0 4px;padding:14px 16px;background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:14px;}
.progress-info{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;font-size:13px;color:rgba(255,255,255,0.7);}
.progress-text b{color:#fff;font-weight:700;}
.progress-pct{font-weight:700;color:#a855f7;font-size:15px;}
.progress-bar{height:6px;background:rgba(255,255,255,0.08);border-radius:99px;overflow:hidden;}
.progress-fill{height:100%;background:linear-gradient(90deg,#a855f7,#7c6cf0);border-radius:99px;transition:width .4s ease;box-shadow:0 0 8px rgba(124,108,240,0.5);}
.search-wrap{position:sticky;top:60px;z-index:90;background:transparent;padding:10px 0;}
.search-input{width:100%;padding:12px 16px;background:rgba(255,255,255,0.07);border:1px solid rgba(255,255,255,0.1);border-radius:14px;color:#fff;font-size:15px;font-family:'Outfit',sans-serif;}
.search-input::placeholder{color:rgba(255,255,255,0.3);}
.search-input:focus{outline:none;border-color:rgba(124,108,240,0.6);}
.cat-label{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:rgba(255,255,255,0.4);margin:18px 0 8px;}
.product-item{background:rgba(255,255,255,0.07);border:1px solid rgba(255,255,255,0.1);border-radius:14px;padding:14px 16px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center;cursor:pointer;transition:all 0.15s;}
.product-item:active{transform:scale(0.98);}
.product-item:hover{background:rgba(255,255,255,0.1);}
.p-name{font-weight:500;font-size:15px;color:#fff;}
.p-meta{font-size:12px;color:rgba(255,255,255,0.4);margin-top:3px;}
.badge{padding:5px 10px;border-radius:8px;font-size:13px;font-weight:700;min-width:36px;text-align:center;}
.badge-neutral{background:rgba(124,108,240,0.3);color:#a5b4fc;}
.badge-green{background:rgba(16,185,129,0.25);color:#6ee7b7;}
.badge-yellow{background:rgba(251,191,36,0.25);color:#fcd34d;}
.badge-red{background:rgba(239,68,68,0.25);color:#fca5a5;}
.empty-msg{text-align:center;color:rgba(255,255,255,0.5);padding:48px 20px;font-size:14px;border:1px dashed rgba(255,255,255,0.12);border-radius:18px;margin:24px auto;max-width:420px;}
.empty-msg .ico{font-size:52px;display:block;margin-bottom:14px;opacity:0.7;}
.empty-msg .ttl{color:rgba(255,255,255,0.85);font-size:17px;font-weight:600;margin-bottom:6px;}
.empty-msg .hint{font-size:13px;color:rgba(255,255,255,0.5);}
.finish-btn{position:fixed;bottom:20px;left:20px;right:20px;background:linear-gradient(135deg,#7c6cf0,#a855f7);color:#fff;border:none;padding:16px;border-radius:16px;font-size:16px;font-weight:700;cursor:pointer;font-family:'Outfit',sans-serif;box-shadow:0 8px 24px rgba(124,108,240,0.45);z-index:90;}
.finish-btn:disabled{opacity:0.5;cursor:not-allowed;}
/* Modal */
.modal{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(10,8,20,0.7);backdrop-filter:blur(6px);z-index:1000;align-items:flex-end;}
.modal.active{display:flex;animation:fadein 0.2s;}
@keyframes fadein{from{opacity:0;}to{opacity:1;}}
.modal-sheet{background:rgba(20,16,38,0.97);border:1px solid rgba(255,255,255,0.1);width:100%;border-radius:24px 24px 0 0;padding:24px;box-shadow:0 -10px 40px rgba(0,0,0,0.5);animation:slideup 0.3s cubic-bezier(0.16,1,0.3,1);}
@keyframes slideup{from{transform:translateY(100%);}to{transform:translateY(0);}}
.modal-center{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(10,8,20,0.7);backdrop-filter:blur(6px);z-index:1000;align-items:center;justify-content:center;padding:24px;}
.modal-center.active{display:flex;}
.modal-box{background:rgba(20,16,38,0.97);border:1px solid rgba(255,255,255,0.1);border-radius:24px;padding:32px 24px;width:100%;max-width:360px;text-align:center;box-shadow:0 20px 50px rgba(0,0,0,0.6);animation:popIn 0.3s cubic-bezier(0.16,1,0.3,1);}
@keyframes popIn{from{opacity:0;transform:scale(0.9);}to{opacity:1;transform:scale(1);}}
.modal-box h2{color:#fff;font-size:20px;margin-bottom:8px;}
.modal-box p{color:rgba(255,255,255,0.5);font-size:14px;margin-bottom:24px;}
.modal-box .icon{font-size:42px;margin-bottom:12px;}
.modal-btns{display:flex;gap:10px;}
.mbtn{flex:1;padding:13px;border-radius:12px;font-size:15px;font-weight:600;cursor:pointer;font-family:'Outfit',sans-serif;border:none;}
.mbtn-no{background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.12);color:rgba(255,255,255,0.7);}
.mbtn-yes{background:linear-gradient(135deg,#7c6cf0,#a855f7);color:#fff;box-shadow:0 4px 15px rgba(124,108,240,0.4);}
.mbtn-ok{width:100%;background:linear-gradient(135deg,#7c6cf0,#a855f7);color:#fff;box-shadow:0 4px 15px rgba(124,108,240,0.4);}
/* Calculator */
.calc-header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:16px;}
.calc-title{font-size:16px;font-weight:700;color:#fff;max-width:80%;}
.calc-close{background:rgba(255,255,255,0.1);border:none;color:#fff;width:32px;height:32px;border-radius:8px;cursor:pointer;font-size:16px;}
.calc-display{width:100%;font-size:30px;padding:10px 4px;text-align:right;border:none;border-bottom:2px solid rgba(255,255,255,0.1);color:#a5b4fc;background:transparent;font-family:'Outfit',monospace;margin-bottom:16px;}
.calc-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;}
.cbtn{padding:18px 4px;border-radius:14px;border:none;font-size:22px;font-weight:600;cursor:pointer;font-family:'Outfit',sans-serif;touch-action:manipulation;transition:transform 0.1s, background 0.1s;}
.cbtn:active{transform:scale(0.94);}
.cbtn-num{background:rgba(255,255,255,0.08);color:#fff;}
.cbtn-num:active{background:rgba(255,255,255,0.18);}
.cbtn-op{background:rgba(124,108,240,0.22);color:#a5b4fc;}
.cbtn-op:active{background:rgba(124,108,240,0.4);}
.cbtn-add{grid-column:span 2;background:rgba(168,85,247,0.25);color:#e9d5ff;font-weight:700;font-size:16px;letter-spacing:0.5px;}
.cbtn-add:active{background:rgba(168,85,247,0.45);}
.cbtn-save{grid-column:span 2;background:linear-gradient(135deg,#7c6cf0,#a855f7);color:#fff;font-weight:700;font-size:16px;box-shadow:0 4px 15px rgba(124,108,240,0.4);letter-spacing:0.5px;}
@media (max-width:430px){.cbtn{padding:22px 4px;font-size:24px;}.cbtn-add,.cbtn-save{font-size:17px;padding:20px 4px;}}
.total-row{margin-top:14px;text-align:center;font-size:15px;color:rgba(255,255,255,0.6);}
.total-num{color:#a5b4fc;font-weight:700;font-size:18px;}
/* Toast */
.toast{position:fixed;left:50%;bottom:90px;transform:translateX(-50%) translateY(20px);padding:12px 20px;border-radius:12px;font-size:14px;font-weight:600;color:#fff;z-index:9999;pointer-events:none;opacity:0;transition:all 0.25s ease-out;backdrop-filter:blur(12px);box-shadow:0 8px 30px rgba(0,0,0,0.3);}
.toast.active{opacity:1;transform:translateX(-50%) translateY(0);}
.toast-success{background:rgba(34,197,94,0.85);}
.toast-error{background:rgba(239,68,68,0.9);}
.toast-info{background:rgba(124,108,240,0.85);}
/* Prompt input in modals */
.prompt-input{width:100%;padding:14px;font-size:18px;border-radius:10px;border:1px solid rgba(255,255,255,0.15);background:rgba(255,255,255,0.06);color:#fff;text-align:center;font-family:'Outfit',sans-serif;margin:8px 0 4px;outline:none;}
.prompt-input:focus{border-color:#7c6cf0;background:rgba(124,108,240,0.1);}
.hist-log{margin-top:12px;border-top:1px solid rgba(255,255,255,0.08);padding-top:10px;}
.hist-label{font-size:11px;color:rgba(255,255,255,0.35);margin-bottom:6px;}
.hist-scroll{max-height:80px;overflow-y:auto;}
.hist-item{display:flex;justify-content:space-between;align-items:center;padding:4px 0;border-bottom:1px solid rgba(255,255,255,0.05);font-size:11px;color:rgba(255,255,255,0.5);}
.hist-item:last-child{border-bottom:none;}
.hist-del{background:none;border:none;color:rgba(239,68,68,0.6);cursor:pointer;font-size:14px;padding:0 4px;}
.hist-edit{background:none;border:none;color:rgba(124,108,240,0.75);cursor:pointer;font-size:13px;padding:0 4px;}
.hist-edit:hover{color:rgba(124,108,240,1);}
.hist-actions{display:inline-flex;gap:2px;}
</style>
</head>
<body>
<header>
  <div>
    <h1 style="display:inline-flex;align-items:center;gap:8px;">{{ icon('shop', 20)|safe }} {{ org.name }}</h1>
    <p>Ревизия · {{ username }}</p>
  </div>
  <div class="hbtns">
    {% if is_admin %}<a href="/admin" class="hbtn hbtn-outline" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('sliders', 14)|safe }} Админ</a>{% endif %}
    <a href="/logout" class="hbtn hbtn-outline" onclick="return confirm('Выйти из аккаунта?');" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('logout', 14)|safe }} Выйти</a>
  </div>
</header>

{% if not locations %}
  <div class="empty-msg">
    <span class="ico">{{ icon('mappin', 38)|safe }}</span>
    <div class="ttl">Нет локаций</div>
    <div class="hint">Попросите администратора добавить локации (склад, кухня и т.д.).</div>
  </div>
{% else %}
<div class="tabs">
  {% for loc in locations %}
  <a href="/revision?location={{ loc.name|urlencode }}" class="tab {% if selected and selected.id == loc.id %}active{% endif %}">{{ loc.name }}</a>
  {% endfor %}
</div>

{% if selected %}
<div class="container">
  {% if total_products > 0 %}
  <div class="progress-wrap">
    <div class="progress-info">
      <span class="progress-text">Посчитано <b>{{ counted_products }}</b> из <b>{{ total_products }}</b></span>
      <span class="progress-pct">{{ progress_pct }}%</span>
    </div>
    <div class="progress-bar"><div class="progress-fill" style="width:{{ progress_pct }}%"></div></div>
  </div>
  {% endif %}
  <div class="search-wrap">
    <input id="search" class="search-input" type="text" placeholder="Поиск товара…" oninput="filterProducts()">
  </div>

  {% if not grouped %}
    <div class="empty-msg">
      <span class="ico">{{ icon('box', 38)|safe }}</span>
      <div class="ttl">Нет товаров</div>
      <div class="hint">Добавьте товары в админ-панели или загрузите Excel-файл.</div>
    </div>
  {% else %}
  <div id="productList">
  {% for cat, prods in grouped %}
    <div class="product-group">
      <div class="cat-label">{{ cat.name if cat else 'Без категории' }}</div>
      {% for p in prods %}
        {% set qty = qty_map.get(p.id, 0) %}
        {% set norm = norms_map.get(p.id) %}
        {% if norm and norm > 0 %}
          {% if qty >= norm %}
            {% set badge_cls = 'badge-green' %}
          {% elif qty >= norm * 0.5 %}
            {% set badge_cls = 'badge-yellow' %}
          {% else %}
            {% set badge_cls = 'badge-red' %}
          {% endif %}
          {% set badge_text = qty|qty + ' / ' + norm|int|string %}
        {% else %}
          {% set badge_cls = 'badge-neutral' %}
          {% set badge_text = qty|qty if qty > 0 else '' %}
        {% endif %}
        <div class="product-item" data-name="{{ p.name|lower }}"
          onclick="openCalc({{ p.id }}, {{ selected.id }}, '{{ p.name|replace("'","\\'")|e }}', '{{ p.unit }}')">
          <div>
            <div class="p-name">{{ p.name }}</div>
            <div class="p-meta">{{ p.unit }}{% if norm and norm > 0 %} · норма {{ norm|int }}{% endif %}</div>
          </div>
          {% if qty > 0 or (norm and norm > 0) %}
          <div class="badge {{ badge_cls }}">{{ badge_text if badge_text else '0' }}</div>
          {% endif %}
        </div>
      {% endfor %}
    </div>
  {% endfor %}
  </div>
  {% endif %}
</div>

{% if rev_status == 'pending' %}
  <button class="finish-btn" disabled>⏳ Ожидание подтверждения...</button>
{% else %}
  <button class="finish-btn" onclick="requestFinish()">Завершить ревизию</button>
{% endif %}
{% endif %}
{% endif %}

<!-- Calculator modal -->
<div class="modal" id="calcModal" onclick="if(event.target===this)closeCalc()">
<div class="modal-sheet">
  <div class="calc-header">
    <div class="calc-title" id="calcTitle"></div>
    <button class="calc-close" onclick="closeCalc()">✕</button>
  </div>
  <input id="calcDisplay" class="calc-display" readonly value="0">
  <div class="calc-grid">
    <button class="cbtn cbtn-num" onclick="num('7')">7</button>
    <button class="cbtn cbtn-num" onclick="num('8')">8</button>
    <button class="cbtn cbtn-num" onclick="num('9')">9</button>
    <button class="cbtn cbtn-op" onclick="setOp('/')">÷</button>
    <button class="cbtn cbtn-num" onclick="num('4')">4</button>
    <button class="cbtn cbtn-num" onclick="num('5')">5</button>
    <button class="cbtn cbtn-num" onclick="num('6')">6</button>
    <button class="cbtn cbtn-op" onclick="setOp('*')">×</button>
    <button class="cbtn cbtn-num" onclick="num('1')">1</button>
    <button class="cbtn cbtn-num" onclick="num('2')">2</button>
    <button class="cbtn cbtn-num" onclick="num('3')">3</button>
    <button class="cbtn cbtn-op" onclick="setOp('-')">−</button>
    <button class="cbtn cbtn-num" onclick="num(',')">,</button>
    <button class="cbtn cbtn-num" onclick="num('0')">0</button>
    <button class="cbtn cbtn-op" onclick="clr()">C</button>
    <button class="cbtn cbtn-op" onclick="setOp('+')">+</button>
    <button class="cbtn cbtn-op" onclick="calculate()">=</button>
    <button class="cbtn cbtn-add" onclick="addToTotal()">+ ДОБАВИТЬ</button>
    <button class="cbtn cbtn-save" onclick="closeCalcDone()">✓ ГОТОВО</button>
  </div>
  <div class="total-row">Итого: <span id="total" class="total-num">0</span> <span id="unitLabel"></span></div>
  <div class="hist-log">
    <div class="hist-label">История (текущая сессия):</div>
    <div class="hist-scroll" id="histLog"></div>
  </div>
</div>
</div>

<!-- Confirm finish modal -->
<div class="modal-center" id="confirmModal">
<div class="modal-box">
  <div class="icon">{{ icon('clipboard', 40)|safe }}</div>
  <h2>Завершить ревизию?</h2>
  <p>На проверку уйдут все посчитанные локации одним запросом. После этого добавить позиции будет нельзя, пока админ не подтвердит.</p>
  <div class="modal-btns">
    <button class="mbtn mbtn-no" onclick="cancelFinish()">Нет</button>
    <button class="mbtn mbtn-yes" onclick="confirmFinish()">Да</button>
  </div>
</div>
</div>

<!-- Sent modal -->
<div class="modal-center" id="sentModal">
<div class="modal-box">
  <div class="icon" style="color:#34d399;">{{ icon('check-circle', 40)|safe }}</div>
  <h2>Запрос отправлен</h2>
  <p>Ожидайте подтверждения администратором</p>
  <button class="mbtn mbtn-ok" onclick="closeSentModal()">ОК</button>
</div>
</div>

<script>
let curProdId = null, curLocId = null, curUnit = '';
let val = '0', op = null, prev = null, total = 0;

function filterProducts() {
  const q = document.getElementById('search').value.toLowerCase();
  document.querySelectorAll('.product-item').forEach(el => {
    const vis = el.dataset.name.includes(q);
    el.style.display = vis ? 'flex' : 'none';
  });
  document.querySelectorAll('.product-group').forEach(g => {
    const hasVis = [...g.querySelectorAll('.product-item')].some(e => e.style.display !== 'none');
    g.style.display = hasVis ? 'block' : 'none';
  });
}

async function openCalc(prodId, locId, name, unit) {
  curProdId = prodId; curLocId = locId; curUnit = unit;
  val = '0'; op = null; prev = null; total = 0;
  document.getElementById('calcTitle').innerText = name;
  document.getElementById('unitLabel').innerText = unit;
  document.getElementById('calcDisplay').value = '0';
  document.getElementById('total').innerText = '0';
  await loadHistory(prodId, locId);
  document.getElementById('calcModal').classList.add('active');
}

async function loadHistory(prodId, locId) {
  try {
    const res = await fetch(`/revision/product_history/${prodId}?location_id=${locId}`);
    const data = await res.json();
    const el = document.getElementById('histLog');
    const items = (data && data.items) ? data.items : (Array.isArray(data) ? data : []);
    if (data && typeof data.total === 'number') {
      total = data.total;
      document.getElementById('total').innerText = fmt(total);
    }
    if (!items.length) { el.innerHTML = '<span style="color:rgba(255,255,255,0.3)">Пусто</span>'; return; }
    el.innerHTML = items.map(h =>
      `<div class="hist-item">
        <span>${h.text}</span>
        <span class="hist-actions">
          <button class="hist-edit" onclick="editHistItem(${h.id}, ${h.quantity})" title="Изменить"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:-2px;"><path d="M12 20h9"/><path d="M16.5 3.5a2.12 2.12 0 0 1 3 3L7 19l-4 1 1-4z"/></svg></button>
          <button class="hist-del" onclick="deleteHistItem(${h.id})" title="Удалить">×</button>
        </span>
      </div>`
    ).join('');
  } catch(e) { }
}

function showToast(msg, type) {
  let t = document.getElementById('toast');
  if (!t) {
    t = document.createElement('div');
    t.id = 'toast';
    t.className = 'toast';
    document.body.appendChild(t);
  }
  t.textContent = msg;
  t.className = 'toast toast-' + (type || 'info') + ' active';
  clearTimeout(window._toastTimer);
  window._toastTimer = setTimeout(() => { t.className = 'toast'; }, 1800);
}

function uiConfirm(message) {
  return new Promise(resolve => {
    const overlay = document.createElement('div');
    overlay.className = 'modal-center active';
    overlay.innerHTML = `
      <div class="modal-box">
        <div class="icon"><svg width="38" height="38" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><path d="M9.1 9a3 3 0 0 1 5.8 1c0 2-3 3-3 3"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg></div>
        <h2>${message}</h2>
        <div class="modal-btns">
          <button class="mbtn mbtn-no">Нет</button>
          <button class="mbtn mbtn-yes">Да</button>
        </div>
      </div>`;
    overlay.querySelector('.mbtn-no').onclick = () => { document.body.removeChild(overlay); resolve(false); };
    overlay.querySelector('.mbtn-yes').onclick = () => { document.body.removeChild(overlay); resolve(true); };
    document.body.appendChild(overlay);
  });
}

function uiPrompt(message, defaultVal) {
  return new Promise(resolve => {
    const overlay = document.createElement('div');
    overlay.className = 'modal-center active';
    overlay.innerHTML = `
      <div class="modal-box">
        <div class="icon"><svg width="34" height="34" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M12 20h9"/><path d="M16.5 3.5a2.12 2.12 0 0 1 3 3L7 19l-4 1 1-4z"/></svg></div>
        <h2>${message}</h2>
        <input class="prompt-input" type="text" inputmode="decimal" value="${defaultVal || ''}">
        <div class="modal-btns">
          <button class="mbtn mbtn-no">Отмена</button>
          <button class="mbtn mbtn-yes">OK</button>
        </div>
      </div>`;
    const inp = overlay.querySelector('input');
    overlay.querySelector('.mbtn-no').onclick = () => { document.body.removeChild(overlay); resolve(null); };
    overlay.querySelector('.mbtn-yes').onclick = () => { const v = inp.value; document.body.removeChild(overlay); resolve(v); };
    inp.addEventListener('keydown', (e) => { if (e.key === 'Enter') overlay.querySelector('.mbtn-yes').click(); });
    document.body.appendChild(overlay);
    setTimeout(() => { inp.focus(); inp.select(); }, 50);
  });
}

async function editHistItem(itemId, currentQty) {
  const raw = await uiPrompt('Новое значение:', String(currentQty).replace('.', ','));
  if (raw === null) return;
  const n = parseNum(raw);
  if (isNaN(n) || n <= 0) { showToast('Введите корректное число', 'error'); return; }
  const fd = new FormData();
  fd.append('quantity', round3(n));
  const res = await fetch(`/revision/edit_item/${itemId}`, {method: 'POST', body: fd});
  const data = await res.json();
  if (!data.ok) { showToast('Ошибка: ' + (data.error || ''), 'error'); return; }
  await loadHistory(curProdId, curLocId);
  showToast('Изменено', 'success');
}

async function deleteHistItem(itemId) {
  const ok = await uiConfirm('Удалить запись?');
  if (!ok) return;
  await fetch(`/revision/delete_item/${itemId}`, {method: 'POST'});
  await loadHistory(curProdId, curLocId);
  showToast('Удалено', 'success');
}

function closeCalc() { document.getElementById('calcModal').classList.remove('active'); }

function parseNum(s){ return parseFloat(String(s).replace(',', '.')); }
function round3(n){ return Math.round(n * 1000) / 1000; }
function num(n) {
  // Разрешаем только одну запятую/точку в числе
  if ((n === ',' || n === '.') && (val.indexOf(',') !== -1 || val.indexOf('.') !== -1)) return;
  if(val==='0'||val==='Error') val=n==='.'||n===','?'0,':n; else val+=n;
  document.getElementById('calcDisplay').value=val;
}
function setOp(o) { prev=parseNum(val); val='0'; op=o; }
function calculate() {
  if(op&&prev!=null){
    const cur=parseNum(val); let r;
    switch(op){case'+':r=prev+cur;break;case'-':r=prev-cur;break;case'*':r=prev*cur;break;case'/':r=cur!==0?prev/cur:'Error';break;}
    if(typeof r === 'number') r = round3(r);
    val=(typeof r === 'number' ? r.toString().replace('.', ',') : r); op=null; prev=null; document.getElementById('calcDisplay').value=val;
  }
}
function clr() { val='0'; prev=null; op=null; document.getElementById('calcDisplay').value='0'; }
function fmt(n){ return String(n).replace('.', ','); }

async function addToTotal() {
  // Автосохранение: каждое +ДОБАВИТЬ сразу пишется на сервер
  calculate();
  const n = round3(parseNum(val));
  if(isNaN(n) || n <= 0) { showToast('Введите число', 'error'); return; }
  await sendToServer(n);
  val='0'; document.getElementById('calcDisplay').value='0';
}

async function sendToServer(n) {
  const fd = new FormData();
  fd.append('product_id', curProdId);
  fd.append('location_id', curLocId);
  fd.append('count', n);
  try {
    const res = await fetch('/revision/add', {method:'POST', body:fd});
    const data = await res.json();
    if (data.ok) {
      total = round3((total || 0) + n);
      document.getElementById('total').innerText = fmt(total);
      await loadHistory(curProdId, curLocId);
      showToast('+' + fmt(n) + ' сохранено', 'success');
    } else {
      showToast('Ошибка: ' + (data.error || ''), 'error');
    }
  } catch(e) {
    showToast('Ошибка сети', 'error');
  }
}

async function closeCalcDone() {
  // Если есть несохранённое значение на дисплее — сохраняем перед закрытием
  const n = round3(parseNum(val));
  if (!isNaN(n) && n > 0) {
    await sendToServer(n);
  }
  closeCalc();
  location.reload();
}

function requestFinish() { document.getElementById('confirmModal').classList.add('active'); }
function cancelFinish() { document.getElementById('confirmModal').classList.remove('active'); }
async function confirmFinish() {
  document.getElementById('confirmModal').classList.remove('active');
  const fd = new FormData();
  fd.append('location_id', {{ selected.id if selected else 0 }});
  try {
    const res = await fetch('/revision/finish', {method:'POST', body:fd});
    const data = await res.json();
    if (data.ok) {
      document.getElementById('sentModal').classList.add('active');
    } else {
      alert(data.error || 'Не удалось завершить ревизию');
    }
  } catch(e) {
    alert('Ошибка: ' + e);
  }
}
function closeSentModal() { document.getElementById('sentModal').classList.remove('active'); location.reload(); }
</script>
</body>
</html>'''


# ============== ПОДДЕРЖКА: КЛИЕНТСКИЕ ШАБЛОНЫ ==============
_SUPPORT_CSS = '''
* { box-sizing: border-box; -webkit-tap-highlight-color: transparent; margin: 0; padding: 0; }
body {
  font-family: 'Outfit', sans-serif;
  background: linear-gradient(135deg, #13111C 0%, #1d1635 50%, #231b50 100%);
  background-attachment: fixed; min-height: 100vh; color: white; padding-bottom: 40px;
}
.blob { position: fixed; border-radius: 50%; filter: blur(80px); opacity: 0.25; pointer-events: none; z-index: 0; }
.blob-1 { width: 380px; height: 380px; background: radial-gradient(circle, #7c6cf0, #a855f7); top: -100px; left: -80px; }
.blob-2 { width: 320px; height: 320px; background: radial-gradient(circle, #a855f7, #6d28d9); bottom: -80px; right: -80px; }
.s-header {
  position: sticky; top: 0; z-index: 20;
  background: rgba(19,17,28,0.75); backdrop-filter: blur(20px);
  border-bottom: 1px solid rgba(255,255,255,0.08);
  padding: 14px 20px; display: flex; align-items: center; justify-content: space-between;
}
.s-header .name { font-weight: 700; font-size: 16px; }
.s-header .sub { font-size: 11px; color: rgba(255,255,255,0.5); text-transform: uppercase; letter-spacing: 0.6px; }
.s-back { background: rgba(255,255,255,0.07); border: 1px solid rgba(255,255,255,0.12); color: white; padding: 8px 14px; border-radius: 10px; font-size: 13px; font-weight: 600; text-decoration: none; }
.s-back:hover { background: rgba(255,255,255,0.12); }
.container { max-width: 760px; margin: 22px auto; padding: 0 16px; position: relative; z-index: 1; }
.card { background: rgba(255,255,255,0.07); border: 1px solid rgba(255,255,255,0.12); border-radius: 18px; padding: 20px; margin-bottom: 14px; backdrop-filter: blur(20px); }
.card h1 { font-size: 20px; font-weight: 700; margin-bottom: 8px; }
.card h2 { font-size: 16px; font-weight: 700; margin-bottom: 10px; }
.muted { color: rgba(255,255,255,0.55); font-size: 13px; }
.input, .area { width: 100%; background: rgba(255,255,255,0.07); border: 1px solid rgba(255,255,255,0.12); border-radius: 12px; color: white; padding: 11px 14px; font-family: 'Outfit', sans-serif; font-size: 14px; margin-bottom: 10px; }
.area { min-height: 140px; resize: vertical; }
.input:focus, .area:focus { outline: none; border-color: rgba(124,108,240,0.6); }
.btn-primary { background: linear-gradient(135deg, #7c6cf0, #a855f7); border: none; color: white; padding: 12px 18px; border-radius: 12px; font-weight: 600; font-family: 'Outfit', sans-serif; font-size: 14px; cursor: pointer; }
.btn-primary:hover { filter: brightness(1.08); }
.flash { padding: 12px 14px; border-radius: 12px; margin-bottom: 12px; font-size: 13px; }
.flash-success { background: rgba(34,197,94,0.14); color: #86efac; border: 1px solid rgba(34,197,94,0.3); }
.flash-error { background: rgba(239,68,68,0.14); color: #fca5a5; border: 1px solid rgba(239,68,68,0.3); }
.tlist a { display: flex; justify-content: space-between; align-items: center; padding: 14px 16px; background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08); border-radius: 12px; margin-bottom: 8px; text-decoration: none; color: white; }
.tlist a:hover { background: rgba(255,255,255,0.08); }
.tstat { font-size: 11px; font-weight: 700; padding: 3px 10px; border-radius: 999px; text-transform: uppercase; letter-spacing: 0.5px; }
.tstat-open { background: rgba(239,68,68,0.2); color: #fca5a5; }
.tstat-answered { background: rgba(124,108,240,0.2); color: #a78bfa; }
.tstat-closed { background: rgba(100,116,139,0.2); color: #94a3b8; }
.empty { padding: 40px 20px; text-align: center; color: rgba(255,255,255,0.5); border: 1px dashed rgba(255,255,255,0.12); border-radius: 14px; }
.empty .ico { font-size: 42px; opacity: 0.55; margin-bottom: 10px; }
.reply { padding: 14px 16px; border-radius: 14px; margin-bottom: 10px; }
.reply-mine { background: rgba(16,185,129,0.08); border-left: 3px solid #10b981; }
.reply-them { background: rgba(168,85,247,0.08); border-left: 3px solid #a855f7; }
.reply-meta { font-size: 11px; color: rgba(255,255,255,0.55); margin-bottom: 6px; }
.reply-body { white-space: pre-wrap; font-size: 14px; line-height: 1.55; }
'''


support_list_html = '''<!DOCTYPE html>
<html lang="ru"><head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Поддержка — {{ org.name }}</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _SUPPORT_CSS + '''</style>
</head><body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>

<div class="s-header">
  <div>
    <div class="name" style="display:flex;align-items:center;gap:6px;">{{ icon('ticket', 16)|safe }} Поддержка</div>
    <div class="sub">{{ org.name }}</div>
  </div>
  <a class="s-back" href="/admin">← В админ-панель</a>
</div>

<div class="container">

  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="flash flash-{{ cat }}">{{ msg }}</div>
    {% endfor %}
  {% endwith %}

  <div class="card">
    <h1>Создать обращение</h1>
    <p class="muted" style="margin-bottom:14px;">Опишите проблему или вопрос — владелец Revisi ответит лично.</p>
    <form method="post">
      <input class="input" type="text" name="subject" placeholder="Тема обращения" required maxlength="300">
      <textarea class="area" name="body" placeholder="Подробное описание..." required></textarea>
      <button class="btn-primary" type="submit" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('send', 15)|safe }} Отправить</button>
    </form>
  </div>

  <div class="card">
    <h2>Ваши обращения</h2>
    {% if items %}
      <div class="tlist" style="margin-top:12px;">
      {% for t in items %}
        <a href="/support/{{ t.id }}">
          <div style="flex:1;min-width:0;">
            <div style="font-weight:600;">{{ t.subject }}</div>
            <div class="muted" style="font-size:12px;">обновлено {{ t.when }}</div>
          </div>
          <span class="tstat tstat-{{ t.status }}">{{ t.status }}</span>
        </a>
      {% endfor %}
      </div>
    {% else %}
      <div class="empty">
        <div class="ico">{{ icon('inbox', 34)|safe }}</div>
        Здесь появятся ваши обращения. Создайте первое выше.
      </div>
    {% endif %}
  </div>

</div>
</body></html>'''


support_view_html = '''<!DOCTYPE html>
<html lang="ru"><head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ t.subject }} — Поддержка</title>
{{ pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _SUPPORT_CSS + '''</style>
</head><body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>

<div class="s-header">
  <div>
    <div class="name" style="display:flex;align-items:center;gap:6px;">{{ icon('ticket', 16)|safe }} Тикет #{{ t.id }}</div>
    <div class="sub">{{ org.name }}</div>
  </div>
  <a class="s-back" href="/support">← К списку</a>
</div>

<div class="container">

  <div class="card">
    <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:10px;flex-wrap:wrap;">
      <h1 style="flex:1;min-width:0;">{{ t.subject }}</h1>
      <span class="tstat tstat-{{ t.status }}">{{ t.status }}</span>
    </div>
    <div class="muted" style="margin-top:6px;">Создано {{ t.created_at.strftime('%d.%m.%Y %H:%M') if t.created_at else '' }}</div>
    <div style="margin-top:16px;white-space:pre-wrap;font-size:14px;line-height:1.55;color:rgba(255,255,255,0.9);">{{ t.body }}</div>
  </div>

  {% for r in replies %}
  <div class="reply {{ 'reply-mine' if r.author_type == 'user' else 'reply-them' }}">
    <div class="reply-meta">
      <b>{{ r.author_label or ('Вы' if r.author_type == 'user' else 'Поддержка') }}</b>
      · {{ r.created_at.strftime('%d.%m.%Y %H:%M') if r.created_at else '' }}
    </div>
    <div class="reply-body">{{ r.body }}</div>
  </div>
  {% endfor %}

  {% if t.status != 'closed' %}
  <div class="card">
    <h2>Дополнить</h2>
    <form method="post">
      <textarea class="area" name="body" placeholder="Добавить детали..." required></textarea>
      <button class="btn-primary" type="submit" style="display:inline-flex;align-items:center;gap:6px;">{{ icon('send', 15)|safe }} Отправить</button>
    </form>
  </div>
  {% else %}
    <div class="empty">Тикет закрыт. Если вопрос остался — создайте новый.</div>
  {% endif %}

</div>
</body></html>'''


# ============== СТАРЫЕ ШАБЛОНЫ (для следующих фаз миграции) ==============
# Сохранены как есть — НЕ используются до миграции /admin и /revision на БД.

legacy_login_html = '''<!-- Старый login-шаблон сохранён в истории git -->'''

legacy_revision_html = '''<!-- Старый revision-шаблон (см. git) -->'''

legacy_admin_html = '''<!-- Старый admin-шаблон (см. git) -->'''


# ============== OWNER PANEL HTML ==============

_OWNER_BASE_CSS = '''
* { box-sizing: border-box; margin: 0; padding: 0; -webkit-tap-highlight-color: transparent; }
body {
    font-family: 'Outfit', sans-serif;
    background: linear-gradient(135deg, #13111C 0%, #1d1635 50%, #231b50 100%);
    background-attachment: fixed;
    min-height: 100vh;
    color: white;
}
.blob { position: fixed; border-radius: 50%; filter: blur(80px); opacity: 0.35; pointer-events: none; z-index: 0; }
.blob-1 { width: 400px; height: 400px; background: radial-gradient(circle, #7c6cf0, #a855f7); top: -100px; left: -100px; }
.blob-2 { width: 350px; height: 350px; background: radial-gradient(circle, #a855f7, #6d28d9); bottom: -80px; right: -80px; }
.glass {
    background: rgba(255,255,255,0.07);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 16px;
}
.owner-header {
    position: sticky; top: 0; z-index: 100;
    background: rgba(19,17,28,0.85);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border-bottom: 1px solid rgba(255,255,255,0.08);
    padding: 0 24px;
    display: flex; align-items: center; justify-content: space-between;
    height: 60px;
}
.owner-header .brand { font-weight: 700; font-size: 18px; }
.owner-header .brand span { background: linear-gradient(135deg, #7c6cf0, #a855f7); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
.owner-header .right { display: flex; align-items: center; gap: 16px; }
.owner-email { font-size: 13px; color: rgba(255,255,255,0.5); }
.btn-logout {
    background: rgba(239,68,68,0.15); color: #fca5a5;
    border: 1px solid rgba(239,68,68,0.25); border-radius: 10px;
    padding: 7px 14px; font-size: 13px; font-family: 'Outfit', sans-serif;
    cursor: pointer; text-decoration: none; font-weight: 600;
}
.btn-logout:hover { background: rgba(239,68,68,0.25); }
.owner-nav {
    display: flex; gap: 4px; padding: 16px 24px 0;
}
.nav-tab {
    display: inline-flex; align-items: center; gap: 7px;
    padding: 9px 18px; border-radius: 12px; font-size: 14px; font-weight: 600;
    text-decoration: none; color: rgba(255,255,255,0.55);
    transition: all 0.2s;
}
.nav-tab svg { opacity: 0.78; }
.nav-tab.active svg { opacity: 1; }
.nav-tab:hover { background: rgba(255,255,255,0.07); color: white; }
.nav-tab.active {
    background: rgba(124,108,240,0.2); color: #a78bfa;
    border: 1px solid rgba(124,108,240,0.3);
}
.page-content { padding: 20px 24px 40px; max-width: 1200px; margin: 0 auto; position: relative; z-index: 1; }
.metrics-grid {
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 24px;
}
@media(max-width: 768px) { .metrics-grid { grid-template-columns: repeat(2, 1fr); } }
@media(max-width: 480px) { .metrics-grid { grid-template-columns: repeat(2, 1fr); } }
.metric-card {
    background: rgba(255,255,255,0.07);
    backdrop-filter: blur(20px); -webkit-backdrop-filter: blur(20px);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 16px; padding: 20px;
}
.metric-num { font-size: 32px; font-weight: 700; color: white; line-height: 1; }
.metric-label { font-size: 13px; color: rgba(255,255,255,0.5); margin-top: 6px; }
.section-title { font-size: 15px; font-weight: 700; margin-bottom: 12px; color: rgba(255,255,255,0.85); }
.section-card {
    background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 16px; padding: 20px; margin-bottom: 20px;
}
.expiring-row {
    display: flex; align-items: center; justify-content: space-between;
    padding: 10px 0; border-bottom: 1px solid rgba(255,255,255,0.06);
    font-size: 14px;
}
.expiring-row:last-child { border-bottom: none; }
.expiring-name { font-weight: 600; }
.expiring-email { color: rgba(255,255,255,0.5); font-size: 12px; }
.days-badge {
    background: rgba(251,146,60,0.2); color: #fb923c;
    border: 1px solid rgba(251,146,60,0.3);
    border-radius: 8px; padding: 2px 8px; font-size: 12px; font-weight: 600;
    margin-right: 10px;
}
.btn-sm {
    padding: 6px 12px; border-radius: 8px; font-size: 12px; font-weight: 600;
    font-family: 'Outfit', sans-serif; cursor: pointer; border: none; text-decoration: none;
    display: inline-block;
}
.btn-extend { background: rgba(124,108,240,0.25); color: #a78bfa; border: 1px solid rgba(124,108,240,0.3); }
.btn-extend:hover { background: rgba(124,108,240,0.4); }
.btn-pro { background: rgba(16,185,129,0.2); color: #6ee7b7; border: 1px solid rgba(16,185,129,0.3); }
.btn-pro:hover { background: rgba(16,185,129,0.35); }
.btn-free { background: rgba(100,116,139,0.2); color: #94a3b8; border: 1px solid rgba(100,116,139,0.3); }
.btn-free:hover { background: rgba(100,116,139,0.35); }
.btn-danger { background: rgba(239,68,68,0.15); color: #fca5a5; border: 1px solid rgba(239,68,68,0.25); }
.btn-danger:hover { background: rgba(239,68,68,0.3); }
.btn-warn { background: rgba(251,191,36,0.15); color: #fcd34d; border: 1px solid rgba(251,191,36,0.25); }
.btn-warn:hover { background: rgba(251,191,36,0.3); }
.badge {
    display: inline-block; padding: 3px 10px; border-radius: 8px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.5px; text-transform: uppercase;
}
.badge-free { background: rgba(100,116,139,0.3); color: #94a3b8; }
.badge-trial { background: rgba(124,108,240,0.3); color: #a5b4fc; }
.badge-pro { background: rgba(16,185,129,0.3); color: #6ee7b7; }
.badge-business { background: rgba(251,191,36,0.3); color: #fcd34d; }
.recent-table { width: 100%; border-collapse: collapse; font-size: 14px; }
.recent-table th {
    text-align: left; padding: 8px 12px; font-size: 11px;
    color: rgba(255,255,255,0.4); text-transform: uppercase; letter-spacing: 0.8px;
    border-bottom: 1px solid rgba(255,255,255,0.08);
}
.recent-table td { padding: 10px 12px; border-bottom: 1px solid rgba(255,255,255,0.05); }
.recent-table tr:last-child td { border-bottom: none; }
.flash-messages { position: relative; z-index: 2; padding: 12px 24px 0; }
.flash-msg {
    padding: 12px 16px; border-radius: 12px; margin-bottom: 8px; font-size: 14px; font-weight: 600;
}
.flash-success { background: rgba(16,185,129,0.15); color: #6ee7b7; border: 1px solid rgba(16,185,129,0.3); }
.flash-error { background: rgba(239,68,68,0.12); color: #fca5a5; border: 1px solid rgba(239,68,68,0.3); }
.status-dot { display: inline-block; width: 8px; height: 8px; border-radius: 50%; margin-right: 5px; }
.dot-green { background: #10b981; }
.dot-red { background: #ef4444; }
.search-box {
    width: 100%; background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.12); border-radius: 12px;
    color: white; padding: 12px 16px;
    font-family: 'Outfit', sans-serif; font-size: 14px; margin-bottom: 16px;
}
.search-box::placeholder { color: rgba(255,255,255,0.3); }
.search-box:focus { outline: none; border-color: rgba(124,108,240,0.5); background: rgba(124,108,240,0.1); }
.orgs-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.orgs-table th {
    text-align: left; padding: 10px 12px; font-size: 11px;
    color: rgba(255,255,255,0.4); text-transform: uppercase; letter-spacing: 0.8px;
    border-bottom: 1px solid rgba(255,255,255,0.1); white-space: nowrap;
}
.orgs-table td { padding: 10px 12px; border-bottom: 1px solid rgba(255,255,255,0.05); vertical-align: middle; }
.orgs-table tr:last-child td { border-bottom: none; }
.orgs-table tr:hover td { background: rgba(255,255,255,0.03); }
.actions-cell { display: flex; gap: 4px; flex-wrap: wrap; }
@media(max-width: 900px) {
    .orgs-table { display: none; }
    .mobile-cards { display: block; }
}
@media(min-width: 901px) {
    .mobile-cards { display: none; }
}
.mobile-card {
    background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 16px; padding: 16px; margin-bottom: 12px;
}
.mobile-card .org-name { font-size: 15px; font-weight: 700; margin-bottom: 4px; }
.mobile-card .org-email { font-size: 12px; color: rgba(255,255,255,0.5); margin-bottom: 10px; }
.org-card-link { display: block; color: inherit; text-decoration: none; padding: 2px 0 6px; margin: -2px 0 4px; border-radius: 8px; }
.org-card-link:active { background: rgba(124,108,240,0.08); }
.card-chev { float: right; color: rgba(255,255,255,0.3); font-weight: 400; font-size: 18px; }
.mobile-card .details { font-size: 12px; color: rgba(255,255,255,0.6); margin-bottom: 8px; display: flex; gap: 12px; flex-wrap: wrap; }
.mobile-card .actions { display: flex; gap: 6px; flex-wrap: wrap; margin-top: 8px; }
.owner-nav { flex-wrap: wrap; row-gap: 6px; }
.metrics-grid-8 { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin-bottom: 24px; }
@media(max-width: 768px) { .metrics-grid-8 { grid-template-columns: repeat(2, 1fr); } }
.metric-trend { font-size: 11px; color: #6ee7b7; margin-top: 4px; font-weight: 600; }
.metric-num.small { font-size: 24px; }
.charts-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }
@media(max-width: 900px) { .charts-grid { grid-template-columns: 1fr; } }
.chart-title { font-size: 13px; color: rgba(255,255,255,0.7); margin-bottom: 8px; font-weight: 600; }
.chart-total { font-size: 24px; font-weight: 700; color: #fff; margin-bottom: 8px; }
.spark { display: block; width: 100%; height: 60px; }
.alert-section { background: rgba(239,68,68,0.07); border: 1px solid rgba(239,68,68,0.2); border-radius: 14px; padding: 14px 16px; margin-bottom: 12px; }
.alert-section.warn { background: rgba(251,146,60,0.07); border-color: rgba(251,146,60,0.2); }
.alert-section.info { background: rgba(59,130,246,0.07); border-color: rgba(59,130,246,0.2); }
.alert-section h3 { font-size: 14px; margin-bottom: 8px; color: rgba(255,255,255,0.85); }
.alert-row { display: flex; justify-content: space-between; align-items: center; padding: 8px 0; border-bottom: 1px solid rgba(255,255,255,0.05); font-size: 13px; }
.alert-row:last-child { border-bottom: none; }
.audit-table { width: 100%; border-collapse: collapse; font-size: 12px; }
.audit-table th { text-align: left; padding: 8px 10px; color: rgba(255,255,255,0.4); font-size: 10px; text-transform: uppercase; letter-spacing: 0.6px; border-bottom: 1px solid rgba(255,255,255,0.08); }
.audit-table td { padding: 8px 10px; border-bottom: 1px solid rgba(255,255,255,0.05); }
.audit-table .action-pill { display: inline-block; padding: 2px 8px; border-radius: 6px; font-size: 11px; font-weight: 600; background: rgba(124,108,240,0.2); color: #a78bfa; }
.ticket-row { display: flex; justify-content: space-between; align-items: center; padding: 14px 16px; background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08); border-radius: 12px; margin-bottom: 8px; cursor: pointer; text-decoration: none; color: inherit; }
.ticket-row:hover { background: rgba(255,255,255,0.07); }
.ticket-status { font-size: 11px; font-weight: 700; padding: 3px 10px; border-radius: 999px; text-transform: uppercase; letter-spacing: 0.5px; }
.status-open { background: rgba(239,68,68,0.2); color: #fca5a5; }
.status-answered { background: rgba(124,108,240,0.2); color: #a78bfa; }
.status-closed { background: rgba(100,116,139,0.2); color: #94a3b8; }
.feature-toggle { display: flex; align-items: center; justify-content: space-between; padding: 10px 0; border-bottom: 1px solid rgba(255,255,255,0.06); }
.feature-toggle:last-child { border-bottom: none; }
.toggle-switch { position: relative; width: 44px; height: 24px; }
.toggle-switch input { opacity: 0; width: 0; height: 0; }
.toggle-slider { position: absolute; cursor: pointer; inset: 0; background: rgba(255,255,255,0.15); border-radius: 24px; transition: .3s; }
.toggle-slider:before { content: ""; position: absolute; height: 18px; width: 18px; left: 3px; bottom: 3px; background: white; border-radius: 50%; transition: .3s; }
input:checked + .toggle-slider { background: #7c6cf0; }
input:checked + .toggle-slider:before { transform: translateX(20px); }
.input-row { display: flex; gap: 8px; align-items: center; margin-bottom: 12px; }
.input-row input, .input-row select, .input-row textarea {
    background: rgba(255,255,255,0.07); border: 1px solid rgba(255,255,255,0.12);
    border-radius: 10px; color: white; padding: 10px 12px;
    font-family: 'Outfit', sans-serif; font-size: 14px;
}
.input-row input:focus, .input-row select:focus, .input-row textarea:focus { outline: none; border-color: rgba(124,108,240,0.5); }
.input-row .grow { flex: 1; }
textarea.broadcast-area { width: 100%; min-height: 200px; background: rgba(255,255,255,0.07); border: 1px solid rgba(255,255,255,0.12); border-radius: 12px; color: white; padding: 12px 14px; font-family: 'Outfit', sans-serif; font-size: 14px; resize: vertical; }
textarea.broadcast-area:focus { outline: none; border-color: rgba(124,108,240,0.5); }
.btn-primary-sm { background: linear-gradient(135deg, #7c6cf0, #a855f7); color: white; border: none; padding: 10px 20px; border-radius: 10px; font-family: 'Outfit', sans-serif; font-size: 14px; font-weight: 600; cursor: pointer; }
.btn-primary-sm:hover { filter: brightness(1.1); }
.code-block { background: rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.08); border-radius: 8px; padding: 10px 12px; font-family: ui-monospace, Menlo, Consolas, monospace; font-size: 12px; color: #c4b5fd; overflow-x: auto; }
.dash-bottom-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
.tbl-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }

/* === Универсальные стили скролла навигации (без скроллбара) === */
.owner-nav { -webkit-overflow-scrolling: touch; }
.owner-nav::-webkit-scrollbar { display: none; }
.owner-nav { scrollbar-width: none; }

/* === НИЖНЯЯ НАВИГАЦИЯ (мобильная) === */
.bottom-nav {
    display: none;
    position: fixed; bottom: 0; left: 0; right: 0;
    background: rgba(13,11,22,0.96);
    backdrop-filter: blur(20px); -webkit-backdrop-filter: blur(20px);
    border-top: 1px solid rgba(255,255,255,0.1);
    padding: 6px 4px calc(6px + env(safe-area-inset-bottom, 0px));
    z-index: 99;
    justify-content: space-around;
    box-shadow: 0 -8px 24px rgba(0,0,0,0.4);
}
.btab {
    flex: 1;
    display: flex; flex-direction: column; align-items: center;
    gap: 2px; padding: 6px 2px; border-radius: 12px;
    background: none; border: none;
    color: rgba(255,255,255,0.55);
    font-family: 'Outfit', sans-serif;
    text-decoration: none; cursor: pointer;
    min-height: 50px;
    transition: color 0.15s;
}
.btab:active { background: rgba(124,108,240,0.1); }
.btab.active { color: #a78bfa; }
.btab .bicon { display: inline-flex; align-items: center; justify-content: center; height: 22px; }
.btab .blabel { font-size: 10px; font-weight: 600; }

/* === BOTTOM SHEET «ЕЩЁ» === */
.more-modal {
    position: fixed; inset: 0; z-index: 100;
    pointer-events: none; visibility: hidden;
}
.more-modal.open { pointer-events: auto; visibility: visible; }
.more-overlay {
    position: absolute; inset: 0;
    background: rgba(0,0,0,0.55);
    opacity: 0;
    transition: opacity 0.25s;
}
.more-modal.open .more-overlay { opacity: 1; }
.more-sheet {
    position: absolute; bottom: 0; left: 0; right: 0;
    background: linear-gradient(180deg, #1d1635 0%, #13111c 100%);
    border-top: 1px solid rgba(255,255,255,0.12);
    border-radius: 24px 24px 0 0;
    padding: 8px 18px calc(20px + env(safe-area-inset-bottom, 0px));
    transform: translateY(100%);
    transition: transform 0.3s cubic-bezier(0.32, 0.72, 0, 1);
    box-shadow: 0 -12px 32px rgba(0,0,0,0.5);
}
.more-modal.open .more-sheet { transform: translateY(0); }
.more-handle {
    width: 44px; height: 4px;
    background: rgba(255,255,255,0.3);
    border-radius: 2px; margin: 4px auto 14px;
}
.more-title {
    font-size: 11px;
    color: rgba(255,255,255,0.5);
    text-transform: uppercase; letter-spacing: 0.8px;
    margin-bottom: 14px; font-weight: 700;
}
.more-grid {
    display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px;
}
.mtab {
    display: flex; flex-direction: column; align-items: center;
    gap: 8px; padding: 18px 8px; border-radius: 16px;
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.08);
    text-decoration: none;
    color: rgba(255,255,255,0.85);
    font-size: 12px; font-weight: 600;
    text-align: center;
    transition: all 0.15s;
}
.mtab:active { transform: scale(0.97); background: rgba(255,255,255,0.1); }
.mtab.active {
    background: rgba(124,108,240,0.22);
    border-color: rgba(124,108,240,0.4);
    color: #a78bfa;
}
.mtab .micon { display: inline-flex; align-items: center; justify-content: center; height: 26px; }

/* === ПЛАНШЕТ (≤900px) === */
@media (max-width: 900px) {
    .page-content { padding: 16px 14px 80px; }
    .owner-header { padding: 0 14px; height: 56px; }
    .owner-header .brand { font-size: 16px; }
    .owner-email { display: none; }
    .btn-logout { padding: 6px 11px; font-size: 12px; }
    .flash-messages { padding: 10px 14px 0; }
    .owner-nav { padding: 12px 12px 0; flex-wrap: nowrap; overflow-x: auto; }
    .nav-tab { padding: 8px 14px; font-size: 13px; white-space: nowrap; flex-shrink: 0; }
    .dash-bottom-grid { grid-template-columns: 1fr; }
    .charts-grid { grid-template-columns: 1fr; }
    .recent-table { font-size: 13px; }
    .recent-table th, .recent-table td { padding: 8px 8px; }
}

/* === ТЕЛЕФОН (≤640px) === */
@media (max-width: 640px) {
    /* Переключение nav: скрываем верхнюю, показываем нижнюю */
    .owner-nav { display: none; }
    .bottom-nav { display: flex; }
    .page-content { padding: 12px 12px 90px; }

    /* Компактнее метрики */
    .metrics-grid, .metrics-grid-8 { grid-template-columns: repeat(2, 1fr); gap: 8px; margin-bottom: 10px; }
    .metric-card { padding: 10px 12px; border-radius: 12px; }
    .metric-num { font-size: 22px; }
    .metric-num.small { font-size: 18px; }
    .metric-label { font-size: 10.5px; margin-top: 2px; line-height: 1.25; }
    /* Компактнее карточки секций */
    .section-card { padding: 12px; border-radius: 14px; margin-bottom: 10px; }
    .section-title { font-size: 13px; margin-bottom: 8px; }
    .charts-grid { gap: 10px; margin-bottom: 10px; }
    .chart-title { font-size: 12px; margin-bottom: 4px; }
    .chart-total { font-size: 18px; margin-bottom: 4px; }
    .spark { height: 40px; }
    .dash-bottom-grid { gap: 10px; }
    /* Таблицы — горизонтальный скролл */
    .audit-table { font-size: 11px; min-width: 560px; }
    .recent-table { min-width: 480px; }
    .orgs-table { min-width: 720px; }
    /* Sticky search */
    .search-box { position: sticky; top: 8px; z-index: 5; backdrop-filter: blur(10px); background: rgba(20,15,40,0.85); }
    /* Алерты: вертикальная компоновка */
    .alert-row { flex-direction: column; align-items: stretch; gap: 6px; padding: 10px 0; }
    .alert-row > div:first-child { width: 100%; }
    .alert-row > :last-child:not(:first-child) { display: flex; gap: 6px; align-items: center; flex-wrap: wrap; }
    .expiring-row { flex-direction: column; align-items: stretch; gap: 8px; padding: 10px 0; }
    .expiring-row > :last-child:not(:first-child) { display: flex; gap: 6px; align-items: center; flex-wrap: wrap; }
    /* Input-row — стек на мобильном */
    .input-row { flex-direction: column; align-items: stretch; gap: 8px; }
    .input-row input, .input-row select, .input-row textarea { width: 100%; }
    .input-row .grow { flex: 1 1 auto; }
    /* Touch targets */
    .btn-sm { padding: 8px 14px; font-size: 13px; min-height: 36px; line-height: 1.1; }
    .btn-primary-sm { padding: 12px 22px; font-size: 14px; min-height: 44px; width: 100%; }
    .nav-tab { padding: 8px 13px; font-size: 13px; }
    .actions-cell { gap: 6px; }
    .actions-cell .btn-sm { min-height: 34px; }
    /* Badges — больше воздуха */
    .badge { font-size: 10px; padding: 3px 8px; }
    /* Form labels — лучше читаются */
    .feature-toggle { padding: 12px 0; }
    .feature-toggle > div:first-child { flex: 1; min-width: 0; padding-right: 10px; }
    /* Тикеты */
    .ticket-row { padding: 12px 14px; }
    /* Текстовые области выше */
    textarea.broadcast-area { min-height: 160px; font-size: 14px; }
    /* H1 в страницах */
    .page-content h1 { font-size: 19px !important; }
}

/* === МАЛЕНЬКИЙ ТЕЛЕФОН (≤380px) === */
@media (max-width: 380px) {
    .metrics-grid, .metrics-grid-8 { grid-template-columns: 1fr; }
    .owner-header { padding: 0 10px; }
    .owner-header .brand { font-size: 15px; }
    .nav-tab { padding: 7px 11px; font-size: 12px; }
    .page-content { padding: 12px 10px 80px; }
    .section-card { padding: 12px; }
    .metric-card { padding: 12px; }
}
'''

_OWNER_PAGE_HEAD = '''<!DOCTYPE html>
<html lang="ru"><head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__TITLE__ — Revisi Owner</title>
{{ owner_pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _OWNER_BASE_CSS + '''</style>
</head><body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<header class="owner-header">
  <div class="brand" style="display:flex;align-items:center;gap:10px;">{{ icon('chart', 22)|safe }} <span>Панель владельца</span></div>
  <div class="right">
    <span class="owner-email">{{ owner_email }}</span>
    <a class="btn-logout" href="/owner/logout">Выйти</a>
  </div>
</header>
{% with messages = get_flashed_messages(with_categories=true) %}
{% if messages %}
<div class="flash-messages">
  {% for cat, msg in messages %}
  <div class="flash-msg flash-{{ cat }}">{{ msg }}</div>
  {% endfor %}
</div>
{% endif %}
{% endwith %}
'''

_OWNER_PAGE_FOOT = '''
</body></html>'''


# Список вкладок для top + bottom nav (иконка = ключ в _ICONS, генерится через _icon())
_OWNER_TABS = [
    ('dashboard', '/owner', 'chart',      'Дашборд'),
    ('orgs', '/owner/orgs', 'building',   'Компании'),
    ('alerts', '/owner/alerts', 'alert',  'Алерты'),
    ('tickets', '/owner/tickets', 'ticket', 'Тикеты'),
    ('audit', '/owner/audit', 'file',     'Аудит'),
    ('broadcast', '/owner/broadcast', 'megaphone', 'Рассылка'),
    ('catalog', '/owner/catalog', 'box',  'Каталог'),
    ('backup', '/owner/backup', 'database', 'Бэкап'),
    ('health', '/owner/health', 'heartpulse', 'Health'),
    ('security', '/owner/security', 'lock', 'Безопасность'),
    ('sql', '/owner/sql', 'terminal',     'SQL'),
]
# Какие 4 главные показываются в нижнем баре, остальные — в "Ещё"
_OWNER_PRIMARY_TABS = ['dashboard', 'orgs', 'alerts', 'tickets']


def _owner_nav(active):
    """Возвращает HTML с верхним и нижним nav + bottom-sheet «Ещё»."""
    primary = set(_OWNER_PRIMARY_TABS)
    # Верхний nav (виден на планшете+ через CSS)
    top = ['<nav class="owner-nav">']
    for key, url, icon_name, label in _OWNER_TABS:
        cls = 'nav-tab active' if key == active else 'nav-tab'
        top.append(f'<a class="{cls}" href="{url}">{_icon(icon_name, 16)} {label}</a>')
    top.append('</nav>')

    # Нижний nav (виден на телефоне)
    bottom = ['<nav class="bottom-nav">']
    for key in _OWNER_PRIMARY_TABS:
        td = next(t for t in _OWNER_TABS if t[0] == key)
        cls = 'btab active' if key == active else 'btab'
        bottom.append(
            f'<a class="{cls}" href="{td[1]}">'
            f'<span class="bicon">{_icon(td[2], 20)}</span>'
            f'<span class="blabel">{td[3]}</span></a>'
        )
    is_more_active = active not in primary
    more_cls = 'btab active' if is_more_active else 'btab'
    bottom.append(
        f'<button class="{more_cls}" type="button" onclick="ownerToggleMore()">'
        f'<span class="bicon">{_icon("menu", 20)}</span><span class="blabel">Ещё</span></button>'
    )
    bottom.append('</nav>')

    # Bottom sheet с остальными вкладками
    sheet = ['<div class="more-modal" id="moreModal">',
             '<div class="more-overlay" onclick="ownerToggleMore()"></div>',
             '<div class="more-sheet">',
             '<div class="more-handle"></div>',
             '<div class="more-title">Ещё</div>',
             '<div class="more-grid">']
    for key, url, icon_name, label in _OWNER_TABS:
        if key in primary:
            continue
        cls = 'mtab active' if key == active else 'mtab'
        sheet.append(
            f'<a class="{cls}" href="{url}">'
            f'<span class="micon">{_icon(icon_name, 22)}</span><span>{label}</span></a>'
        )
    sheet.append('</div></div></div>')

    js = '''<script>
function ownerToggleMore() {
  var m = document.getElementById('moreModal');
  if (m) m.classList.toggle('open');
}
</script>'''
    return ''.join(top) + ''.join(bottom) + ''.join(sheet) + js


owner_login_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Панель владельца — Revisi</title>
{{ owner_pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''
</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="blob blob-3"></div>
<div class="card">
  <div class="icon-box">📊</div>
  <div class="title">Панель владельца</div>
  <div class="subtitle">Вход для супер-администратора системы</div>
  {% if error %}<div class="error">{{ error }}</div>{% endif %}
  <form method="post">
    <div class="form-group">
      <label>Email</label>
      <input class="input" type="email" name="email" required autocomplete="email" placeholder="owner@example.com">
    </div>
    <div class="form-group">
      <label>Пароль</label>
      <input class="input" type="password" name="password" required autocomplete="current-password" placeholder="Введите пароль">
    </div>
    <button class="btn-primary" type="submit">Войти →</button>
  </form>
</div>
</body>
</html>'''


# ============== ШАБЛОНЫ 2FA (login-style — компактная карточка) ==============

owner_2fa_verify_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>2FA — Revisi Owner</title>
{{ owner_pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''
.input-code{letter-spacing:8px;font-size:24px;text-align:center;font-family:'SF Mono','Monaco',monospace;}
.hint{font-size:13px;color:rgba(255,255,255,0.55);margin-top:14px;text-align:center;line-height:1.5;}
</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="blob blob-3"></div>
<div class="card">
  <div class="icon-box">🔐</div>
  <div class="title">Двухфакторная аутентификация</div>
  <div class="subtitle">Введите 6-значный код из приложения</div>
  {% if error %}<div class="error">{{ error }}</div>{% endif %}
  <form method="post">
    <div class="form-group">
      <input class="input input-code" type="text" name="code" required autocomplete="one-time-code"
             inputmode="numeric" pattern="[0-9 ]*" maxlength="9" placeholder="000000" autofocus>
    </div>
    <button class="btn-primary" type="submit">Войти →</button>
  </form>
  <div class="hint">Если потеряли телефон — введите один из recovery-кодов<br>(формат XXXX-XXXX).</div>
</div>
</body>
</html>'''


owner_2fa_setup_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Включение 2FA — Revisi Owner</title>
{{ owner_pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''
.qr-box{background:#fff;border-radius:12px;padding:14px;display:flex;justify-content:center;margin:18px auto;max-width:240px;}
.qr-box svg{width:100%;height:auto;display:block;}
.secret-box{background:rgba(0,0,0,0.3);border:1px dashed rgba(255,255,255,0.18);border-radius:8px;padding:10px 14px;font-family:'SF Mono','Monaco',monospace;font-size:13px;color:#fed7aa;text-align:center;letter-spacing:1px;margin-top:10px;word-break:break-all;}
.steps{font-size:14px;color:rgba(255,255,255,0.75);line-height:1.6;text-align:left;margin:14px 0;}
.steps ol{padding-left:22px;margin:0;}
.steps li{margin-bottom:6px;}
.input-code{letter-spacing:8px;font-size:22px;text-align:center;font-family:'SF Mono','Monaco',monospace;}
.cancel-link{display:block;margin-top:14px;text-align:center;font-size:13px;color:rgba(255,255,255,0.55);}
.cancel-link:hover{color:#fb923c;}
</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="card" style="max-width:480px;">
  <div class="icon-box">🔐</div>
  <div class="title">Включить 2FA</div>
  <div class="subtitle">Защита owner-панели вторым фактором</div>

  <div class="steps">
    <ol>
      <li>Откройте Google Authenticator, Authy, 1Password или Yandex Key</li>
      <li>Отсканируйте QR-код ниже</li>
      <li>Введите 6-значный код из приложения для подтверждения</li>
    </ol>
  </div>

  <div class="qr-box">{{ qr_svg|safe }}</div>

  <div class="hint" style="font-size:12px;color:rgba(255,255,255,0.5);text-align:center;">
    Не видите QR? Введите ключ вручную:
  </div>
  <div class="secret-box">{{ secret }}</div>

  {% if error %}<div class="error" style="margin-top:14px;">{{ error }}</div>{% endif %}
  <form method="post" style="margin-top:18px;">
    <div class="form-group">
      <label>Код из приложения</label>
      <input class="input input-code" type="text" name="code" required autocomplete="one-time-code"
             inputmode="numeric" pattern="[0-9 ]*" maxlength="9" placeholder="000000" autofocus>
    </div>
    <button class="btn-primary" type="submit">Подтвердить и включить</button>
  </form>
  <a class="cancel-link" href="/owner/security">← Отмена</a>
</div>
</body>
</html>'''


owner_2fa_codes_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Recovery-коды — Revisi Owner</title>
{{ owner_pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _BASE_CSS + '''
.codes-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin:18px 0;font-family:'SF Mono','Monaco',monospace;font-size:15px;}
.code-item{background:rgba(0,0,0,0.3);border:1px solid rgba(251,146,60,0.2);border-radius:8px;padding:10px 12px;text-align:center;color:#fed7aa;letter-spacing:1px;}
.warn-box{background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.3);border-radius:10px;padding:14px;font-size:13px;line-height:1.5;color:#fca5a5;text-align:left;margin:14px 0;}
.warn-box strong{color:#ef4444;display:block;margin-bottom:6px;}
.btn-copy{display:inline-block;margin-top:10px;padding:8px 14px;background:rgba(255,255,255,0.08);color:#fff;border:1px solid rgba(255,255,255,0.15);border-radius:8px;font-size:13px;cursor:pointer;font-family:inherit;}
.btn-copy:hover{background:rgba(255,255,255,0.14);}
</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>
<div class="card" style="max-width:520px;">
  <div class="icon-box">📋</div>
  <div class="title">2FA включена ✓</div>
  <div class="subtitle">Сохраните recovery-коды</div>

  <div class="warn-box">
    <strong>⚠️ Внимание — это единственный раз когда вы видите эти коды</strong>
    Каждый код можно использовать <b>один раз</b> вместо TOTP, если потеряете телефон.
    Запишите их в безопасное место (1Password, заметка с шифрованием, бумажка в сейфе).
  </div>

  <div class="codes-grid">
    {% for code in codes %}<div class="code-item">{{ code }}</div>{% endfor %}
  </div>

  <button type="button" class="btn-copy" onclick="navigator.clipboard.writeText(this.dataset.codes);this.textContent='Скопировано ✓';"
          data-codes="{% for code in codes %}{{ code }}{% if not loop.last %}\\n{% endif %}{% endfor %}">
    📋 Скопировать в буфер
  </button>

  <a class="btn-primary" href="/owner/security" style="display:block;margin-top:18px;text-decoration:none;">
    Я сохранил коды → Перейти к панели
  </a>
</div>
</body>
</html>'''


owner_security_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Безопасность')
    + '''
<div class="page-content" style="max-width:700px;margin:0 auto;">
  <h1 style="font-size:24px;font-weight:700;margin-bottom:18px;">🔐 Безопасность</h1>

  <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:14px;padding:24px;margin-bottom:18px;">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;margin-bottom:12px;">
      <div>
        <div style="font-size:18px;font-weight:600;margin-bottom:4px;">Двухфакторная аутентификация</div>
        <div style="font-size:13px;color:rgba(255,255,255,0.6);">
          {% if totp_enabled %}
          <span style="color:#10b981;">●</span> Включена · {{ remaining_codes }} recovery-кодов осталось
          {% else %}
          <span style="color:#ef4444;">●</span> Отключена
          {% endif %}
        </div>
      </div>
      {% if not totp_enabled %}
      <a href="/owner/2fa/setup" style="background:#fb923c;color:#1a0c00;font-weight:600;padding:10px 18px;border-radius:8px;text-decoration:none;font-size:14px;">Включить 2FA</a>
      {% endif %}
    </div>

    {% if totp_enabled %}
    <div style="font-size:13px;color:rgba(255,255,255,0.6);line-height:1.6;margin-top:14px;">
      При входе после email+пароля будет запрашиваться 6-значный код из приложения-аутентификатора.
      Если потеряли доступ к приложению — используйте один из recovery-кодов.
    </div>

    <details style="margin-top:18px;border-top:1px solid rgba(255,255,255,0.08);padding-top:14px;">
      <summary style="cursor:pointer;color:#ef4444;font-size:14px;font-weight:600;">Отключить 2FA</summary>
      <form method="post" action="/owner/2fa/disable" style="margin-top:14px;display:flex;flex-direction:column;gap:10px;">
        <input type="password" name="password" required placeholder="Текущий пароль"
               style="background:rgba(0,0,0,0.3);border:1px solid rgba(255,255,255,0.15);border-radius:8px;padding:10px 14px;color:#fff;font-family:inherit;font-size:14px;">
        <input type="text" name="code" required placeholder="6-значный код или recovery-код"
               autocomplete="one-time-code" inputmode="numeric"
               style="background:rgba(0,0,0,0.3);border:1px solid rgba(255,255,255,0.15);border-radius:8px;padding:10px 14px;color:#fff;font-family:inherit;font-size:14px;letter-spacing:2px;">
        <button type="submit" style="background:#ef4444;color:#fff;font-weight:600;padding:10px;border-radius:8px;border:none;cursor:pointer;font-size:14px;">
          Отключить 2FA
        </button>
      </form>
    </details>
    {% endif %}
  </div>

  <div style="font-size:12px;color:rgba(255,255,255,0.45);line-height:1.6;">
    <strong style="color:rgba(255,255,255,0.7);">Что ещё защищает owner-панель:</strong><br>
    • Rate-limit: 3 неудачные попытки → блокировка на 15 мин<br>
    • Idle-timeout: 2 часа без активности → автологаут<br>
    • Поисковики не индексируют, браузеры не кешируют, нет referrer-leak
  </div>
</div>
'''
    + _OWNER_PAGE_FOOT
)


owner_dashboard_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Owner Dashboard — Revisi</title>
{{ owner_pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _OWNER_BASE_CSS + '''</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>

<header class="owner-header">
  <div class="brand" style="display:flex;align-items:center;gap:10px;">{{ icon('chart', 22)|safe }} <span>Панель владельца</span></div>
  <div class="right">
    <span class="owner-email">{{ owner_email }}</span>
    <a class="btn-logout" href="/owner/logout">Выйти</a>
  </div>
</header>

{% with messages = get_flashed_messages(with_categories=true) %}
{% if messages %}
<div class="flash-messages">
  {% for cat, msg in messages %}
  <div class="flash-msg flash-{{ cat }}">{{ msg }}</div>
  {% endfor %}
</div>
{% endif %}
{% endwith %}
''' + _owner_nav('dashboard') + '''
<div class="page-content">

  <div class="metrics-grid-8">
    <div class="metric-card">
      <div class="metric-num small">{{ total_orgs }}</div>
      <div class="metric-label">Всего компаний</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ active_orgs }}</div>
      <div class="metric-label">Активных (7д)</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ trial_orgs }}</div>
      <div class="metric-label">На trial</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ paying_orgs }}</div>
      <div class="metric-label">Платящих</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ mrr }}₽</div>
      <div class="metric-label">MRR</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ conv_pct }}%</div>
      <div class="metric-label">Конверсия trial → paid</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ total_users }}</div>
      <div class="metric-label">Пользователей</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ total_revisions }}</div>
      <div class="metric-label">Завершённых ревизий</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ activity_today }}</div>
      <div class="metric-label">Активность сегодня</div>
    </div>
  </div>

  <div class="charts-grid">
    <div class="section-card">
      <div class="chart-title">🆕 Регистрации (30 дней)</div>
      <div class="chart-total">{{ signups_per_day|sum }}</div>
      {% set max_s = signups_per_day|max if signups_per_day|max > 0 else 1 %}
      <svg class="spark" viewBox="0 0 300 60" preserveAspectRatio="none">
        <defs><linearGradient id="g1" x1="0" y1="0" x2="0" y2="1">
          <stop offset="0%" stop-color="#a855f7" stop-opacity="0.55"/>
          <stop offset="100%" stop-color="#7c6cf0" stop-opacity="0"/></linearGradient></defs>
        {% set step = 300 / 29 %}
        {% set pts = [] %}
        {% for v in signups_per_day %}{% set _ = pts.append(((loop.index0 * step)|round(2), (60 - (v / max_s) * 50)|round(2))) %}{% endfor %}
        <polygon fill="url(#g1)" points="0,60 {% for x,y in pts %}{{ x }},{{ y }} {% endfor %}300,60"/>
        <polyline fill="none" stroke="#a855f7" stroke-width="2" points="{% for x,y in pts %}{{ x }},{{ y }} {% endfor %}"/>
      </svg>
    </div>
    <div class="section-card">
      <div class="chart-title">📦 Активность по дням (30 дней)</div>
      <div class="chart-total">{{ activity_per_day|sum }}</div>
      {% set max_a = activity_per_day|max if activity_per_day|max > 0 else 1 %}
      <svg class="spark" viewBox="0 0 300 60" preserveAspectRatio="none">
        <defs><linearGradient id="g2" x1="0" y1="0" x2="0" y2="1">
          <stop offset="0%" stop-color="#22d3ee" stop-opacity="0.55"/>
          <stop offset="100%" stop-color="#0ea5e9" stop-opacity="0"/></linearGradient></defs>
        {% set step2 = 300 / 29 %}
        {% set pts2 = [] %}
        {% for v in activity_per_day %}{% set _ = pts2.append(((loop.index0 * step2)|round(2), (60 - (v / max_a) * 50)|round(2))) %}{% endfor %}
        <polygon fill="url(#g2)" points="0,60 {% for x,y in pts2 %}{{ x }},{{ y }} {% endfor %}300,60"/>
        <polyline fill="none" stroke="#22d3ee" stroke-width="2" points="{% for x,y in pts2 %}{{ x }},{{ y }} {% endfor %}"/>
      </svg>
    </div>
  </div>

  {% if expiring_list %}
  <div class="section-card">
    <div class="section-title">⚠️ Trial заканчивается скоро (3 дня)</div>
    {% for item in expiring_list %}
    <div class="expiring-row">
      <div>
        <div class="expiring-name">{{ item.name }}</div>
        <div class="expiring-email">{{ item.email }}</div>
      </div>
      <div style="display:flex;align-items:center;gap:8px;">
        <span class="days-badge">{{ item.days_left }} д.</span>
        <form method="post" action="/owner/orgs/{{ item.id }}/extend_trial" style="display:inline;">
          <button class="btn-sm btn-extend" type="submit">Продлить +7д</button>
        </form>
      </div>
    </div>
    {% endfor %}
  </div>
  {% endif %}

  {% if recent_list %}
  <div class="section-card">
    <div class="section-title">🆕 Последние регистрации</div>
    <div class="tbl-wrap">
    <table class="recent-table">
      <thead><tr>
        <th>Компания</th><th>Email</th><th>Тариф</th><th>Дата</th>
      </tr></thead>
      <tbody>
      {% for r in recent_list %}
      <tr>
        <td style="font-weight:600;">{{ r.name }}</td>
        <td style="color:rgba(255,255,255,0.6);">{{ r.email }}</td>
        <td><span class="badge badge-{{ r.plan }}">{{ r.plan }}</span></td>
        <td style="color:rgba(255,255,255,0.5);">{{ r.created_at }}</td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    </div>
  </div>
  {% endif %}

</div>
</body>
</html>'''


owner_orgs_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Компании — Owner — Revisi</title>
{{ owner_pwa_head|safe }}
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
<style>''' + _OWNER_BASE_CSS + '''</style>
</head>
<body>
<div class="blob blob-1"></div>
<div class="blob blob-2"></div>

<header class="owner-header">
  <div class="brand" style="display:flex;align-items:center;gap:10px;">{{ icon('chart', 22)|safe }} <span>Панель владельца</span></div>
  <div class="right">
    <span class="owner-email">{{ owner_email }}</span>
    <a class="btn-logout" href="/owner/logout">Выйти</a>
  </div>
</header>

{% with messages = get_flashed_messages(with_categories=true) %}
{% if messages %}
<div class="flash-messages">
  {% for cat, msg in messages %}
  <div class="flash-msg flash-{{ cat }}">{{ msg }}</div>
  {% endfor %}
</div>
{% endif %}
{% endwith %}
''' + _owner_nav('orgs') + '''
<div class="page-content">

  <input class="search-box" type="text" id="searchBox" placeholder="Поиск по названию или email…" oninput="filterOrgs(this.value)">

  <!-- Desktop table -->
  <div class="glass" style="overflow:hidden;">
  <table class="orgs-table" id="orgsTable">
    <thead><tr>
      <th>Компания</th>
      <th>Email</th>
      <th>Тариф</th>
      <th>Истекает</th>
      <th>Польз.</th>
      <th>Активность</th>
      <th>Статус</th>
      <th>Действия</th>
    </tr></thead>
    <tbody id="orgsBody">
    {% for r in org_rows %}
    <tr class="org-row" data-search="{{ r.name|lower }} {{ r.email|lower }}">
      <td style="font-weight:700;">{{ r.name }}</td>
      <td style="color:rgba(255,255,255,0.6);">{{ r.email }}</td>
      <td><span class="badge badge-{{ r.plan }}">{{ r.plan }}</span></td>
      <td style="color:rgba(255,255,255,0.5);font-size:12px;">{{ r.ends }}</td>
      <td style="text-align:center;">{{ r.user_count }}</td>
      <td style="color:rgba(255,255,255,0.5);font-size:12px;">{{ r.last_activity }}</td>
      <td>
        {% if r.is_blocked %}
          <span class="status-dot dot-red"></span>Заблокирована
        {% else %}
          <span class="status-dot dot-green"></span>Активна
        {% endif %}
      </td>
      <td>
        <div class="actions-cell">
          <form method="post" action="/owner/orgs/{{ r.id }}/extend_trial" style="display:inline;">
            <button class="btn-sm btn-extend" type="submit" title="Продлить trial +7д">+7д</button>
          </form>
          <form method="post" action="/owner/orgs/{{ r.id }}/set_plan" style="display:inline;">
            <input type="hidden" name="plan" value="pro">
            <button class="btn-sm btn-pro" type="submit" title="Установить PRO">PRO</button>
          </form>
          <form method="post" action="/owner/orgs/{{ r.id }}/set_plan" style="display:inline;">
            <input type="hidden" name="plan" value="free">
            <button class="btn-sm btn-free" type="submit" title="Установить FREE">FREE</button>
          </form>
          <form method="post" action="/owner/orgs/{{ r.id }}/toggle_block" style="display:inline;">
            <button class="btn-sm btn-warn" type="submit" title="{{ 'Разблокировать' if r.is_blocked else 'Заблокировать' }}">{{ '🔓' if r.is_blocked else '🔒' }}</button>
          </form>
          <form method="post" action="/owner/orgs/{{ r.id }}/delete" style="display:inline;"
                onsubmit="return confirm('Удалить компанию «{{ r.name }}» и ВСЕ её данные? Это действие необратимо!');">
            <button class="btn-sm btn-danger" type="submit" title="Удалить">🗑</button>
          </form>
        </div>
      </td>
    </tr>
    {% endfor %}
    </tbody>
  </table>
  </div>

  <!-- Mobile cards -->
  <div class="mobile-cards" id="mobileCards">
  {% for r in org_rows %}
  <div class="mobile-card org-row" data-search="{{ r.name|lower }} {{ r.email|lower }}">
    <a class="org-card-link" href="/owner/orgs/{{ r.id }}">
      <div class="org-name">{{ r.name }} <span class="card-chev">›</span></div>
      <div class="org-email">{{ r.email }}</div>
    </a>
    <div class="details">
      <span><span class="badge badge-{{ r.plan }}">{{ r.plan }}</span></span>
      <span>До: {{ r.ends }}</span>
      <span>Польз.: {{ r.user_count }}</span>
    </div>
    <div style="font-size:12px;color:rgba(255,255,255,0.5);margin-bottom:6px;">
      Активность: {{ r.last_activity }} &nbsp;|&nbsp;
      {% if r.is_blocked %}
        <span class="status-dot dot-red"></span>Заблокирована
      {% else %}
        <span class="status-dot dot-green"></span>Активна
      {% endif %}
    </div>
    <div class="actions">
      <form method="post" action="/owner/orgs/{{ r.id }}/extend_trial" style="display:inline;">
        <button class="btn-sm btn-extend" type="submit">+7д trial</button>
      </form>
      <form method="post" action="/owner/orgs/{{ r.id }}/set_plan" style="display:inline;">
        <input type="hidden" name="plan" value="pro">
        <button class="btn-sm btn-pro" type="submit">PRO</button>
      </form>
      <form method="post" action="/owner/orgs/{{ r.id }}/set_plan" style="display:inline;">
        <input type="hidden" name="plan" value="free">
        <button class="btn-sm btn-free" type="submit">FREE</button>
      </form>
      <form method="post" action="/owner/orgs/{{ r.id }}/toggle_block" style="display:inline;">
        <button class="btn-sm btn-warn" type="submit">{{ '🔓 Разблок.' if r.is_blocked else '🔒 Блок.' }}</button>
      </form>
      <form method="post" action="/owner/orgs/{{ r.id }}/delete" style="display:inline;"
            onsubmit="return confirm('Удалить «{{ r.name }}»? Все данные будут потеряны!');">
        <button class="btn-sm btn-danger" type="submit">🗑 Удалить</button>
      </form>
    </div>
  </div>
  {% endfor %}
  </div>

</div>

<script>
function filterOrgs(q) {
  q = q.toLowerCase().trim();
  document.querySelectorAll('.org-row').forEach(function(row) {
    var text = row.dataset.search || '';
    row.style.display = (!q || text.indexOf(q) !== -1) ? '' : 'none';
  });
}
</script>
</body>
</html>'''


# ============== ДЕТАЛИ КОМПАНИИ ==============
owner_org_detail_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Компания')
    + _owner_nav('orgs') + '''
<div class="page-content">

  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:18px;flex-wrap:wrap;gap:12px;">
    <div style="flex:1;min-width:0;">
      <h1 style="font-size:22px;font-weight:700;margin-bottom:4px;word-break:break-word;">{{ org.name }}</h1>
      <div style="font-size:13px;color:rgba(255,255,255,0.6);word-break:break-all;">{{ org.owner_email }}</div>
    </div>
    <div style="display:flex;gap:6px;flex-wrap:wrap;">
      <a class="btn-sm btn-extend" href="/owner/orgs">← К списку</a>
      <form method="post" action="/owner/orgs/{{ org.id }}/impersonate" style="display:inline;"
            onsubmit="return confirm('Войти как admin компании «{{ org.name }}»? Действие записывается в аудит.');">
        <button class="btn-sm btn-pro" type="submit">🎭 Войти как admin</button>
      </form>
    </div>
  </div>

  <!-- Метрики -->
  <div class="metrics-grid-8">
    <div class="metric-card">
      <div class="metric-num small">
        <span class="badge badge-{{ org.plan }}">{{ org.plan }}</span>
      </div>
      <div class="metric-label">Тариф · до {{ ends_str }}</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ user_rows|length }}</div>
      <div class="metric-label">Пользователи</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ locations_count }}</div>
      <div class="metric-label">Локаций</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ products_count }}</div>
      <div class="metric-label">Товаров</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ revisions_count }}</div>
      <div class="metric-label">Всего ревизий</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ items_7d }}</div>
      <div class="metric-label">Записей за 7 дней</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ org.monthly_price or 0 }}₽</div>
      <div class="metric-label">Подписка / мес</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{% if org.is_blocked %}<span style="color:#fca5a5;">🔒</span>{% else %}<span style="color:#6ee7b7;">✓</span>{% endif %}</div>
      <div class="metric-label">{% if org.is_blocked %}Заблокирована{% else %}Активна{% endif %}</div>
    </div>
  </div>

  <!-- Sparkline активности -->
  <div class="section-card">
    <div class="chart-title">Активность за 14 дней (записи в ревизиях)</div>
    {% set max_v = spark|max %}
    {% if max_v == 0 %}{% set max_v = 1 %}{% endif %}
    <svg class="spark" viewBox="0 0 280 60" preserveAspectRatio="none">
      <defs>
        <linearGradient id="ga" x1="0" y1="0" x2="0" y2="1">
          <stop offset="0%" stop-color="#a855f7" stop-opacity="0.45"/>
          <stop offset="100%" stop-color="#7c6cf0" stop-opacity="0.0"/>
        </linearGradient>
      </defs>
      <polyline fill="none" stroke="#a855f7" stroke-width="2"
        points="{% for v in spark %}{{ loop.index0 * 280 / 13 }},{{ 56 - (v / max_v) * 50 }} {% endfor %}"/>
      <polygon fill="url(#ga)"
        points="0,60 {% for v in spark %}{{ loop.index0 * 280 / 13 }},{{ 56 - (v / max_v) * 50 }} {% endfor %} 280,60"/>
    </svg>
  </div>

  <!-- Цена подписки -->
  <div class="section-card">
    <div class="section-title">💰 Цена подписки</div>
    <form method="post" action="/owner/orgs/{{ org.id }}/price" class="input-row">
      <input type="number" inputmode="numeric" pattern="[0-9]*" name="monthly_price" min="0" step="100" value="{{ org.monthly_price or 0 }}" class="grow" placeholder="Стоимость в ₽">
      <button type="submit" class="btn-primary-sm">Сохранить</button>
    </form>
    <div style="font-size:12px;color:rgba(255,255,255,0.5);">Используется в расчёте MRR на дашборде.</div>
  </div>

  <!-- Feature-флаги -->
  <div class="section-card">
    <div class="section-title">🚦 Feature-флаги</div>
    <form method="post" action="/owner/orgs/{{ org.id }}/features">
      <div class="feature-toggle">
        <div>
          <div style="font-weight:600;font-size:14px;">Excel-экспорт</div>
          <div style="font-size:12px;color:rgba(255,255,255,0.5);">Кнопка скачивания отчётов в .xlsx</div>
        </div>
        <label class="toggle-switch">
          <input type="checkbox" name="excel_export" {% if features.get('excel_export') %}checked{% endif %}>
          <span class="toggle-slider"></span>
        </label>
      </div>
      <div class="feature-toggle">
        <div>
          <div style="font-weight:600;font-size:14px;">Несколько локаций</div>
          <div style="font-size:12px;color:rgba(255,255,255,0.5);">Иначе только одна локация</div>
        </div>
        <label class="toggle-switch">
          <input type="checkbox" name="multi_location" {% if features.get('multi_location') %}checked{% endif %}>
          <span class="toggle-slider"></span>
        </label>
      </div>
      <div class="feature-toggle">
        <div>
          <div style="font-weight:600;font-size:14px;">История ревизий</div>
          <div style="font-size:12px;color:rgba(255,255,255,0.5);">Доступ к архиву прошлых ревизий</div>
        </div>
        <label class="toggle-switch">
          <input type="checkbox" name="history" {% if features.get('history') %}checked{% endif %}>
          <span class="toggle-slider"></span>
        </label>
      </div>
      <div class="feature-toggle">
        <div>
          <div style="font-weight:600;font-size:14px;">Безлимит пользователей</div>
          <div style="font-size:12px;color:rgba(255,255,255,0.5);">Без ограничений по числу операторов</div>
        </div>
        <label class="toggle-switch">
          <input type="checkbox" name="unlimited_users" {% if features.get('unlimited_users') %}checked{% endif %}>
          <span class="toggle-slider"></span>
        </label>
      </div>
      <button type="submit" class="btn-primary-sm" style="margin-top:14px;">Сохранить флаги</button>
    </form>
  </div>

  <!-- Шаблон бланка отчёта -->
  <div class="section-card">
    <div class="section-title">📄 Шаблон бланка отчёта (.xlsx)</div>
    <div style="font-size:12px;color:rgba(255,255,255,0.6);margin-bottom:12px;">
      При завершении ревизии остатки подставятся в этот шаблон.
      Колонка <b>B</b> = код товара, колонка <b>C</b> = название (фолбэк),
      колонка <b>G</b> — туда пишется фактический остаток.
      Если шаблон не загружен — будет стандартный отчёт.
    </div>
    {% if has_template %}
    <div style="background:rgba(34,197,94,0.12);border:1px solid rgba(34,197,94,0.3);border-radius:10px;padding:10px 12px;font-size:13px;margin-bottom:12px;">
      ✅ Шаблон загружен · <b>{{ template_size }}</b> байт
    </div>
    <form method="post" action="/owner/orgs/{{ org.id }}/template/remove"
          onsubmit="return confirm('Удалить шаблон? Отчёты будут генериться в стандартном виде.');"
          style="display:inline;">
      <button type="submit" class="btn-primary-sm" style="background:#ef4444;">Удалить шаблон</button>
    </form>
    <span style="font-size:12px;color:rgba(255,255,255,0.5);margin-left:8px;">или замените, загрузив новый ниже</span>
    {% else %}
    <div style="background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.3);border-radius:10px;padding:10px 12px;font-size:13px;margin-bottom:12px;">
      ⚠ Шаблон не загружен — отчёт будет в стандартном формате
    </div>
    {% endif %}
    <form method="post" action="/owner/orgs/{{ org.id }}/template/upload"
          enctype="multipart/form-data" style="margin-top:12px;display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
      <input type="file" name="file" accept=".xlsx" required
             style="color:rgba(255,255,255,0.85);font-size:13px;">
      <button type="submit" class="btn-primary-sm">Загрузить шаблон</button>
    </form>
    <div style="font-size:11px;color:rgba(255,255,255,0.4);margin-top:8px;">
      Поддерживается только .xlsx (Excel 2007+). Старые .xls и XML SpreadsheetML нужно пересохранить.
    </div>
  </div>

  <!-- Пользователи -->
  <div class="section-card">
    <div class="section-title">👥 Пользователи компании</div>
    {% if user_rows %}
    <div class="tbl-wrap">
    <table class="recent-table" style="min-width:520px;">
      <thead><tr><th>Логин</th><th>Email</th><th>Роль</th><th>Последний вход</th></tr></thead>
      <tbody>
        {% for u in user_rows %}
        <tr>
          <td style="font-weight:600;">{{ u.username }}</td>
          <td style="color:rgba(255,255,255,0.6);">{{ u.email }}</td>
          <td><span class="badge badge-{{ 'pro' if u.role == 'admin' else 'free' }}">{{ u.role }}</span></td>
          <td style="color:rgba(255,255,255,0.5);font-size:12px;">{{ u.last_login }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
    </div>
    {% else %}
    <div style="color:rgba(255,255,255,0.5);font-size:13px;">Пока нет пользователей.</div>
    {% endif %}
  </div>

</div>
''' + _OWNER_PAGE_FOOT
)


# ============== АУДИТ ==============
owner_audit_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Аудит')
    + _owner_nav('audit') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:14px;">📋 Журнал действий владельца</h1>
  <div style="font-size:13px;color:rgba(255,255,255,0.5);margin-bottom:14px;">Последние 200 значимых действий. Используется для расследования инцидентов и комплаенса.</div>
  <div class="glass tbl-wrap" style="padding:12px;">
    {% if log_rows %}
    <table class="audit-table">
      <thead><tr>
        <th>Когда</th><th>Действие</th><th>Компания</th><th>IP</th><th>Детали</th>
      </tr></thead>
      <tbody>
      {% for r in log_rows %}
      <tr>
        <td style="white-space:nowrap;color:rgba(255,255,255,0.7);">{{ r.when }}</td>
        <td><span class="action-pill">{{ r.action }}</span></td>
        <td style="color:rgba(255,255,255,0.7);">{{ r.org }}</td>
        <td style="color:rgba(255,255,255,0.5);font-family:ui-monospace,Menlo,Consolas,monospace;font-size:11px;">{{ r.ip }}</td>
        <td style="color:rgba(255,255,255,0.65);font-family:ui-monospace,Menlo,Consolas,monospace;font-size:11px;">{{ r.details }}</td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    {% else %}
    <div style="padding:30px;text-align:center;color:rgba(255,255,255,0.4);">Журнал пуст.</div>
    {% endif %}
  </div>
</div>
''' + _OWNER_PAGE_FOOT
)


# ============== АЛЕРТЫ ==============
owner_alerts_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Алерты')
    + _owner_nav('alerts') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:18px;">⚠️ Алерты</h1>

  <div class="alert-section">
    <h3>🔥 Trial истекает в течение 3 дней <span style="color:rgba(255,255,255,0.4);font-size:12px;font-weight:400;">({{ expiring_3|length }})</span></h3>
    {% if expiring_3 %}
      {% for r in expiring_3 %}
      <div class="alert-row">
        <div>
          <a href="/owner/orgs/{{ r.id }}" style="color:white;text-decoration:none;font-weight:600;">{{ r.name }}</a>
          <div style="font-size:11px;color:rgba(255,255,255,0.5);">{{ r.email }}</div>
        </div>
        <div>
          <span class="days-badge">{{ r.days }} дн.</span>
          <form method="post" action="/owner/orgs/{{ r.id }}/extend_trial" style="display:inline;">
            <button class="btn-sm btn-extend" type="submit">+7д</button>
          </form>
        </div>
      </div>
      {% endfor %}
    {% else %}<div style="color:rgba(255,255,255,0.45);font-size:13px;">Никого. 👌</div>{% endif %}
  </div>

  <div class="alert-section warn">
    <h3>⏳ Trial истекает в течение 4–7 дней <span style="color:rgba(255,255,255,0.4);font-size:12px;font-weight:400;">({{ expiring_7|length }})</span></h3>
    {% if expiring_7 %}
      {% for r in expiring_7 %}
      <div class="alert-row">
        <div>
          <a href="/owner/orgs/{{ r.id }}" style="color:white;text-decoration:none;font-weight:600;">{{ r.name }}</a>
          <div style="font-size:11px;color:rgba(255,255,255,0.5);">{{ r.email }}</div>
        </div>
        <div>
          <span class="days-badge">{{ r.days }} дн.</span>
          <form method="post" action="/owner/orgs/{{ r.id }}/extend_trial" style="display:inline;">
            <button class="btn-sm btn-extend" type="submit">+7д</button>
          </form>
        </div>
      </div>
      {% endfor %}
    {% else %}<div style="color:rgba(255,255,255,0.45);font-size:13px;">Пусто.</div>{% endif %}
  </div>

  <div class="alert-section info">
    <h3>💸 Платная подписка истекает в течение 7 дней <span style="color:rgba(255,255,255,0.4);font-size:12px;font-weight:400;">({{ paid_expiring|length }})</span></h3>
    {% if paid_expiring %}
      {% for r in paid_expiring %}
      <div class="alert-row">
        <div>
          <a href="/owner/orgs/{{ r.id }}" style="color:white;text-decoration:none;font-weight:600;">{{ r.name }}</a>
          <div style="font-size:11px;color:rgba(255,255,255,0.5);">{{ r.email }}</div>
        </div>
        <span class="days-badge">{{ r.days }} дн.</span>
      </div>
      {% endfor %}
    {% else %}<div style="color:rgba(255,255,255,0.45);font-size:13px;">Никаких ожидаемых истечений.</div>{% endif %}
  </div>

  <div class="alert-section warn">
    <h3>😴 Неактивные компании <span style="color:rgba(255,255,255,0.4);font-size:12px;font-weight:400;">({{ inactive|length }})</span></h3>
    {% if inactive %}
      {% for r in inactive %}
      <div class="alert-row">
        <div>
          <a href="/owner/orgs/{{ r.id }}" style="color:white;text-decoration:none;font-weight:600;">{{ r.name }}</a>
          <div style="font-size:11px;color:rgba(255,255,255,0.5);">{{ r.email }} · последний вход: {{ r.last }}</div>
        </div>
        {% if r.churn_risk %}
          <span class="ticket-status status-open">churn-risk</span>
        {% else %}
          <span class="ticket-status status-answered">14d+</span>
        {% endif %}
      </div>
      {% endfor %}
    {% else %}<div style="color:rgba(255,255,255,0.45);font-size:13px;">Все активны.</div>{% endif %}
  </div>

</div>
''' + _OWNER_PAGE_FOOT
)


# ============== РАССЫЛКА ==============
owner_broadcast_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Рассылка')
    + _owner_nav('broadcast') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:18px;">📨 Рассылка клиентам</h1>

  <div class="section-card">
    <form method="post" action="/owner/broadcast">
      <div style="margin-bottom:14px;">
        <label style="display:block;font-size:12px;color:rgba(255,255,255,0.6);margin-bottom:6px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Аудитория</label>
        <select name="audience" class="grow" style="background:rgba(255,255,255,0.07);border:1px solid rgba(255,255,255,0.12);border-radius:10px;color:white;padding:10px 12px;font-family:'Outfit',sans-serif;font-size:14px;width:100%;">
          <option value="all">Все компании</option>
          <option value="trial">Только trial (активные)</option>
          <option value="paid">Только платные (pro/business)</option>
          <option value="inactive14">Не заходили 14+ дней</option>
        </select>
      </div>
      <div style="margin-bottom:14px;">
        <label style="display:block;font-size:12px;color:rgba(255,255,255,0.6);margin-bottom:6px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Тема письма</label>
        <input name="subject" required class="grow" style="background:rgba(255,255,255,0.07);border:1px solid rgba(255,255,255,0.12);border-radius:10px;color:white;padding:10px 12px;font-family:'Outfit',sans-serif;font-size:14px;width:100%;" placeholder="Например: Новые функции в Revisi">
      </div>
      <div style="margin-bottom:14px;">
        <label style="display:block;font-size:12px;color:rgba(255,255,255,0.6);margin-bottom:6px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Текст</label>
        <textarea name="body" required class="broadcast-area" placeholder="Здравствуйте! Мы добавили..."></textarea>
      </div>
      <button type="submit" class="btn-primary-sm">📨 Отправить рассылку</button>
    </form>
  </div>

  <div class="section-card" style="font-size:13px;color:rgba(255,255,255,0.6);">
    <div class="section-title">ℹ️ Как это работает</div>
    <p style="margin-bottom:8px;">Сейчас рассылка фиксирует адресатов и пишет запись в аудит. Реальная отправка email подключается отдельно (SMTP / SendGrid / Postmark).</p>
    <p>После настройки SMTP письма будут улетать в фоновом режиме на адреса <code class="code-block" style="display:inline;padding:2px 6px;">organization.owner_email</code>.</p>
  </div>

</div>
''' + _OWNER_PAGE_FOOT
)


# ============== БЭКАП ==============
owner_backup_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Бэкап')
    + _owner_nav('backup') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:18px;">🗄 Бэкап базы данных</h1>

  <div class="section-card">
    <div class="section-title">📦 Скачать снимок данных (JSON)</div>
    <p style="font-size:13px;color:rgba(255,255,255,0.6);margin-bottom:16px;">
      Будет сгенерирован zip со всеми основными таблицами в JSON: компании, пользователи, локации, категории, товары, нормы, ревизии, тикеты, каталог, аудит. Бинарные данные (Excel-шаблоны) исключены — фиксируется только размер.
    </p>
    <a class="btn-primary-sm" href="/owner/backup/download" style="text-decoration:none;display:inline-block;">⬇️ Скачать ZIP</a>
  </div>

  <div class="section-card" style="font-size:13px;color:rgba(255,255,255,0.6);">
    <div class="section-title">💡 Рекомендация</div>
    <p>Для production-окружения настройте автоматические бэкапы PostgreSQL на стороне Render (или своего хостинга). Этот ручной экспорт — удобный быстрый снимок для отладки и переноса.</p>
  </div>
</div>
''' + _OWNER_PAGE_FOOT
)


# ============== ТИКЕТЫ (СПИСОК) ==============
owner_tickets_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Тикеты')
    + _owner_nav('tickets') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:14px;">🎫 Тикеты поддержки</h1>
  <div style="font-size:13px;color:rgba(255,255,255,0.5);margin-bottom:18px;">Сначала открытые, затем по дате обновления.</div>

  {% if items %}
    {% for t in items %}
    <a href="/owner/tickets/{{ t.id }}" class="ticket-row">
      <div style="flex:1;min-width:0;">
        <div style="font-weight:600;font-size:14px;margin-bottom:3px;">{{ t.subject }}</div>
        <div style="font-size:12px;color:rgba(255,255,255,0.5);">{{ t.org_name }} · {{ t.when }}</div>
      </div>
      <span class="ticket-status status-{{ t.status }}">{{ t.status }}</span>
    </a>
    {% endfor %}
  {% else %}
    <div class="section-card" style="text-align:center;color:rgba(255,255,255,0.5);padding:40px 20px;">
      <div style="font-size:42px;margin-bottom:10px;opacity:0.6;">📭</div>
      <div>Нет тикетов</div>
    </div>
  {% endif %}
</div>
''' + _OWNER_PAGE_FOOT
)


# ============== ТИКЕТ (ДЕТАЛИ) ==============
owner_ticket_view_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Тикет')
    + _owner_nav('tickets') + '''
<div class="page-content">
  <a href="/owner/tickets" style="color:#a78bfa;text-decoration:none;font-size:13px;">← К списку тикетов</a>

  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin:14px 0 18px;flex-wrap:wrap;gap:12px;">
    <div style="flex:1;min-width:0;">
      <h1 style="font-size:20px;font-weight:700;margin-bottom:6px;">{{ t.subject }}</h1>
      <div style="font-size:13px;color:rgba(255,255,255,0.55);">
        Компания: <a href="/owner/orgs/{{ org.id if org else 0 }}" style="color:#a78bfa;text-decoration:none;">{{ org.name if org else '—' }}</a>
        · Создан: {{ t.created_at.strftime('%d.%m.%Y %H:%M') if t.created_at else '—' }}
      </div>
    </div>
    <span class="ticket-status status-{{ t.status }}">{{ t.status }}</span>
  </div>

  <div class="section-card">
    <div style="font-size:11px;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:8px;">Исходное обращение</div>
    <div style="white-space:pre-wrap;font-size:14px;line-height:1.55;color:rgba(255,255,255,0.9);">{{ t.body }}</div>
  </div>

  {% for r in replies %}
  <div class="section-card" style="border-left: 3px solid {{ '#a855f7' if r.author_type == 'owner' else '#10b981' }};">
    <div style="font-size:11px;color:rgba(255,255,255,0.5);margin-bottom:6px;">
      <b>{{ r.author_label or ('Владелец' if r.author_type == 'owner' else 'Клиент') }}</b>
      · {{ r.created_at.strftime('%d.%m.%Y %H:%M') if r.created_at else '' }}
    </div>
    <div style="white-space:pre-wrap;font-size:14px;line-height:1.55;">{{ r.body }}</div>
  </div>
  {% endfor %}

  {% if t.status != 'closed' %}
  <div class="section-card">
    <div class="section-title">Ответить</div>
    <form method="post">
      <textarea name="body" class="broadcast-area" placeholder="Ваш ответ клиенту..." style="min-height:120px;"></textarea>
      <div style="display:flex;gap:8px;margin-top:12px;flex-wrap:wrap;">
        <button type="submit" name="action" value="reply" class="btn-primary-sm">💬 Отправить ответ</button>
        <button type="submit" name="action" value="close" class="btn-sm btn-warn"
                onclick="return confirm('Закрыть тикет?');">🗙 Закрыть тикет</button>
      </div>
    </form>
  </div>
  {% else %}
  <div class="section-card" style="text-align:center;color:rgba(255,255,255,0.5);">Тикет закрыт.</div>
  {% endif %}
</div>
''' + _OWNER_PAGE_FOOT
)


# ============== КАТАЛОГ ==============
owner_catalog_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Каталог')
    + _owner_nav('catalog') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:6px;">📦 Глобальный каталог</h1>
  <div style="font-size:13px;color:rgba(255,255,255,0.55);margin-bottom:18px;">Стартовый набор товаров. Клиенты смогут импортировать его одной кнопкой при первом входе.</div>

  <div class="section-card">
    <div class="section-title">Добавить товар</div>
    <form method="post" action="/owner/catalog">
      <div class="input-row">
        <input name="name" required placeholder="Название товара" class="grow">
        <input name="code" placeholder="Код (опц.)" style="width:120px;">
        <input name="unit" placeholder="Ед." value="шт" style="width:80px;">
      </div>
      <div class="input-row">
        <input name="category" placeholder="Категория (например, Напитки)" class="grow">
        <button type="submit" class="btn-primary-sm">+ Добавить</button>
      </div>
    </form>
  </div>

  <div class="section-card tbl-wrap" style="padding:0;">
    {% if rows %}
    <table class="recent-table" style="margin:0;min-width:560px;">
      <thead><tr>
        <th>Категория</th><th>Название</th><th>Код</th><th>Ед.</th><th></th>
      </tr></thead>
      <tbody>
      {% for r in rows %}
      <tr>
        <td style="color:rgba(255,255,255,0.6);">{{ r.category_name or '—' }}</td>
        <td style="font-weight:600;">{{ r.name }}</td>
        <td style="color:rgba(255,255,255,0.5);font-family:ui-monospace,Menlo,Consolas,monospace;font-size:12px;">{{ r.code or '—' }}</td>
        <td>{{ r.unit }}</td>
        <td style="text-align:right;">
          <form method="post" action="/owner/catalog/{{ r.id }}/delete" style="display:inline;"
                onsubmit="return confirm('Удалить «{{ r.name }}» из каталога?');">
            <button class="btn-sm btn-danger" type="submit">🗑</button>
          </form>
        </td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    {% else %}
    <div style="padding:30px;text-align:center;color:rgba(255,255,255,0.45);">Каталог пуст. Добавьте первые товары выше.</div>
    {% endif %}
  </div>
</div>
''' + _OWNER_PAGE_FOOT
)


# ============== HEALTH ==============
owner_health_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'Health')
    + _owner_nav('health') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:18px;">🏥 Health-check</h1>

  <div class="metrics-grid-8">
    <div class="metric-card">
      <div class="metric-num small" style="color: {{ '#6ee7b7' if h.db_ok else '#fca5a5' }};">
        {{ '✓ OK' if h.db_ok else '✗ FAIL' }}
      </div>
      <div class="metric-label">База данных</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ h.db_latency_ms if h.db_latency_ms is not none else '—' }} мс</div>
      <div class="metric-label">Latency БД</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ h.db_size }}</div>
      <div class="metric-label">Размер БД</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ h.uptime }}</div>
      <div class="metric-label">Uptime приложения</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ h.orgs }}</div>
      <div class="metric-label">Организаций</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ h.users }}</div>
      <div class="metric-label">Пользователей</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ h.products }}</div>
      <div class="metric-label">Товаров</div>
    </div>
    <div class="metric-card">
      <div class="metric-num small">{{ h.revisions }} / {{ h.items }}</div>
      <div class="metric-label">Ревизий / записей</div>
    </div>
  </div>

  <div class="section-card" style="font-size:13px;color:rgba(255,255,255,0.6);">
    <div class="section-title">ℹ️ Запущено</div>
    <code class="code-block">{{ h.app_started }} UTC</code>
  </div>
</div>
''' + _OWNER_PAGE_FOOT
)


# ============== SQL КОНСОЛЬ ==============
owner_sql_html = (
    _OWNER_PAGE_HEAD.replace('__TITLE__', 'SQL')
    + _owner_nav('sql') + '''
<div class="page-content">
  <h1 style="font-size:22px;font-weight:700;margin-bottom:6px;">🔧 SQL-консоль</h1>
  <div style="font-size:13px;color:rgba(255,255,255,0.55);margin-bottom:18px;">По умолчанию разрешены только SELECT. Все запросы пишутся в аудит.</div>

  <div class="section-card">
    <form method="post">
      <textarea name="query" class="broadcast-area" style="min-height:120px;font-family:ui-monospace,Menlo,Consolas,monospace;font-size:13px;" placeholder="SELECT id, name, plan FROM organizations ORDER BY created_at DESC LIMIT 20;">{{ query }}</textarea>
      <div style="display:flex;gap:14px;align-items:center;margin-top:12px;flex-wrap:wrap;">
        <label style="display:flex;align-items:center;gap:8px;font-size:13px;cursor:pointer;">
          <input type="checkbox" name="allow_write" style="width:18px;height:18px;accent-color:#7c6cf0;">
          <span style="color:#fcd34d;">⚠️ Разрешить запись (INSERT/UPDATE/DELETE/DDL)</span>
        </label>
        <button type="submit" class="btn-primary-sm">▶ Выполнить</button>
      </div>
    </form>
  </div>

  {% if executed %}
    {% if error %}
    <div class="section-card" style="background:rgba(239,68,68,0.08);border-color:rgba(239,68,68,0.3);">
      <div class="section-title" style="color:#fca5a5;">❌ Ошибка</div>
      <code class="code-block" style="color:#fca5a5;white-space:pre-wrap;">{{ error }}</code>
    </div>
    {% else %}
    <div class="section-card tbl-wrap" style="padding:0;">
      <div style="padding:14px 16px 8px;font-size:12px;color:rgba(255,255,255,0.55);">Результат: {{ result_rows|length }} строк</div>
      <table class="recent-table" style="margin:0;font-family:ui-monospace,Menlo,Consolas,monospace;font-size:12px;min-width:520px;">
        <thead><tr>{% for c in result_cols %}<th>{{ c }}</th>{% endfor %}</tr></thead>
        <tbody>
          {% for row in result_rows %}
          <tr>{% for v in row %}<td style="color:rgba(255,255,255,0.85);">{{ v if v is not none else 'NULL' }}</td>{% endfor %}</tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
    {% endif %}
  {% endif %}
</div>
''' + _OWNER_PAGE_FOOT
)


# Создаём таблицы и owner при первом запуске
with app.app_context():
    db.create_all()

    # Миграция: добавляем колонку excel_template если её нет (PostgreSQL)
    try:
        with db.engine.connect() as conn:
            conn.execute(db.text(
                'ALTER TABLE organizations ADD COLUMN IF NOT EXISTS excel_template BYTEA'
            ))
            conn.execute(db.text(
                'ALTER TABLE organizations ADD COLUMN IF NOT EXISTS features JSONB'
            ))
            conn.execute(db.text(
                'ALTER TABLE organizations ADD COLUMN IF NOT EXISTS monthly_price INTEGER DEFAULT 0'
            ))
            conn.execute(db.text(
                "ALTER TABLE organizations ADD COLUMN IF NOT EXISTS timezone VARCHAR(50) DEFAULT 'Europe/Moscow'"
            ))
            conn.execute(db.text(
                "ALTER TABLE organizations ADD COLUMN IF NOT EXISTS pos_system VARCHAR(20)"
            ))
            # 2FA для owner_users
            conn.execute(db.text(
                'ALTER TABLE owner_users ADD COLUMN IF NOT EXISTS totp_secret VARCHAR(64)'
            ))
            conn.execute(db.text(
                'ALTER TABLE owner_users ADD COLUMN IF NOT EXISTS totp_enabled BOOLEAN DEFAULT FALSE NOT NULL'
            ))
            conn.execute(db.text(
                'ALTER TABLE owner_users ADD COLUMN IF NOT EXISTS recovery_codes_json TEXT'
            ))
            # Чтобы удаление пользователя/локации не уничтожало историю ревизий,
            # делаем revisions.user_id и revision_items.location_id nullable.
            # IF EXISTS — на случай если миграция уже была применена.
            conn.execute(db.text('ALTER TABLE revisions ALTER COLUMN user_id DROP NOT NULL'))
            conn.execute(db.text('ALTER TABLE revision_items ALTER COLUMN location_id DROP NOT NULL'))
            conn.commit()
    except Exception as _e:
        print(f'⚠️  Миграция колонок organizations: {_e}')

    # Привязка кастомного бланка отчёта к ООО "Прайд". Срабатывает один раз —
    # после успешной привязки в features ставится флаг и хук больше не лезет,
    # даже если админ потом сам удалит шаблон через owner-панель.
    try:
        _tpl_path = os.path.join(os.path.dirname(__file__), 'template_pride.xlsx')
        if os.path.exists(_tpl_path):
            _pride = Organization.query.filter(
                Organization.name.ilike('ООО%прайд%')
            ).first()
            # dict(...) — копия, иначе SQLAlchemy не заметит мутацию JSON-поля
            _feats = dict(_pride.features or {}) if _pride else {}
            if (
                _pride and not _pride.excel_template
                and not _feats.get('pride_template_seeded')
            ):
                with open(_tpl_path, 'rb') as _f:
                    _pride.excel_template = _f.read()
                _feats['pride_template_seeded'] = True
                _pride.features = _feats
                db.session.commit()
                print(f'✅ Шаблон отчёта привязан к {_pride.name!r}')
    except Exception as _e:
        db.session.rollback()
        print(f'⚠️  Привязка шаблона Прайда: {_e}')

    # Схлопывание дублей открытых ревизий. Если на одной локации зависло
    # несколько in_progress/pending ревизий (наследие гонок и старого бага,
    # когда вторая сотрудница создавала отдельную ревизию), СЛИВАЕМ их позиции
    # в одну, а опустевшие отменяем. Данные НЕ теряются — это ключевое отличие
    # от прежней версии, которая удаляла позиции «дублей».
    # Идемпотентно: после первого прохода дублей не остаётся.
    try:
        from sqlalchemy import func as _sa_func2
        _merged = 0
        _dup_pairs = (
            db.session.query(Revision.org_id, Revision.location_id)
            .filter(Revision.status.in_(['in_progress', 'pending']))
            .filter(Revision.location_id.isnot(None))
            .group_by(Revision.org_id, Revision.location_id)
            .having(_sa_func2.count(Revision.id) > 1)
            .all()
        )
        for _org_id, _loc_id in _dup_pairs:
            _candidates = Revision.query.filter(
                Revision.org_id == _org_id,
                Revision.location_id == _loc_id,
                Revision.status.in_(['in_progress', 'pending']),
            ).order_by(
                # главной делаем pending (её уже отправили на проверку),
                # иначе — самую раннюю по дате создания
                (Revision.status == 'pending').desc(),
                Revision.created_at.asc(),
            ).all()
            _keep = _candidates[0]
            for _r in _candidates[1:]:
                RevisionItem.query.filter_by(revision_id=_r.id).update(
                    {'revision_id': _keep.id}, synchronize_session=False
                )
                _r.status = 'cancelled'
                _r.finished_at = datetime.utcnow()
                _merged += 1
        if _merged:
            db.session.commit()
            print(f'Схлопнуто дублей ревизий (с сохранением данных): {_merged}')
    except Exception as _e:
        db.session.rollback()
        print(f'Схлопывание дублей ревизий: {_e}')

    # Партиал-уникальный индекс: не больше одной ОТКРЫТОЙ (in_progress/pending)
    # ревизии на локацию. Создаётся после схлопывания дублей выше, иначе упадёт.
    # С ним _get_or_create_active_revision ловит IntegrityError и возвращает
    # существующую ревизию вместо создания второй — гонка двух операторов
    # больше не плодит параллельные ревизии.
    try:
        with db.engine.connect() as conn:
            conn.execute(db.text(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_open_revision_per_loc "
                "ON revisions (org_id, location_id) "
                "WHERE status IN ('in_progress', 'pending') AND location_id IS NOT NULL"
            ))
            conn.commit()
    except Exception as _e:
        print(f'Индекс uq_open_revision_per_loc: {_e}')

    # Чистка: убираем сокращение " в асс" / " в ассорт" / " в ассортименте"
    # из названий товаров. Звучит двусмысленно, и пользователь явно попросил
    # его убрать. Идемпотентно — повторные запуски ничего не делают.
    try:
        import re as _re_clean
        # Требуем, чтобы после "асс[.орт.]" шёл разделитель — иначе слова
        # "ассорти", "ассистент" и т.п. не пострадают.
        _pat = _re_clean.compile(
            r'\s*\(?\s*в\s+асс(?:\.|орт\.?|ортименте)?(?=\s|[,.)\]]|$)\s*\)?',
            _re_clean.IGNORECASE,
        )
        _cleaned = 0
        for _p in Product.query.all():
            if not _p.name:
                continue
            _new = _pat.sub(' ', _p.name)
            _new = _re_clean.sub(r'\s+', ' ', _new).strip()
            _new = _re_clean.sub(r'\s+([,.)\]])', r'\1', _new)
            _new = _re_clean.sub(r'\(\s+', '(', _new)
            if _new and _new != _p.name:
                _p.name = _new
                _cleaned += 1
        if _cleaned:
            db.session.commit()
            print(f'🧹 Очищено названий товаров от "в асс": {_cleaned}')
    except Exception as _e:
        db.session.rollback()
        print(f'⚠️  Чистка "в асс": {_e}')

    # Миграция: переименовываем admin-пользователя Мобара в TimurSaipov
    try:
        _mobar_email = os.environ.get('MOBAR_EMAIL')
        if _mobar_email:
            _mobar_org = Organization.query.filter_by(owner_email=_mobar_email).first()
            if _mobar_org:
                _admin_user = User.query.filter_by(
                    org_id=_mobar_org.id, username='admin', role='admin'
                ).first()
                if _admin_user:
                    _exists = User.query.filter_by(
                        org_id=_mobar_org.id, username='TimurSaipov'
                    ).first()
                    if not _exists:
                        _admin_user.username = 'TimurSaipov'
                        db.session.commit()
                        print(f'✅ Мобар admin → TimurSaipov')
    except Exception as _e:
        db.session.rollback()
        print(f'⚠️  Миграция Мобар username: {_e}')

    # Автосоздание owner из переменных окружения (для Render/VPS)
    owner_email = os.environ.get('OWNER_EMAIL')
    owner_password = os.environ.get('OWNER_PASSWORD')
    if owner_email and owner_password:
        existing = OwnerUser.query.filter_by(email=owner_email).first()
        if not existing:
            owner = OwnerUser()
            owner.email = owner_email
            owner.set_password(owner_password)
            db.session.add(owner)
            db.session.commit()
            print(f'✅ Owner создан: {owner_email}')

    # Автосоздание демо-компании из переменных окружения
    demo_email = os.environ.get('DEMO_EMAIL')
    demo_password = os.environ.get('DEMO_PASSWORD')
    demo_name = os.environ.get('DEMO_ORG_NAME', 'Демо Бургерная')
    if demo_email and demo_password:
        existing_org = Organization.query.filter_by(owner_email=demo_email).first()
        if not existing_org:
            from datetime import timedelta
            org = Organization(
                name=demo_name,
                owner_email=demo_email,
                email_verified=True,
                plan='trial',
                trial_ends_at=datetime.utcnow() + timedelta(days=30)
            )
            db.session.add(org)
            db.session.flush()
            user = User(org_id=org.id, username='admin', role='admin', email=demo_email)
            user.set_password(demo_password)
            db.session.add(user)
            for loc_name in ['Склад', 'Кухня', 'Островок']:
                db.session.add(Location(org_id=org.id, name=loc_name))
            db.session.commit()
            print(f'✅ Демо-компания создана: {demo_email}')

    # Автосоздание ООО "Мобар" с товарами из template.xlsx
    mobar_email = os.environ.get('MOBAR_EMAIL')
    mobar_password = os.environ.get('MOBAR_PASSWORD')
    if mobar_email and mobar_password:
        existing_mobar = Organization.query.filter_by(owner_email=mobar_email).first()
        if not existing_mobar:
            template_path = os.path.join(os.path.dirname(__file__), 'template.xlsx')
            template_bytes = None
            if os.path.exists(template_path):
                with open(template_path, 'rb') as f:
                    template_bytes = f.read()

            org = Organization(
                name='ООО "Мобар"',
                owner_email=mobar_email,
                email_verified=True,
                plan='pro',
                subscription_ends_at=datetime.utcnow() + timedelta(days=365),
                excel_template=template_bytes
            )
            db.session.add(org)
            db.session.flush()

            # Создаём admin-пользователя
            user = User(org_id=org.id, username='TimurSaipov', role='admin', email=mobar_email)
            user.set_password(mobar_password)
            db.session.add(user)

            # Локации
            for loc_name in ['Склад', 'Кухня', 'Островок']:
                db.session.add(Location(org_id=org.id, name=loc_name))
            db.session.flush()

            count = 0

            # Импортируем товары из template.xlsx
            if os.path.exists(template_path):
                import openpyxl as _oxl
                wb = _oxl.load_workbook(template_path, data_only=True)
                ws = wb.active
                current_cat_name = 'Разное'
                cat_cache = {}
                count = 0
                skip_headers = {'Код', 'Товар', ''}

                for row in ws.iter_rows(values_only=True):
                    code_raw = row[1] if len(row) > 1 else None
                    name_raw = row[2] if len(row) > 2 else None
                    unit_raw = row[5] if len(row) > 5 else None

                    # Строка-категория
                    if code_raw and isinstance(code_raw, str) and not name_raw:
                        val = code_raw.strip()
                        if val and not val.startswith('Дата печати') and val not in skip_headers:
                            current_cat_name = val
                        continue

                    # Строка-товар
                    if code_raw and name_raw and str(code_raw).strip() not in skip_headers:
                        code = str(code_raw).strip()
                        name = str(name_raw).strip()
                        unit = str(unit_raw).strip() if unit_raw else 'шт'

                        # Получаем или создаём категорию
                        if current_cat_name not in cat_cache:
                            cat = Category(org_id=org.id, name=current_cat_name)
                            db.session.add(cat)
                            db.session.flush()
                            cat_cache[current_cat_name] = cat.id

                        product = Product(
                            org_id=org.id,
                            category_id=cat_cache[current_cat_name],
                            name=name,
                            code=code,
                            unit=unit
                        )
                        db.session.add(product)
                        count += 1

            db.session.commit()
            print(f'✅ ООО Мобар создан: {mobar_email}, товаров: {count}')

if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host="0.0.0.0", port=port, debug=debug)
