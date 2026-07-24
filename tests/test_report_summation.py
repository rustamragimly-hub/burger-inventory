"""Тесты суммирования в отчёте по ревизии — самое опасное место
(баг здесь = потерянные подсчёты у клиента).

Проверяем стандартный отчёт (без кастомного шаблона):
колонки A=Код, B=Наименование, C=Ед, D=Категория, E=Остаток фактический.
"""
import io
from datetime import datetime
from openpyxl import load_workbook
import app
import models

db = app.db


def _completed_rev(org, loc, when):
    rev = models.Revision(org_id=org.id, location_id=loc.id, user_id=None,
                          status='completed', finished_at=when)
    db.session.add(rev)
    db.session.flush()
    return rev


def _item(rev, loc, prod, qty):
    db.session.add(models.RevisionItem(revision_id=rev.id, product_id=prod.id,
                                       location_id=loc.id, quantity=qty))


def _read_report(buf):
    """Вернуть {наименование: остаток} из стандартного отчёта."""
    wb = load_workbook(io.BytesIO(buf.getvalue()) if hasattr(buf, 'getvalue') else buf)
    ws = wb.active
    out = {}
    for row in range(9, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        qty = ws.cell(row=row, column=5).value
        # Пропускаем строки-заголовки категорий (только A заполнена)
        if name and ws.cell(row=row, column=3).value:
            out[name] = qty
    return out


def test_sums_multiple_items_same_product_one_location(session, org_factory):
    org, locs, prods = org_factory(products=[('1001', 'Кола', 'шт')], locations=('Склад',))
    sklad = locs[0]; kola = prods[0]
    when = datetime(2026, 6, 1, 10, 0)
    rev = _completed_rev(org, sklad, when)
    # Один товар посчитан тремя записями: 3 + 2 + 5 = 10
    _item(rev, sklad, kola, 3)
    _item(rev, sklad, kola, 2)
    _item(rev, sklad, kola, 5)
    db.session.commit()

    buf = app._build_revision_xlsx([rev])
    report = _read_report(buf)
    assert report.get('Кола') == 10


def test_sums_across_locations(session, org_factory):
    org, locs, prods = org_factory(products=[('1001', 'Кола', 'шт')], locations=('Склад', 'Кухня'))
    sklad, kuhnya = locs; kola = prods[0]
    when = datetime(2026, 6, 1, 10, 0)
    r1 = _completed_rev(org, sklad, when)
    r2 = _completed_rev(org, kuhnya, when)
    _item(r1, sklad, kola, 5)   # склад
    _item(r2, kuhnya, kola, 4)  # кухня
    db.session.commit()

    # Общий отчёт ресторана: 5 + 4 = 9
    buf = app._build_revision_xlsx([r1, r2])
    report = _read_report(buf)
    assert report.get('Кола') == 9


def test_counted_product_not_in_assortment_still_in_report(session, org_factory):
    # Товар посчитан на складе, но НЕ добавлен в LocationProduct ассортимент —
    # он всё равно должен попасть в отчёт (иначе подсчёт теряется).
    org, locs, prods = org_factory(products=[('1001', 'Кола', 'шт')], locations=('Склад',))
    sklad = locs[0]; kola = prods[0]
    # ассортимент НЕ настраиваем (LocationProduct пуст для этого товара)
    when = datetime(2026, 6, 1, 10, 0)
    rev = _completed_rev(org, sklad, when)
    _item(rev, sklad, kola, 7)
    db.session.commit()

    buf = app._build_revision_xlsx([rev])
    report = _read_report(buf)
    assert report.get('Кола') == 7


def test_uncounted_assortment_product_shows_zero(session, org_factory):
    # Товар из ассортимента локации, который не посчитали, должен присутствовать
    # в отчёте с остатком 0 — не пустой ячейкой и не выпадать из отчёта.
    org, locs, prods = org_factory(
        products=[('1001', 'Кола', 'шт'), ('1002', 'Фанта', 'шт')], locations=('Склад',))
    sklad = locs[0]; kola, fanta = prods
    db.session.add(models.LocationProduct(location_id=sklad.id, product_id=kola.id))
    db.session.add(models.LocationProduct(location_id=sklad.id, product_id=fanta.id))
    when = datetime(2026, 6, 1, 10, 0)
    rev = _completed_rev(org, sklad, when)
    _item(rev, sklad, kola, 5)  # Фанту не считали
    db.session.commit()

    report = _read_report(app._build_revision_xlsx([rev]))
    assert report.get('Кола') == 5
    assert report.get('Фанта') == 0


def test_assortment_union_across_locations(session, org_factory):
    # Товар есть хотя бы в одной локации ревизии → он в отчёте.
    # Кола только на складе, Фанта только на кухне — в общем отчёте оба (с 0).
    org, locs, prods = org_factory(
        products=[('1001', 'Кола', 'шт'), ('1002', 'Фанта', 'шт')],
        locations=('Склад', 'Кухня'))
    sklad, kuhnya = locs; kola, fanta = prods
    db.session.add(models.LocationProduct(location_id=sklad.id, product_id=kola.id))
    db.session.add(models.LocationProduct(location_id=kuhnya.id, product_id=fanta.id))
    when = datetime(2026, 6, 1, 10, 0)
    r1 = _completed_rev(org, sklad, when)
    r2 = _completed_rev(org, kuhnya, when)
    _item(r1, sklad, kola, 3)  # Фанту нигде не считали
    db.session.commit()

    report = _read_report(app._build_revision_xlsx([r1, r2]))
    assert report.get('Кола') == 3
    assert report.get('Фанта') == 0

    # А в отчёте только по складу кухонная Фанта появляться не должна.
    report_sklad = _read_report(app._build_revision_xlsx([r1]))
    assert report_sklad.get('Кола') == 3
    assert 'Фанта' not in report_sklad


def test_category_with_nothing_counted_still_in_report(session, org_factory):
    # Категория, в которой ничего не посчитали, всё равно выводится —
    # её товары из ассортимента показываются с нулями.
    org, locs, prods = org_factory(products=[('1001', 'Кола', 'шт')], locations=('Склад',))
    sklad = locs[0]; kola = prods[0]
    cat = models.Category(org_id=org.id, name='Снеки')
    db.session.add(cat)
    db.session.flush()
    chips = models.Product(org_id=org.id, category_id=cat.id, name='Чипсы', code='2001', unit='шт')
    db.session.add(chips)
    db.session.flush()
    db.session.add(models.LocationProduct(location_id=sklad.id, product_id=kola.id))
    db.session.add(models.LocationProduct(location_id=sklad.id, product_id=chips.id))
    when = datetime(2026, 6, 1, 10, 0)
    rev = _completed_rev(org, sklad, when)
    _item(rev, sklad, kola, 4)  # в «Снеках» ничего не считали
    db.session.commit()

    report = _read_report(app._build_revision_xlsx([rev]))
    assert report.get('Чипсы') == 0
