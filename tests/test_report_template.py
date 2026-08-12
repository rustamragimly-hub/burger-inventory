"""Тесты пути «заполнение загруженного Excel-шаблона» (org.excel_template).

Это формат Mobar/Прайд. Проверяем укрепления против реальных багов:
  • «осиротевшая» строка шаблона (код есть, имя пустое) не остаётся призраком
    и не «съезжает» список — либо восстанавливается из каталога, либо удаляется;
  • дробные суммы округляются (нет 0.7250000000000001);
  • обычные позиции ассортимента заполняются корректно.
"""
import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
import app
import models

db = app.db


def _make_iiko_template():
    """Мини-«Бланк инвентаризации» iiko: B=Код, C=Наименование, F=Ед, G=Остаток.
    Содержит осиротевшую строку (код без имени) — как в реальном файле Mobar."""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Бланк инвентаризации'
    ws.cell(3, 2, 'Товар'); ws.cell(3, 6, 'Ед. изм.'); ws.cell(3, 7, 'Остаток фактический')
    ws.cell(4, 2, 'Код'); ws.cell(4, 3, 'Наименование')
    ws.cell(5, 2, 'Десерты')                      # категория
    ws.cell(6, 2, '1001'); ws.cell(6, 3, 'Кола'); ws.cell(6, 6, 'шт')   # в каталоге
    ws.cell(7, 2, '9999')                          # ОСИРОТЕВШАЯ (нет в каталоге)
    ws.cell(8, 2, '1002'); ws.cell(8, 3, 'Сок'); ws.cell(8, 6, 'л')     # в каталоге
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _read_filled(buf):
    """{наименование -> остаток(G)} из заполненного шаблона + список кодов."""
    wb = load_workbook(io.BytesIO(buf.getvalue()) if hasattr(buf, 'getvalue') else buf)
    ws = wb.active
    by_name, codes = {}, []
    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        if code is not None:
            codes.append(str(code).strip())
        if name and ws.cell(r, 6).value:  # товарная строка (есть имя + единица)
            by_name[str(name).strip()] = ws.cell(r, 7).value
    return by_name, codes


def _setup(org_factory):
    org, locs, prods = org_factory(
        products=[('1001', 'Кола', 'шт'), ('1002', 'Сок', 'л')],
        locations=('Склад',),
    )
    loc = locs[0]
    for p in prods:
        db.session.add(models.LocationProduct(location_id=loc.id, product_id=p.id))
    org.excel_template = _make_iiko_template()
    db.session.commit()
    return org, loc, prods


def test_orphan_row_removed_and_fractions_rounded(session, org_factory):
    org, loc, prods = _setup(org_factory)
    kola, sok = prods
    when = datetime(2026, 8, 11, 10, 0)
    rev = models.Revision(org_id=org.id, location_id=loc.id, status='completed', finished_at=when)
    db.session.add(rev); db.session.flush()
    db.session.add(models.RevisionItem(revision_id=rev.id, product_id=kola.id, location_id=loc.id, quantity=5))
    # 0.5 + 0.225 = 0.7250000000000001 без округления
    db.session.add(models.RevisionItem(revision_id=rev.id, product_id=sok.id, location_id=loc.id, quantity=0.5))
    db.session.add(models.RevisionItem(revision_id=rev.id, product_id=sok.id, location_id=loc.id, quantity=0.225))
    db.session.commit()

    buf = app._build_revision_xlsx([rev])
    by_name, codes = _read_filled(buf)

    # Осиротевшая строка (код 9999, нет в каталоге) должна быть удалена
    assert '9999' not in codes
    # Обычные позиции на месте
    assert by_name.get('Кола') == 5
    # Дробь округлена, без float-мусора
    assert by_name.get('Сок') == 0.725


def test_orphan_row_name_restored_when_in_catalog(session, org_factory):
    """Если у осиротевшей строки код ЕСТЬ в каталоге — имя/единица
    восстанавливаются, строка остаётся полноценной, а не удаляется."""
    org, locs, prods = org_factory(
        products=[('1001', 'Кола', 'шт'), ('9999', 'Вода', 'бут')],
        locations=('Склад',),
    )
    loc = locs[0]
    for p in prods:
        db.session.add(models.LocationProduct(location_id=loc.id, product_id=p.id))
    org.excel_template = _make_iiko_template()  # строка 9999 без имени
    db.session.commit()

    rev = models.Revision(org_id=org.id, location_id=loc.id, status='completed',
                          finished_at=datetime(2026, 8, 11, 10, 0))
    db.session.add(rev); db.session.flush()
    db.session.add(models.RevisionItem(revision_id=rev.id, product_id=prods[1].id, location_id=loc.id, quantity=7))
    db.session.commit()

    buf = app._build_revision_xlsx([rev])
    by_name, codes = _read_filled(buf)
    assert '9999' in codes            # не удалена — код есть в каталоге
    assert by_name.get('Вода') == 7   # имя восстановлено из каталога + количество
