"""Тесты парсеров загрузок: iiko-бланк и выгрузка остатков.
Excel-файлы генерируются в памяти, проверяется корректность распознавания.
"""
import io
from openpyxl import Workbook
import app


def _iiko_bank_xlsx():
    wb = Workbook(); ws = wb.active; ws.title = 'Page1'
    ws['A1'] = 'Бланк инвентаризации'
    ws['B13'] = 'Товар'; ws['F13'] = 'Ед. изм.'; ws['G13'] = 'Остаток фактический'
    ws['B14'] = 'Код'; ws['C14'] = 'Наименование'
    ws['B15'] = 'Напитки б/а'
    ws['B16'] = '1001'; ws['C16'] = 'Кола 0,5 л'; ws['F16'] = 'шт'
    ws['B17'] = '1002'; ws['C17'] = 'Сок 1 л'; ws['F17'] = 'шт'
    ws['B18'] = 'Кофе'
    ws['B19'] = '2001'; ws['C19'] = 'Зерно эспрессо'; ws['F19'] = 'кг'
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_iiko_bank_products_and_categories():
    rows, xlsx = app._parse_iiko_bank(_iiko_bank_xlsx())
    assert len(rows) == 3
    cats = {r['Категория'] for r in rows}
    assert cats == {'Напитки б/а', 'Кофе'}
    by_code = {r['Код']: r for r in rows}
    assert by_code['1001']['Название'] == 'Кола 0,5 л'
    assert by_code['1001']['Ед. изм.'] == 'шт'
    assert by_code['2001']['Категория'] == 'Кофе'
    # Нормализованный шаблон должен читаться openpyxl
    from openpyxl import load_workbook
    load_workbook(io.BytesIO(xlsx))


def test_iiko_bank_sanitizes_via_import_path():
    # «в асс» в названии — парсер сам не чистит, но importrows должен (проверяем,
    # что парсер сохраняет исходное имя, чистка — отдельный слой)
    wb = Workbook(); ws = wb.active
    ws['B13'] = 'Товар'; ws['F13'] = 'Ед. изм.'
    ws['B14'] = 'Код'; ws['C14'] = 'Наименование'
    ws['B15'] = 'Десерты'
    ws['B16'] = '5606'; ws['C16'] = 'Драже в асс М&М'; ws['F16'] = 'шт'
    buf = io.BytesIO(); wb.save(buf)
    rows, _ = app._parse_iiko_bank(buf.getvalue())
    assert rows[0]['Название'] == 'Драже в асс М&М'  # парсер не трогает
    assert app._sanitize_product_name(rows[0]['Название']) == 'Драже М&М'  # чистка убирает


def test_stock_export_universal_excel_with_price():
    wb = Workbook(); ws = wb.active
    ws.append(['Код', 'Наименование', 'Остаток', 'Цена'])
    ws.append(['1001', 'Кола 0,5 л', 40, 55])
    ws.append(['1002', 'Сок 1 л', 12.5, 120])
    buf = io.BytesIO(); wb.save(buf)
    rows = app._parse_stock_export(buf.getvalue())
    assert len(rows) == 2
    d = {r['code']: r for r in rows}
    assert d['1001']['qty'] == 40
    assert d['1001']['price'] == 55
    assert d['1002']['qty'] == 12.5


def test_stock_export_handles_comma_decimal():
    wb = Workbook(); ws = wb.active
    ws.append(['Код', 'Остаток'])
    ws.append(['3001', '3,5'])  # запятая как десятичный разделитель
    buf = io.BytesIO(); wb.save(buf)
    rows = app._parse_stock_export(buf.getvalue())
    assert rows[0]['qty'] == 3.5


def test_stock_export_from_filled_iiko_bank():
    wb = Workbook(); ws = wb.active
    ws['B13'] = 'Товар'; ws['F13'] = 'Ед. изм.'; ws['G13'] = 'Остаток'
    ws['B14'] = 'Код'; ws['C14'] = 'Наименование'
    ws['B15'] = 'Напитки'
    ws['B16'] = '1001'; ws['C16'] = 'Кола'; ws['F16'] = 'шт'; ws['G16'] = 40
    buf = io.BytesIO(); wb.save(buf)
    rows = app._parse_stock_export(buf.getvalue())
    assert len(rows) == 1
    assert rows[0]['code'] == '1001'
    assert rows[0]['qty'] == 40
